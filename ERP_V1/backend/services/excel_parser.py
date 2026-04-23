"""
Excel parsing logic for client and factory Excel files.

Extracts row data from uploaded Excel files, matches against existing products,
and detects duplicates/variants.
"""
import json
import logging
import os
from collections import defaultdict
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import openpyxl
from sqlalchemy.orm import Session

from config import UPLOAD_DIR
from enums import JobStatus, ApprovalStatus
from models import (
    ProcessingJob, Order, Product, OrderItem, ClientProductBarcode,
    ProductImage,
)
from services.image_extractor import extract_images_for_rows

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------

def _detect_column_map(header, ai_col_map: Optional[dict] = None) -> dict:
    """Detect column positions from header row, with optional AI override."""
    if ai_col_map and ai_col_map.get("part_no_1") is not None:
        return ai_col_map

    col_map = {
        "part_no_1": None, "part_no_2": None, "description": None,
        "chinese_name": None, "qty": None, "price": None,
        "part_type": None, "dimension": None, "material": None,
        "variant_note": None, "weight": None, "category": None,
    }
    part_no_indices = []
    price_indices = []

    for ci, cell in enumerate(header or []):
        val = str(cell or "").strip().lower()
        if "description" in val or "goods" in val or "part name" in val or (val == "name"):
            col_map["description"] = ci
        elif "part no" in val or "part code" in val or "mfr" in val:
            part_no_indices.append(ci)
        elif any(kw in val for kw in ["unit price", "price", "\u5355\u4ef7"]) and "total" not in val:
            price_indices.append(ci)
        elif val in ("quantity", "qty"):
            col_map["qty"] = ci
        elif "\u540d\u79f0" in val or "chinese" in val:
            col_map["chinese_name"] = ci
        elif any(kw in val for kw in ["part type", "type", "original", "copy"]) and "part no" not in val:
            col_map["part_type"] = ci
        elif any(kw in val for kw in ["dimension", "size", "spec", "\u89c4\u683c"]):
            col_map["dimension"] = ci
        elif any(kw in val for kw in ["material", "\u6750\u8d28", "\u6750\u6599"]):
            col_map["material"] = ci
        elif any(kw in val for kw in ["variant", "note", "remark", "\u5907\u6ce8"]) and "variant" in val:
            col_map["variant_note"] = ci
        elif any(kw in val for kw in ["weight", "weigh", "\u91cd\u91cf"]):
            col_map["weight"] = ci
        elif any(kw in val for kw in ["category", "\u7c7b\u522b", "\u5206\u7c7b"]):
            col_map["category"] = ci

    if len(part_no_indices) >= 2:
        col_map["part_no_1"] = part_no_indices[0]
        col_map["part_no_2"] = part_no_indices[1]
    elif len(part_no_indices) == 1:
        col_map["part_no_1"] = part_no_indices[0]
        col_map["part_no_2"] = part_no_indices[0]

    if price_indices:
        col_map["price"] = price_indices[-1]

    if col_map["part_no_1"] is None:
        col_map = {
            "description": 1, "part_no_1": 2, "part_no_2": 3,
            "chinese_name": 5, "qty": 7, "price": 8,
        }

    return col_map


def _safe_col(row, col_map, key):
    """Safely extract a string value from a row using the column map."""
    col = col_map.get(key)
    if col is None or len(row) <= col:
        return ""
    return str(row[col] or "").strip()


def _parse_weight(raw: str) -> Optional[float]:
    """Parse a weight string like '1.5kg' into a float."""
    if not raw:
        return None
    try:
        cleaned = raw.lower().replace("kg", "").replace("lbs", "").replace("lb", "").strip()
        return round(float(cleaned), 3) if cleaned else None
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Client Excel parsing
# ---------------------------------------------------------------------------

def process_client_excel(job: ProcessingJob, db: Session):
    """Parse client request Excel: Col A=barcode, Col B=manufacturer_code, Col C=quantity."""
    wb = openpyxl.load_workbook(job.file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    total = len(rows)
    job.total_rows = total
    db.commit()

    results = []
    summary = {"matched": 0, "new_products": 0, "duplicates": 0, "ambiguous": 0, "errors": 0, "bin_matches": 0}
    bin_product_ids = []
    seen_codes = {}

    for i, row in enumerate(rows):
        if len(row) < 3:
            continue

        barcode = str(row[0] or "").strip()
        mfr_code = str(row[1] or "").strip()
        qty_raw = row[2]
        try:
            qty = int(float(qty_raw)) if qty_raw else 0
        except (ValueError, TypeError):
            qty = 0

        if not mfr_code and not barcode:
            continue

        match_status = "NEW_PRODUCT"
        product_id = None
        product_name = None

        products_by_code = db.query(Product).filter(
            Product.product_code == mfr_code,
            Product.is_active == True,
            Product.deleted_at.is_(None),
        ).all()

        if len(products_by_code) == 1:
            match_status = "MATCHED"
            product_id = products_by_code[0].id
            product_name = products_by_code[0].product_name
            summary["matched"] += 1
        elif len(products_by_code) > 1:
            match_status = "AMBIGUOUS"
            summary["ambiguous"] += 1
        else:
            if barcode and job.order_id:
                order = db.query(Order).filter(Order.id == job.order_id).first()
                if order and order.client_id:
                    barcode_map = db.query(ClientProductBarcode).filter(
                        ClientProductBarcode.client_id == order.client_id,
                        ClientProductBarcode.barcode_code == barcode,
                    ).first()
                    if barcode_map:
                        match_status = "MATCHED"
                        product_id = barcode_map.product_id
                        p = db.query(Product).filter(Product.id == product_id).first()
                        product_name = p.product_name if p else None
                        summary["matched"] += 1

            if match_status == "NEW_PRODUCT":
                bin_product = db.query(Product).filter(
                    Product.product_code == mfr_code,
                    Product.deleted_at.isnot(None),
                ).first() if mfr_code else None
                if bin_product:
                    summary["bin_matches"] += 1
                    bin_product_ids.append(bin_product.id)
                summary["new_products"] += 1

        code_key = mfr_code or barcode
        if code_key in seen_codes:
            match_status = "DUPLICATE"
            summary["duplicates"] += 1
            first_idx = seen_codes[code_key]
            if results[first_idx]["match_status"] != "DUPLICATE":
                results[first_idx]["match_status"] = "DUPLICATE"
        seen_codes[code_key] = i

        results.append({
            "row": i + 2,
            "barcode": barcode,
            "manufacturer_code": mfr_code,
            "quantity": qty,
            "match_status": match_status,
            "product_id": product_id,
            "product_name": product_name,
        })

        if i % 5 == 0 or i == total - 1:
            job.processed_rows = i + 1
            job.progress = min(int((i + 1) / max(total, 1) * 100), 100)
            db.commit()

    summary["bin_product_ids"] = bin_product_ids
    job.status = JobStatus.COMPLETED.value
    job.progress = 100
    job.processed_rows = total
    job.result_summary = json.dumps(summary)
    job.result_data = json.dumps(results)
    job.completed_at = datetime.utcnow()
    db.commit()


# ---------------------------------------------------------------------------
# Factory Excel parsing
# ---------------------------------------------------------------------------

def process_factory_excel(job: ProcessingJob, db: Session):
    """Parse factory response Excel with auto-detected column layout.
    Two passes: 1) read_only for row data, 2) full load for images."""
    # PASS 1: Read row data
    wb = openpyxl.load_workbook(job.file_path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not all_rows:
        job.status = JobStatus.FAILED.value
        job.error_message = "Empty Excel file"
        db.commit()
        return

    header = all_rows[0]

    # Check for AI-provided column mapping
    ai_col_map = _load_ai_column_mapping(job.file_path, header)
    col_map = _detect_column_map(header, ai_col_map)

    rows = all_rows[1:]
    total = len(rows)
    job.total_rows = total
    db.commit()

    results, summary, merged_row_map = _parse_factory_rows(
        rows, col_map, job, db, total,
    )

    # PASS 2: Extract images
    job.progress = 50
    db.commit()

    excel_row_to_result = {}
    for ridx, r in enumerate(results):
        excel_row_to_result[r["row"]] = ridx
    excel_row_to_result.update(merged_row_map)

    product_image_hashes = extract_images_for_rows(
        file_path=job.file_path,
        excel_row_to_result=excel_row_to_result,
        results=results,
        summary=summary,
        job_order_id=job.order_id,
        db=db,
        ProductImage=ProductImage,
        on_progress=lambda p: _update_progress(job, db, p),
    )

    # Flush so newly extracted images are queryable for conflict detection
    db.flush()

    # PASS 3: Detect image conflicts with existing DB images
    _detect_image_conflicts(results, summary, product_image_hashes, db, job_started_at=job.started_at or job.created_at)

    db.commit()

    job.status = JobStatus.COMPLETED.value
    job.progress = 100
    job.processed_rows = total
    job.result_summary = json.dumps(summary)
    job.result_data = json.dumps(results)
    job.completed_at = datetime.utcnow()
    db.commit()


def _load_ai_column_mapping(file_path: str, header) -> Optional[dict]:
    """Load AI column mapping if a .column_mapping.json sidecar exists."""
    mapping_path = file_path + ".column_mapping.json"
    if not os.path.exists(mapping_path):
        return None
    try:
        with open(mapping_path, "r", encoding="utf-8") as mf:
            ai_mapping = json.load(mf)
        from services.column_mapper import build_col_map
        headers_str = [str(h) if h is not None else "" for h in header]
        ai_col_map = build_col_map(headers_str, ai_mapping, "product")
        logger.info("Using AI column mapping: %s", ai_col_map)
        return ai_col_map
    except Exception as exc:
        logger.warning("Failed to load AI column mapping, falling back: %s", exc)
        return None


def _update_progress(job: ProcessingJob, db: Session, progress: int):
    """Update job progress and commit."""
    job.progress = progress
    db.commit()


def _parse_factory_rows(
    rows: list,
    col_map: dict,
    job: ProcessingJob,
    db: Session,
    total: int,
) -> Tuple[list, dict, dict]:
    """Parse factory Excel data rows and match against products.

    Returns: (results, summary, merged_row_map)
    """
    results = []
    summary = {
        "matched": 0, "new_products": 0, "new_variants": 0, "no_price": 0,
        "images_extracted": 0, "errors": 0, "bin_matches": 0, "merged_rows": 0,
    }
    bin_product_ids = []
    seen_codes = {}
    duplicate_codes = {}
    seen_variant_keys = {}
    merged_row_map = {}
    new_product_codes = {}

    for i, row in enumerate(rows):
        if not row or len(row) < 4:
            continue

        sl = row[0]
        description = _safe_col(row, col_map, "description")
        pn1 = col_map.get("part_no_1")
        pn2 = col_map.get("part_no_2")
        part_no_1 = str(row[pn1] or "").strip() if pn1 is not None and len(row) > pn1 else ""
        part_no_2 = str(row[pn2] or "").strip() if pn2 is not None and len(row) > pn2 else ""
        mfr_code = part_no_2 or part_no_1
        barcode = part_no_1 if part_no_1 != mfr_code else ""
        chinese_name = _safe_col(row, col_map, "chinese_name")
        part_type = _safe_col(row, col_map, "part_type")
        dimension = _safe_col(row, col_map, "dimension")
        material = _safe_col(row, col_map, "material")
        variant_note = _safe_col(row, col_map, "variant_note")
        weight_raw = _safe_col(row, col_map, "weight")
        weight_val = _parse_weight(weight_raw)
        category = _safe_col(row, col_map, "category")

        qty_col = col_map.get("qty")
        price_col = col_map.get("price")
        qty_raw = row[qty_col] if qty_col is not None and len(row) > qty_col else None
        price_raw = row[price_col] if price_col is not None and len(row) > price_col else None

        try:
            qty = int(float(qty_raw)) if qty_raw else 0
        except (ValueError, TypeError):
            qty = 0

        try:
            factory_price = float(price_raw) if price_raw else None
        except (ValueError, TypeError):
            factory_price = None

        # Variant-aware duplicate handling
        variant_key = (mfr_code, description, dimension) if mfr_code else None

        if variant_key and variant_key in seen_variant_keys:
            existing_idx = seen_variant_keys[variant_key]
            existing_cat = results[existing_idx].get("category", "")
            existing_cat_list = [c.strip() for c in existing_cat.split("; ")] if existing_cat else []
            if category and category not in existing_cat_list:
                results[existing_idx]["category"] = "; ".join(filter(None, [existing_cat, category]))
            if weight_val is not None and results[existing_idx].get("weight") is None:
                results[existing_idx]["weight"] = weight_val
            merged_row_map[i + 2] = existing_idx
            summary["merged_rows"] += 1
            if i % 5 == 0:
                job.processed_rows = i + 1
                job.progress = min(int((i + 1) / max(total, 1) * 50), 50)
                db.commit()
            continue

        if not mfr_code and not barcode:
            continue

        # Match product
        match_status, product_id, product_name, existing_variants_list = _match_product(
            mfr_code, barcode, description, chinese_name, part_type,
            dimension, material, variant_note, weight_val, category,
            summary, bin_product_ids, new_product_codes, len(results), db,
        )

        if factory_price is None:
            summary["no_price"] += 1

        results.append({
            "row": i + 2,
            "sl": sl,
            "description": description,
            "barcode": barcode,
            "manufacturer_code": mfr_code,
            "chinese_name": chinese_name,
            "part_type": part_type,
            "dimension": dimension,
            "material": material,
            "variant_note": variant_note,
            "weight": weight_val,
            "category": category,
            "quantity": qty,
            "factory_price_usd": factory_price,
            "match_status": match_status,
            "product_id": product_id,
            "product_name": product_name,
            "has_image": False,
            "is_duplicate_in_file": False,
            "existing_variants": existing_variants_list,
        })

        result_idx = len(results) - 1
        if mfr_code:
            if mfr_code in seen_codes:
                if mfr_code not in duplicate_codes:
                    duplicate_codes[mfr_code] = [seen_codes[mfr_code]]
                duplicate_codes[mfr_code].append(result_idx)
            else:
                seen_codes[mfr_code] = result_idx
        if variant_key:
            seen_variant_keys[variant_key] = len(results) - 1

        if i % 5 == 0:
            job.processed_rows = i + 1
            job.progress = min(int((i + 1) / max(total, 1) * 50), 50)
            db.commit()

    summary["bin_product_ids"] = bin_product_ids
    summary["duplicate_codes"] = [
        {"code": code, "indices": indices, "rows": [results[i]["row"] for i in indices]}
        for code, indices in duplicate_codes.items()
    ]
    summary["duplicate_code_count"] = len(duplicate_codes)

    return results, summary, merged_row_map


def _match_product(
    mfr_code, barcode, description, chinese_name, part_type,
    dimension, material, variant_note, weight_val, category,
    summary, bin_product_ids, new_product_codes, current_result_len, db,
):
    """Match a row against existing products. Returns (match_status, product_id, product_name, existing_variants)."""
    match_status = "NEW_PRODUCT"
    product_id = None
    product_name = None

    children_by_code = db.query(Product).filter(
        Product.product_code == mfr_code,
        Product.parent_id.isnot(None),
        Product.is_active == True,
        Product.deleted_at.is_(None),
    ).all()

    if children_by_code:
        exact_match = None
        for child in children_by_code:
            type_ok = (not part_type) or (child.part_type or "") == part_type
            dim_ok = (not dimension) or (child.dimension or "") == dimension
            mat_ok = (not material) or (child.material or "") == material
            name_ok = (not description) or child.product_name == description
            if type_ok and dim_ok and mat_ok and name_ok:
                exact_match = child
                break

        if exact_match:
            match_status = "MATCHED"
            product_id = exact_match.id
            product_name = exact_match.product_name
            summary["matched"] += 1
            if description and exact_match.product_name == exact_match.product_code:
                exact_match.product_name = description
                product_name = description
            if part_type and not exact_match.part_type:
                exact_match.part_type = part_type
            if dimension and not exact_match.dimension:
                exact_match.dimension = dimension
            if material and not exact_match.material:
                exact_match.material = material
            if variant_note and not exact_match.variant_note:
                exact_match.variant_note = variant_note
            if weight_val is not None and not exact_match.unit_weight_kg:
                exact_match.unit_weight_kg = weight_val
            if category and not exact_match.category:
                exact_match.category = category
            if chinese_name and not exact_match.product_name_chinese:
                exact_match.product_name_chinese = chinese_name
            db.commit()
        elif len(children_by_code) == 1 and not part_type and not dimension and not material:
            child = children_by_code[0]
            if not description or child.product_name == description or child.product_name == child.product_code:
                match_status = "MATCHED"
                product_id = child.id
                product_name = child.product_name
                summary["matched"] += 1
                if description and child.product_name == child.product_code:
                    child.product_name = description
                    product_name = description
                if chinese_name and not child.product_name_chinese:
                    child.product_name_chinese = chinese_name
                    db.commit()
            else:
                match_status = "NEW_VARIANT"
                summary["new_variants"] = summary.get("new_variants", 0) + 1
        else:
            match_status = "NEW_VARIANT"
            summary["new_variants"] = summary.get("new_variants", 0) + 1
    else:
        legacy = db.query(Product).filter(
            Product.product_code == mfr_code,
            Product.parent_id.is_(None),
            Product.is_active == True,
            Product.deleted_at.is_(None),
        ).first()
        if legacy and legacy.product_name.startswith("["):
            match_status = "NEW_VARIANT"
            summary["new_variants"] = summary.get("new_variants", 0) + 1
        elif legacy:
            match_status = "MATCHED"
            product_id = legacy.id
            product_name = legacy.product_name
            summary["matched"] += 1
        else:
            match_status = "NEW_PRODUCT"
            bin_product = db.query(Product).filter(
                Product.product_code == mfr_code,
                Product.deleted_at.isnot(None),
            ).first() if mfr_code else None
            if bin_product:
                summary["bin_matches"] += 1
                bin_product_ids.append(bin_product.id)
            summary["new_products"] += 1

    if match_status == "NEW_PRODUCT" and mfr_code in new_product_codes:
        match_status = "NEW_VARIANT"
        summary["new_products"] -= 1
        summary["new_variants"] = summary.get("new_variants", 0) + 1
    elif match_status == "NEW_PRODUCT":
        new_product_codes[mfr_code] = current_result_len

    # Build existing variants list for NEW_VARIANT rows
    if match_status == "NEW_VARIANT" and children_by_code:
        existing_variants_list = [{
            "id": c.id,
            "product_name": c.product_name,
            "material": c.material or "",
            "dimension": c.dimension or "",
            "part_type": c.part_type or "",
            "is_default": getattr(c, 'is_default', False),
        } for c in children_by_code]
    else:
        existing_variants_list = []

    return match_status, product_id, product_name, existing_variants_list


def _detect_image_conflicts(
    results: list,
    summary: dict,
    product_image_hashes: Dict[str, set],
    db: Session,
    job_started_at=None,
):
    """Detect image conflicts between new Excel images and existing DB images."""
    image_conflicts = []
    matched_pids_with_images = set()
    for r in results:
        pid = r.get("product_id")
        if pid and r.get("has_image"):
            matched_pids_with_images.add(pid)

    if matched_pids_with_images:
        q = db.query(ProductImage).filter(
            ProductImage.product_id.in_(list(matched_pids_with_images)),
            ProductImage.source_type == "FACTORY_EXCEL",
        )
        # Only compare against images that existed BEFORE this extraction
        if job_started_at:
            q = q.filter(ProductImage.created_at < job_started_at)
        existing_db_images = q.all()

        existing_hash_map = {}
        for db_img in existing_db_images:
            if db_img.image_hash:
                existing_hash_map[db_img.product_id] = {
                    "hash": db_img.image_hash,
                    "thumbnail_url": f"/api/products/file/?path={db_img.thumbnail_path or db_img.image_path}",
                }

        for pid, new_hashes in product_image_hashes.items():
            if pid in existing_hash_map:
                existing = existing_hash_map[pid]
                if existing["hash"] not in new_hashes:
                    code = ""
                    for r in results:
                        if r.get("product_id") == pid:
                            code = r.get("manufacturer_code", "")
                            break
                    # Find the newly extracted image's thumbnail (most recent first)
                    # First try: exact hash match from current extraction
                    new_thumbnail_url = None
                    new_img = None
                    if new_hashes:
                        new_img = db.query(ProductImage).filter(
                            ProductImage.product_id == pid,
                            ProductImage.image_hash.in_(list(new_hashes)),
                        ).order_by(ProductImage.created_at.desc()).first()
                    # Fallback: if no new image saved (dedup or extraction failed),
                    # show the most recent non-existing image as preview
                    if not new_img:
                        new_img = db.query(ProductImage).filter(
                            ProductImage.product_id == pid,
                            ProductImage.image_hash != existing["hash"],
                        ).order_by(ProductImage.created_at.desc()).first()
                    if new_img:
                        new_thumbnail_url = f"/api/products/file/?path={new_img.thumbnail_path or new_img.image_path}"

                    image_conflicts.append({
                        "product_id": pid,
                        "code": code,
                        "existing_hash": existing["hash"],
                        "existing_thumbnail_url": existing["thumbnail_url"],
                        "new_thumbnail_url": new_thumbnail_url,
                        "new_hashes": list(new_hashes),
                    })

    summary["image_conflicts"] = image_conflicts
    summary["image_conflict_count"] = len(image_conflicts)
