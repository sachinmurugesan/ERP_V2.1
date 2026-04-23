"""
Landed Cost Excel generation — produces a professional cost sheet.
"""
from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers


_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_HEADER_FONT = Font(bold=True, size=10)
_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11)
_CURRENCY_FMT = '#,##0.00'
_THIN_BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC'),
)


def generate_landed_cost_excel(data: dict) -> BytesIO:
    """Generate Excel file from landed cost API response dict."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Landed Cost"

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    row = 1

    # ── Title ──
    ws.cell(row=row, column=1, value="LANDED COST BREAKDOWN").font = _TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Order: {data.get('order_number', '—')}")
    ws.cell(row=row, column=2, value=f"Client: {data.get('client_name', '—')}")
    ws.cell(row=row, column=3, value=f"Date: {date.today().isoformat()}")
    row += 2

    # ── Invoice ──
    _section_header(ws, row, "INVOICE")
    row += 1
    invoice = data.get("invoice", {})
    _cost_row(ws, row, invoice.get("label", "Invoice"), invoice.get("amount_inr", 0))
    row += 2

    # ── Expenses ──
    _section_header(ws, row, "EXPENSES")
    row += 1
    for exp in data.get("expenses", []):
        _cost_row(ws, row, exp.get("label", ""), exp.get("amount_inr", 0))
        row += 1
    row += 1

    # ── Summary ──
    summary = data.get("summary", {})
    _section_header(ws, row, "SUMMARY")
    row += 1
    _cost_row(ws, row, "Total Bill", summary.get("total_bill_inr", 0), bold=True)
    row += 1
    _cost_row(ws, row, "Total Expenses", summary.get("total_expenses_inr", 0), bold=True)
    row += 1
    _cost_row(ws, row, "Grand Total", summary.get("grand_total_inr", 0), bold=True)
    row += 1
    c = ws.cell(row=row, column=1, value="Expense %")
    c.font = Font(bold=True)
    c2 = ws.cell(row=row, column=2, value=f"{summary.get('expense_percent', 0):.4f}%")
    c2.alignment = Alignment(horizontal='right')
    row += 2

    # ── Per-Item Breakdown ──
    items = data.get("items", [])
    if items:
        _section_header(ws, row, "PER-ITEM BREAKDOWN")
        row += 1

        # Headers
        headers = ["#", "Product", "Code", "Qty", "Client Factory (CNY)",
                   "Value (INR)", "Freight Share", "Duty Share",
                   "Clearance Share", "Commission Share",
                   "Landed Cost", "Per Unit"]
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = Alignment(horizontal='center' if col_idx <= 4 else 'right')
        row += 1

        # Set wider columns for item table
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 14
        ws.column_dimensions['K'].width = 16
        ws.column_dimensions['L'].width = 14

        for idx, item in enumerate(items, 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item.get("product_name", ""))
            ws.cell(row=row, column=3, value=item.get("product_code", ""))
            ws.cell(row=row, column=4, value=item.get("quantity", 0))
            _num_cell(ws, row, 5, item.get("client_factory_price_cny", 0))
            _num_cell(ws, row, 6, item.get("item_value_inr", 0))
            _num_cell(ws, row, 7, item.get("freight_share", 0))
            _num_cell(ws, row, 8, item.get("duty_share", 0))
            _num_cell(ws, row, 9, item.get("clearance_share", 0))
            _num_cell(ws, row, 10, item.get("commission_share", 0))
            _num_cell(ws, row, 11, item.get("total_landed_cost", 0))
            _num_cell(ws, row, 12, item.get("landed_cost_per_unit", 0))
            row += 1

        # Grand total row
        c = ws.cell(row=row, column=1, value="TOTAL")
        c.font = Font(bold=True)
        _num_cell(ws, row, 11, summary.get("grand_total_inr", 0), bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _section_header(ws, row: int, title: str):
    """Write a section header with grey background."""
    c = ws.cell(row=row, column=1, value=title)
    c.font = _SECTION_FONT
    c.fill = _HEADER_FILL
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = _HEADER_FILL


def _cost_row(ws, row: int, label: str, amount: float, bold: bool = False):
    """Write a label + amount row."""
    c1 = ws.cell(row=row, column=1, value=label)
    if bold:
        c1.font = Font(bold=True)
    c2 = ws.cell(row=row, column=2, value=amount)
    c2.number_format = _CURRENCY_FMT
    c2.alignment = Alignment(horizontal='right')
    if bold:
        c2.font = Font(bold=True)


def _num_cell(ws, row: int, col: int, val: float, bold: bool = False):
    """Write a right-aligned number cell."""
    c = ws.cell(row=row, column=col, value=val)
    c.number_format = _CURRENCY_FMT
    c.alignment = Alignment(horizontal='right')
    if bold:
        c.font = Font(bold=True)
