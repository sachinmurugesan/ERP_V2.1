"""Product Catalog API endpoints"""
import logging
import os
from math import ceil
import shutil
import hashlib
from typing import Optional, List

logger = logging.getLogger(__name__)
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse
from pathlib import Path
from sqlalchemy.orm import Session
import mimetypes
from sqlalchemy import or_, func

from config import UPLOAD_DIR
from database import get_db
from datetime import datetime
from core.security import CurrentUser, get_current_user
from models import (
    Product, ProductImage, OrderItem, ClientProductBarcode,
    UnloadedItem, WarehouseStock, PackingListItem,
    ProductVerification, AfterSalesItem, Order, ProductRequest,
)
from schemas.products import (
    ProductCreate, ProductOut, ProductImageOut,
    ProductListResponse, ProductGroupedListResponse, ProductGroupItem,
    BulkDeleteRequest, BulkUpdateRequest,
    CodeValidationRequest, CodeValidationResult,
    MapProductRequestBody, RejectProductRequestBody,
    ApproveProductRequestBody,
)

router = APIRouter()
public_router = APIRouter()  # No auth — used for public file serving


def _prod_url(path: str) -> str:
    """Return the authenticated API URL for a product image/file.
    Used in all API responses so frontend components never need bare /uploads/ paths.
    """
    return f"/api/products/file/?path={path}"


# ========================================
# Product File Proxy (authenticated)
# ========================================

@public_router.get("/file/")
def download_product_file(
    path: str = Query(..., description="Relative path within uploads, e.g. products/abc/image.jpg"),
):
    """Public download of a product catalog image.
    Path is whitelisted to the products/ subdirectory only; UUID dirs prevent enumeration.
    """
    # Strict whitelist — only the products/ subdirectory
    if not path.startswith("products/") or ".." in path:
        raise HTTPException(status_code=400, detail="Invalid file path")

    full_path = UPLOAD_DIR / path
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    filename = Path(path).name
    mime_type = mimetypes.guess_type(str(full_path))[0] or "application/octet-stream"
    return FileResponse(path=str(full_path), filename=filename, media_type=mime_type)


# ========================================
# Endpoints
# ========================================

# ========================================
# Endpoints
# ========================================
@router.get("/")
def list_products(
    search: Optional[str] = None,
    category: Optional[str] = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    show_parents: bool = Query(False),
    group: bool = Query(False),
    sort_by: Optional[str] = Query(None, description="Sort field: product_code, product_name, category, hs_code, variants, created_at"),
    sort_dir: Optional[str] = Query("asc", description="Sort direction: asc or desc"),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """List products with search and pagination.
    By default only shows variant children (orderable products).
    ?show_parents=true  → include parent products
    ?group=true         → return parents with nested variants
    """
    query = db.query(Product).filter(
        Product.is_active == True,
        Product.deleted_at.is_(None)
    )

    # By default: only show children (orderable variants)
    if not show_parents and not group:
        query = query.filter(Product.parent_id.isnot(None))

    if search:
        query = query.filter(
            or_(
                Product.product_code.ilike(f"%{search}%"),
                Product.product_name.ilike(f"%{search}%"),
                Product.material.ilike(f"%{search}%"),
                Product.part_type.ilike(f"%{search}%"),
                Product.dimension.ilike(f"%{search}%"),
            )
        )

    if category:
        query = query.filter(Product.category == category)

    # CLIENT brand-based access filtering
    if current_user.role == "CLIENT" and current_user.client_id:
        from models import ClientBrandAccess
        records = db.query(ClientBrandAccess).filter(
            ClientBrandAccess.client_id == current_user.client_id
        ).all()
        if records:
            allowed_brands = [r.brand for r in records if r.brand != "__NO_BRAND__"]
            include_no_brand = any(r.include_no_brand for r in records)
            conditions = []
            if allowed_brands:
                conditions.append(Product.brand.in_(allowed_brands))
            if include_no_brand:
                conditions.append(Product.brand.is_(None))
                conditions.append(Product.brand == "")
            # Also include per-product access exceptions
            from models import ClientProductAccess
            product_exceptions = db.query(ClientProductAccess.product_id).filter(
                ClientProductAccess.client_id == current_user.client_id
            ).subquery()
            conditions.append(Product.id.in_(product_exceptions))

            if conditions:
                query = query.filter(or_(*conditions))
            else:
                return {"products": [], "total": 0, "page": page, "per_page": per_page, "pages": 0}
        else:
            # No brands assigned — check for per-product exceptions
            from models import ClientProductAccess
            product_exceptions = db.query(ClientProductAccess.product_id).filter(
                ClientProductAccess.client_id == current_user.client_id
            ).subquery()
            has_exceptions = db.query(func.count(ClientProductAccess.id)).filter(
                ClientProductAccess.client_id == current_user.client_id
            ).scalar()
            if has_exceptions:
                query = query.filter(Product.id.in_(product_exceptions))
            else:
                return {"products": [], "total": 0, "page": page, "per_page": per_page, "pages": 0}

    # --- Grouped mode: return parents with nested children ---
    if group:
        # Get parent products
        parent_q = db.query(Product).filter(
            Product.is_active == True,
            Product.deleted_at.is_(None),
            Product.parent_id.is_(None),  # parents only
        )
        if search:
            # Search should match parent code/name OR any child's data
            matching_child_parent_ids = db.query(Product.parent_id).filter(
                Product.parent_id.isnot(None),
                Product.is_active == True,
                Product.deleted_at.is_(None),
                or_(
                    Product.product_code.ilike(f"%{search}%"),
                    Product.product_name.ilike(f"%{search}%"),
                    Product.material.ilike(f"%{search}%"),
                    Product.part_type.ilike(f"%{search}%"),
                    Product.dimension.ilike(f"%{search}%"),
                ),
            ).distinct().subquery()
            parent_q = parent_q.filter(
                or_(
                    Product.product_code.ilike(f"%{search}%"),
                    Product.product_name.ilike(f"%{search}%"),
                    Product.id.in_(db.query(matching_child_parent_ids)),
                )
            )
        if category:
            parent_q = parent_q.filter(Product.category == category)

        total = parent_q.count()

        # Sorting for grouped mode
        # Note: parent product_name is auto-generated "[CODE]", so for product_name/category/hs_code
        # we sort by the first child's value (which is what the UI actually displays).
        from sqlalchemy.orm import aliased
        _C = aliased(Product, name="child_sort")
        _child_filter = [_C.parent_id.isnot(None), _C.is_active == True, _C.deleted_at.is_(None)]

        if sort_by == "variants":
            child_agg = (
                db.query(_C.parent_id.label("pid"), func.count(_C.id).label("val"))
                .filter(*_child_filter)
                .group_by(_C.parent_id)
                .subquery()
            )
            parent_q = parent_q.outerjoin(child_agg, Product.id == child_agg.c.pid)
            order_expr = child_agg.c.val.desc() if sort_dir == "desc" else child_agg.c.val.asc()
        elif sort_by in ("product_name", "category", "hs_code"):
            _col = getattr(_C, sort_by)
            child_agg = (
                db.query(_C.parent_id.label("pid"), func.min(_col).label("val"))
                .filter(*_child_filter)
                .group_by(_C.parent_id)
                .subquery()
            )
            parent_q = parent_q.outerjoin(child_agg, Product.id == child_agg.c.pid)
            order_expr = child_agg.c.val.desc() if sort_dir == "desc" else child_agg.c.val.asc()
        else:
            _sort_map = {
                "product_code": Product.product_code,
                "created_at": Product.created_at,
            }
            sort_col = _sort_map.get(sort_by, Product.product_code)
            order_expr = sort_col.desc() if sort_dir == "desc" else sort_col.asc()
        parents = parent_q.order_by(order_expr).offset((page - 1) * per_page).limit(per_page).all()

        parent_ids = [p.id for p in parents]
        # Fetch all children for these parents in one query
        children = db.query(Product).filter(
            Product.parent_id.in_(parent_ids),
            Product.is_active == True,
            Product.deleted_at.is_(None),
        ).all() if parent_ids else []

        children_map: dict[str, list] = {}
        child_ids = []
        for c in children:
            children_map.setdefault(c.parent_id, []).append(c)
            child_ids.append(c.id)

        # Thumbnails for children
        all_ids = parent_ids + child_ids
        thumbnail_map = _build_thumbnail_map(all_ids, db)

        items = []
        for p in parents:
            out = ProductOut.model_validate(p)
            out.thumbnail_url = thumbnail_map.get(p.id)
            kids = children_map.get(p.id, [])
            out.variant_count = len(kids)
            kid_outs = []
            for c in kids:
                co = ProductOut.model_validate(c)
                co.thumbnail_url = thumbnail_map.get(c.id)
                kid_outs.append(co)
            items.append({
                "parent": out,
                "variants": kid_outs,
            })

        return {"items": items, "total": total, "page": page, "per_page": per_page, "grouped": True}

    # --- Flat mode (default) ---
    total = query.count()
    products = query.offset((page - 1) * per_page).limit(per_page).all()

    product_ids = [p.id for p in products]
    thumbnail_map = _build_thumbnail_map(product_ids, db)

    items = []
    for p in products:
        out = ProductOut.model_validate(p)
        out.thumbnail_url = thumbnail_map.get(p.id)
        items.append(out.model_dump())

    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": ceil(total / per_page) if per_page else 0,
    }


def _build_thumbnail_map(product_ids: list[str], db: Session) -> dict[str, str]:
    """Batch-fetch thumbnail URLs for a list of product IDs.

    For parent products that have no direct images, falls back to the
    first child variant's image so the parent row shows a thumbnail.
    """
    thumbnail_map: dict[str, str] = {}
    if not product_ids:
        return thumbnail_map

    # Also fetch child product IDs for any parents in the list
    children = db.query(Product.id, Product.parent_id).filter(
        Product.parent_id.in_(product_ids),
        Product.is_active == True,
        Product.deleted_at.is_(None),
    ).all()
    child_to_parent: dict[str, str] = {c.id: c.parent_id for c in children}
    all_ids = list(set(product_ids) | set(child_to_parent.keys()))

    images = db.query(ProductImage).filter(
        ProductImage.product_id.in_(all_ids)
    ).order_by(ProductImage.is_primary.desc(), ProductImage.created_at).all()

    for img in images:
        rel_path = img.thumbnail_path or img.image_path
        if not rel_path or not (UPLOAD_DIR / rel_path).exists():
            continue
        url = _prod_url(rel_path)
        # Direct product thumbnail
        if img.product_id not in thumbnail_map:
            thumbnail_map[img.product_id] = url
        # Bubble up to parent if parent has no thumbnail yet
        parent_id = child_to_parent.get(img.product_id)
        if parent_id and parent_id not in thumbnail_map:
            thumbnail_map[parent_id] = url

    return thumbnail_map


@router.get("/categories/")
def list_categories(db: Session = Depends(get_db)):
    """Get all distinct product categories from the products table"""
    rows = db.query(Product.category).filter(
        Product.is_active == True,
        Product.deleted_at.is_(None),
        Product.category.isnot(None),
        Product.category != '',
    ).distinct().order_by(Product.category).all()
    return [r[0] for r in rows]


@router.get("/brands/")
def list_brands(db: Session = Depends(get_db)):
    """Get all distinct product brands from the products table"""
    rows = db.query(Product.brand).filter(
        Product.is_active == True,
        Product.deleted_at.is_(None),
        Product.brand.isnot(None),
        Product.brand != '',
    ).distinct().order_by(Product.brand).all()
    return [r[0] for r in rows]


@router.get("/subcategories/")
def list_subcategories(db: Session = Depends(get_db)):
    """Get all distinct product subcategories from the products table"""
    rows = db.query(Product.subcategory).filter(
        Product.is_active == True,
        Product.deleted_at.is_(None),
        Product.subcategory.isnot(None),
        Product.subcategory != '',
    ).distinct().order_by(Product.subcategory).all()
    return [r[0] for r in rows]


@router.get("/materials/")
def list_materials(db: Session = Depends(get_db)):
    """Get all distinct product materials"""
    rows = db.query(Product.material).filter(
        Product.is_active == True,
        Product.deleted_at.is_(None),
        Product.material.isnot(None),
        Product.material != '',
    ).distinct().order_by(Product.material).all()
    return [r[0] for r in rows]


@router.get("/hs-codes/")
def list_hs_codes(db: Session = Depends(get_db)):
    """Get all distinct HS codes"""
    rows = db.query(Product.hs_code).filter(
        Product.is_active == True,
        Product.deleted_at.is_(None),
        Product.hs_code.isnot(None),
        Product.hs_code != '',
    ).distinct().order_by(Product.hs_code).all()
    return [r[0] for r in rows]


@router.get("/part-types/")
def list_part_types(db: Session = Depends(get_db)):
    """Get all distinct part types"""
    rows = db.query(Product.part_type).filter(
        Product.is_active == True,
        Product.deleted_at.is_(None),
        Product.part_type.isnot(None),
        Product.part_type != '',
    ).distinct().order_by(Product.part_type).all()
    return [r[0] for r in rows]


@router.get("/search/")
def search_products(
    q: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Quick search for product selector — only returns orderable variants (children)."""
    from core.serializers import filter_list_for_role
    products = db.query(Product).filter(
        Product.is_active == True,
        Product.deleted_at.is_(None),
        Product.parent_id.isnot(None),  # Only children (orderable)
        or_(
            Product.product_code.ilike(f"%{q}%"),
            Product.product_name.ilike(f"%{q}%"),
            Product.material.ilike(f"%{q}%"),
            Product.part_type.ilike(f"%{q}%"),
            Product.dimension.ilike(f"%{q}%"),
        )
    ).limit(20).all()

    # G-016: apply CLIENT_HIDDEN_FIELDS / FACTORY_HIDDEN_FIELDS stripping.
    # Notes and other sensitive fields must not reach CLIENT/FACTORY callers.
    results = [ProductOut.model_validate(p).model_dump() for p in products]
    return filter_list_for_role(results, current_user.role)


@router.post("/validate-codes/")
def validate_product_codes(req: CodeValidationRequest, db: Session = Depends(get_db)):
    """
    Validate bulk-pasted product codes.
    Returns FOUND, NOT_FOUND, or AMBIGUOUS for each code.
    AMBIGUOUS = same code matches multiple products (different material/name).
    """
    results = []
    for code in req.codes:
        code = code.strip()
        if not code:
            continue

        # Only match children (orderable variants), skip parents
        matches = db.query(Product).filter(
            Product.product_code == code,
            Product.parent_id.isnot(None),  # children only
            Product.is_active == True,
            Product.deleted_at.is_(None),
        ).all()

        def _match_dict(m):
            return {
                "id": m.id,
                "product_code": m.product_code,
                "product_name": m.product_name,
                "material": m.material,
                "part_type": m.part_type,
                "dimension": m.dimension,
                "variant_note": m.variant_note,
                "category": m.category,
                "is_default": m.is_default,
            }

        if len(matches) == 0:
            results.append(CodeValidationResult(code=code, status="NOT_FOUND"))
        elif len(matches) == 1:
            results.append(CodeValidationResult(
                code=code,
                status="FOUND",
                matches=[_match_dict(matches[0])]
            ))
        else:
            # Multiple variants — check if one is marked default
            default_match = next((m for m in matches if m.is_default), None)
            if default_match:
                # Auto-select default variant, still pass all matches for reference
                all_dicts = [_match_dict(m) for m in matches]
                # Put default first
                all_dicts.sort(key=lambda d: (not d["is_default"], d["product_name"]))
                results.append(CodeValidationResult(
                    code=code,
                    status="FOUND",
                    matches=all_dicts,
                ))
            else:
                # No default set — truly ambiguous
                results.append(CodeValidationResult(
                    code=code,
                    status="AMBIGUOUS",
                    matches=[_match_dict(m) for m in matches]
                ))

    return results


@router.post("/bulk-delete/")
def bulk_delete_products(
    req: BulkDeleteRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Soft-delete products (move to bin). Permanent deletion only from bin."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    if not req.product_ids:
        raise HTTPException(status_code=400, detail="No product IDs provided")

    archived = 0
    for pid in req.product_ids:
        product = db.query(Product).filter(
            Product.id == pid,
            Product.deleted_at.is_(None),
        ).first()
        if not product:
            continue

        product.deleted_at = datetime.utcnow()
        product.is_active = False

        # Remove from DRAFT order items (safe — drafts are editable)
        draft_order_ids = [o.id for o in db.query(Order).filter(
            Order.status == "DRAFT", Order.deleted_at.is_(None)
        ).all()]
        if draft_order_ids:
            db.query(OrderItem).filter(
                OrderItem.product_id == pid,
                OrderItem.order_id.in_(draft_order_ids),
            ).delete(synchronize_session='fetch')

        archived += 1

        # If this was the last active child, also soft-delete the orphan parent
        if product.parent_id:
            siblings = db.query(Product).filter(
                Product.parent_id == product.parent_id,
                Product.id != pid,
                Product.is_active == True,
                Product.deleted_at.is_(None),
            ).count()
            if siblings == 0:
                parent = db.query(Product).filter(Product.id == product.parent_id).first()
                if parent and parent.deleted_at is None:
                    parent.deleted_at = datetime.utcnow()
                    parent.is_active = False

    db.commit()
    return {
        "message": f"{archived} product{'s' if archived != 1 else ''} moved to bin",
        "archived": archived,
    }


@router.post("/bulk-update/")
def bulk_update_products(
    req: BulkUpdateRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Update category, material, or hs_code for multiple products at once.
    Only non-null fields in the request are applied."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    if not req.product_ids and not req.product_codes:
        raise HTTPException(status_code=400, detail="No product IDs or codes provided")

    # Build update dict from non-None fields
    updates = {}
    if req.category is not None:
        updates["category"] = req.category
    if req.material is not None:
        updates["material"] = req.material
    if req.hs_code is not None:
        updates["hs_code"] = req.hs_code
    if req.part_type is not None:
        updates["part_type"] = req.part_type
    if req.brand is not None:
        updates["brand"] = req.brand

    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")

    q = db.query(Product).filter(Product.deleted_at.is_(None))
    if req.product_ids:
        q = q.filter(Product.id.in_(req.product_ids))
    elif req.product_codes:
        q = q.filter(Product.product_code.in_(req.product_codes))
    count = q.update(updates, synchronize_session="fetch")

    db.commit()
    return {"updated": count, "fields": list(updates.keys())}


@router.post("/remove-duplicate-images/")
def remove_duplicate_images(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Scan all products — remove duplicate images (same file hash) keeping one per unique image."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    import hashlib

    all_images = db.query(ProductImage).order_by(
        ProductImage.product_id, ProductImage.is_primary.desc(), ProductImage.created_at
    ).all()

    # Group by product_id
    product_images = {}
    for img in all_images:
        product_images.setdefault(img.product_id, []).append(img)

    total_removed = 0
    products_cleaned = 0

    for pid, images in product_images.items():
        seen_hashes = set()
        duplicates = []

        for img in images:
            # Compute hash if missing
            file_hash = img.image_hash
            if not file_hash:
                file_path = os.path.join(str(UPLOAD_DIR), img.image_path)
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as fh:
                        file_hash = hashlib.md5(fh.read(), usedforsecurity=False).hexdigest()
                    img.image_hash = file_hash
                else:
                    # File missing — mark as duplicate to remove DB record
                    duplicates.append(img)
                    continue

            if file_hash in seen_hashes:
                duplicates.append(img)
            else:
                seen_hashes.add(file_hash)

        if duplicates:
            products_cleaned += 1
            for dup in duplicates:
                # Delete file from disk
                file_path = os.path.join(str(UPLOAD_DIR), dup.image_path)
                if os.path.exists(file_path):
                    os.remove(file_path)
                if dup.thumbnail_path:
                    thumb_path = os.path.join(str(UPLOAD_DIR), dup.thumbnail_path)
                    if os.path.exists(thumb_path):
                        os.remove(thumb_path)
                db.delete(dup)
                total_removed += 1

    db.commit()
    return {
        "images_removed": total_removed,
        "products_cleaned": products_cleaned,
        "message": f"Removed {total_removed} duplicate images from {products_cleaned} products",
    }


@router.post("/cleanup-orphan-images/")
def cleanup_orphan_images(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Remove DB records for product images whose files are missing from disk."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    all_images = db.query(ProductImage).all()
    orphaned = 0
    for img in all_images:
        file_path = os.path.join(str(UPLOAD_DIR), img.image_path)
        if not os.path.exists(file_path):
            db.delete(img)
            orphaned += 1
    db.commit()
    return {
        "orphaned_removed": orphaned,
        "total_checked": len(all_images),
        "remaining": len(all_images) - orphaned,
        "message": f"Removed {orphaned} orphaned image records (files missing from disk)",
    }


@router.post("/re-extract-images/")
def re_extract_images(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Re-extract product images from Excel files in the temp directory.
    Supports both floating ('over the cell') and embedded ('fit into cell') images.
    Matches rows to products by product_code column; saves new images, skips duplicates."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")

    import openpyxl
    from services.image_extractor import collect_all_images, save_image_to_disk

    temp_dir = UPLOAD_DIR / "temp"
    xlsx_files = sorted(temp_dir.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if not xlsx_files:
        raise HTTPException(status_code=404, detail="No Excel files found in temp directory")

    # Build product_code → product_id mapping
    all_products = db.query(Product).filter(Product.deleted_at.is_(None)).all()
    code_to_pid = {p.product_code: p.id for p in all_products}

    # Existing image hashes per product for dedup
    existing_hashes: dict[str, set] = {}
    for img_rec in db.query(ProductImage).all():
        existing_hashes.setdefault(img_rec.product_id, set()).add(img_rec.image_hash)

    total_saved = 0
    total_skipped_no_match = 0
    total_skipped_dup = 0
    total_errors = 0
    files_processed = 0

    for excel_path in xlsx_files:
        try:
            # Build row → product_code from the data sheet
            wb_data = openpyxl.load_workbook(str(excel_path), read_only=True)
            ws_data = wb_data.active
            header_row = list(ws_data.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            part_no_indices = [
                ci for ci, val in enumerate(header_row)
                if any(kw in str(val or "").strip().lower() for kw in ("part no", "part code", "mfr"))
            ]
            wb_data.close()

            if not part_no_indices:
                continue

            mfr_col = part_no_indices[1] if len(part_no_indices) >= 2 else part_no_indices[0]

            wb_data2 = openpyxl.load_workbook(str(excel_path), read_only=True)
            ws_data2 = wb_data2.active
            row_to_code: dict[int, str] = {}
            for row_idx, row in enumerate(ws_data2.iter_rows(min_row=2, values_only=True), start=2):
                if row and len(row) > mfr_col and row[mfr_col]:
                    code = str(row[mfr_col]).strip()
                    if code:
                        row_to_code[row_idx] = code
            wb_data2.close()

            if not row_to_code:
                continue

            # Collect images from both drawing and richData styles
            all_images = collect_all_images(str(excel_path))
            if not all_images:
                continue

            saved = 0
            for excel_row, image_data in all_images.items():
                code = row_to_code.get(excel_row)
                if not code:
                    total_skipped_no_match += 1
                    continue
                pid = code_to_pid.get(code)
                if not pid:
                    total_skipped_no_match += 1
                    continue
                try:
                    ok = save_image_to_disk(
                        image_data, pid, "FACTORY_EXCEL", None,
                        db, ProductImage, existing_hashes,
                    )
                    if ok:
                        saved += 1
                    else:
                        total_skipped_dup += 1
                except Exception as e:
                    logger.warning("re_extract row %d: %s", excel_row, e)
                    total_errors += 1

            if saved > 0:
                db.commit()
                total_saved += saved
                files_processed += 1

        except Exception as e:
            logger.warning("re_extract file %s: %s", excel_path.name, e)
            total_errors += 1

    return {
        "images_saved": total_saved,
        "skipped_no_match": total_skipped_no_match,
        "skipped_duplicate": total_skipped_dup,
        "errors": total_errors,
        "files_processed": files_processed,
        "message": f"Extracted {total_saved} images from {files_processed} Excel files",
    }


@router.get("/find-duplicates/")
def find_duplicate_products(db: Session = Depends(get_db)):
    """Find TRUE duplicate children under the same parent (not intentional variants)."""
    from sqlalchemy import func

    # Only check children (parent-child is intentional, not a duplicate)
    # Duplicates = same product_code with >1 child AND no distinguishing variant fields
    dups = db.query(
        Product.product_code, func.count(Product.id).label("count")
    ).filter(
        Product.deleted_at.is_(None),
        Product.product_code.isnot(None),
        Product.product_code != "",
        Product.parent_id.isnot(None),  # Only count children
    ).group_by(Product.product_code).having(func.count(Product.id) > 1).all()

    results = []
    for code, count in dups:
        products = db.query(Product).filter(
            Product.product_code == code,
            Product.parent_id.isnot(None),  # Only children
            Product.deleted_at.is_(None),
        ).order_by(Product.created_at).all()

        # Check which ones are used in orders
        product_ids = [p.id for p in products]
        used_ids = set(
            row[0] for row in db.query(OrderItem.product_id).filter(
                OrderItem.product_id.in_(product_ids)
            ).distinct().all()
        )

        results.append({
            "product_code": code,
            "count": count,
            "products": [{
                "id": p.id,
                "product_code": p.product_code,
                "product_name": p.product_name,
                "created_at": p.created_at.isoformat() if p.created_at else None,
                "used_in_orders": p.id in used_ids,
                "image_count": db.query(ProductImage).filter(ProductImage.product_id == p.id).count(),
            } for p in products],
        })

    return {
        "duplicate_groups": results,
        "total_groups": len(results),
        "total_extra_products": sum(d["count"] - 1 for d in results),
    }


@router.get("/bin/")
def list_bin_products(
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    """List soft-deleted (archived) products"""
    query = db.query(Product).filter(Product.deleted_at.isnot(None))

    if search:
        query = query.filter(
            or_(
                Product.product_code.ilike(f"%{search}%"),
                Product.product_name.ilike(f"%{search}%"),
            )
        )

    total = query.count()
    products = query.order_by(Product.deleted_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    items = []
    for p in products:
        out = ProductOut.model_validate(p)
        out.thumbnail_url = None
        items.append(out)

    return {"items": items, "total": total, "page": page, "per_page": per_page, "pages": ceil(total / per_page) if per_page else 0}


@router.post("/bin/permanent-delete/")
def permanent_delete_bin_products(
    req: BulkDeleteRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Permanently delete soft-deleted products from the bin (removes all data + files)"""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    if not req.product_ids:
        raise HTTPException(status_code=400, detail="No product IDs provided")

    deleted = 0
    images_deleted = 0

    for pid in req.product_ids:
        product = db.query(Product).filter(
            Product.id == pid,
            Product.deleted_at.isnot(None),
        ).first()
        if not product:
            continue

        # Delete all images (files + DB records)
        images = db.query(ProductImage).filter(ProductImage.product_id == pid).all()
        for img in images:
            file_path = os.path.join(str(UPLOAD_DIR), img.image_path)
            if os.path.exists(file_path):
                os.remove(file_path)
            if img.thumbnail_path:
                thumb_file = os.path.join(str(UPLOAD_DIR), img.thumbnail_path)
                if os.path.exists(thumb_file):
                    os.remove(thumb_file)
            db.delete(img)
            images_deleted += 1

        # Preserve order items — snapshot product data, copy image, then nullify the FK
        order_items = db.query(OrderItem).filter(OrderItem.product_id == pid).all()
        for oi in order_items:
            # Populate snapshots if not already set
            if not oi.product_code_snapshot:
                oi.product_code_snapshot = product.product_code
            if not oi.product_name_snapshot:
                oi.product_name_snapshot = product.product_name
            if not oi.material_snapshot:
                oi.material_snapshot = product.material
            if not oi.category_snapshot:
                oi.category_snapshot = product.category
            if not oi.image_path_snapshot:
                primary_img = db.query(ProductImage).filter(
                    ProductImage.product_id == pid
                ).first()
                if primary_img:
                    # Copy image to order folder so it survives product deletion
                    src = UPLOAD_DIR / primary_img.image_path
                    if src.exists():
                        dest_dir = UPLOAD_DIR / "orders" / oi.order_id
                        dest_dir.mkdir(parents=True, exist_ok=True)
                        dest = dest_dir / src.name
                        if not dest.exists():
                            shutil.copy2(str(src), str(dest))
                        oi.image_path_snapshot = f"orders/{oi.order_id}/{src.name}"
                    else:
                        oi.image_path_snapshot = primary_img.image_path
            oi.product_id = None  # Detach from product, keep the order item

        # Delete related records that are NOT order history
        db.query(ClientProductBarcode).filter(ClientProductBarcode.product_id == pid).delete()
        db.query(UnloadedItem).filter(UnloadedItem.product_id == pid).delete()
        db.query(WarehouseStock).filter(WarehouseStock.product_id == pid).delete()
        db.query(PackingListItem).filter(PackingListItem.product_id == pid).delete()
        db.query(ProductVerification).filter(ProductVerification.product_id == pid).delete()
        db.query(AfterSalesItem).filter(AfterSalesItem.product_id == pid).delete()

        # Track parent for orphan cleanup
        parent_id = product.parent_id
        db.delete(product)
        deleted += 1

        # If this was the last child, also delete the orphan parent
        if parent_id:
            remaining = db.query(Product).filter(
                Product.parent_id == parent_id,
            ).count()
            if remaining == 0:
                orphan_parent = db.query(Product).filter(Product.id == parent_id).first()
                if orphan_parent:
                    db.delete(orphan_parent)
                    deleted += 1

    db.commit()

    # Clean up empty product image directories
    try:
        products_dir = os.path.join(str(UPLOAD_DIR), "products")
        if os.path.exists(products_dir):
            for folder in os.listdir(products_dir):
                folder_path = os.path.join(products_dir, folder)
                try:
                    if os.path.isdir(folder_path) and not os.listdir(folder_path):
                        os.rmdir(folder_path)
                except OSError:
                    pass
    except OSError:
        pass

    return {
        "message": f"{deleted} products permanently deleted",
        "deleted": deleted,
        "images_deleted": images_deleted,
    }


@router.post("/bin/restore/")
def restore_bin_products(
    req: BulkDeleteRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Restore soft-deleted products from the bin back to active catalog"""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    if not req.product_ids:
        raise HTTPException(status_code=400, detail="No product IDs provided")

    restored = 0
    for pid in req.product_ids:
        product = db.query(Product).filter(
            Product.id == pid,
            Product.deleted_at.isnot(None),
        ).first()
        if not product:
            continue

        product.deleted_at = None
        product.is_active = True
        restored += 1

    db.commit()
    return {
        "message": f"{restored} product{'s' if restored != 1 else ''} restored",
        "restored": restored,
    }


@router.post("/{product_id}/set-default/")
def set_default_variant(
    product_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Mark a variant child as the default for its parent group.
    Unsets is_default on all siblings, sets it on the target.
    """
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if not product.parent_id:
        raise HTTPException(status_code=400, detail="Cannot set default on a parent product")

    # Unset all siblings
    db.query(Product).filter(
        Product.parent_id == product.parent_id,
        Product.id != product_id,
    ).update({"is_default": False})

    # Set this one
    product.is_default = True
    db.commit()
    db.refresh(product)

    return {"ok": True, "product_id": product_id}


@router.get("/pending-review-list/")
def list_pending_review_v2(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """List product requests pending approval (Admin only). Reads from product_requests table."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    if current_user.role not in ("ADMIN", "SUPER_ADMIN"):
        raise HTTPException(status_code=403, detail="Admin access required")
    from models import Client

    requests = db.query(ProductRequest).filter(
        ProductRequest.status == "PENDING",
    ).order_by(ProductRequest.created_at.desc()).all()

    result = []
    for pr in requests:
        client_name = None
        order_reference = None

        if pr.client_id:
            client = db.query(Client).filter(Client.id == pr.client_id).first()
            client_name = client.company_name if client else None

        if pr.order_id:
            order = db.query(Order).filter(Order.id == pr.order_id).first()
            if order:
                order_reference = order.order_number or order.po_reference or order.id[:8]

        result.append({
            "id": pr.id,
            "product_code": pr.product_code,
            "product_name": pr.product_name,
            "quantity": pr.quantity,
            "status": pr.status,
            "client_name": client_name,
            "order_id": pr.order_id,
            "order_reference": order_reference,
            "requested_by_name": pr.requested_by_name,
            "created_at": pr.created_at.isoformat() if pr.created_at else None,
        })
    return {"products": result, "total": len(result)}


@router.post("/product-requests/{request_id}/map/")
def map_product_request(
    request_id: str,
    data: MapProductRequestBody,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Admin maps a product request to an existing real product.
    Creates OrderItem in the linked order and grants client access."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    if current_user.role not in ("ADMIN", "SUPER_ADMIN"):
        raise HTTPException(status_code=403, detail="Admin access required")
    from models import Notification, ClientProductAccess, ClientBrandAccess, User

    pr = db.query(ProductRequest).filter(ProductRequest.id == request_id).first()
    if not pr or pr.status != "PENDING":
        raise HTTPException(status_code=404, detail="Pending product request not found")

    # Find the target (real) product
    target = db.query(Product).filter(
        Product.id == data.target_product_id,
        Product.deleted_at.is_(None),
    ).first()
    if not target:
        raise HTTPException(status_code=404, detail="Target product not found")

    original_code = pr.product_code

    # Update request status
    pr.status = "MAPPED"
    pr.mapped_product_id = target.id
    pr.reviewed_by = current_user.id
    pr.reviewed_at = datetime.utcnow()

    # Create OrderItem in the linked order
    order_item = OrderItem(
        order_id=pr.order_id,
        product_id=target.id,
        quantity=pr.quantity,
        product_code_snapshot=target.product_code,
        product_name_snapshot=target.product_name,
        material_snapshot=target.material,
        category_snapshot=target.category,
        part_type_snapshot=getattr(target, 'part_type', None),
        dimension_snapshot=getattr(target, 'dimension', None),
        variant_note_snapshot=getattr(target, 'variant_note', None),
        status="ACTIVE",
    )
    db.add(order_item)

    # Grant client per-product access if not already accessible
    client_id = pr.client_id
    if client_id:
        brand_access = db.query(ClientBrandAccess).filter(
            ClientBrandAccess.client_id == client_id,
            ClientBrandAccess.brand == target.brand,
        ).first() if target.brand else None

        if not brand_access:
            existing_access = db.query(ClientProductAccess).filter(
                ClientProductAccess.client_id == client_id,
                ClientProductAccess.product_id == target.id,
            ).first()
            if not existing_access:
                db.add(ClientProductAccess(
                    client_id=client_id,
                    product_id=target.id,
                    added_via="MAPPED",
                    source_code=original_code,
                ))
                if target.parent_id:
                    existing_parent = db.query(ClientProductAccess).filter(
                        ClientProductAccess.client_id == client_id,
                        ClientProductAccess.product_id == target.parent_id,
                    ).first()
                    if not existing_parent:
                        db.add(ClientProductAccess(
                            client_id=client_id,
                            product_id=target.parent_id,
                            added_via="MAPPED",
                            source_code=original_code,
                        ))

        # Notify client
        existing_notif = db.query(Notification).filter(
            Notification.client_id == client_id,
            Notification.notification_type == "PRODUCT_MAPPED",
            Notification.is_read == False,
        ).first()

        if existing_notif:
            existing_notif.count = existing_notif.count + 1
            existing_notif.message = f"{existing_notif.count} products were mapped to correct parts"
            existing_notif.updated_at = datetime.utcnow()
        else:
            db.add(Notification(
                user_id=pr.requested_by,
                client_id=client_id,
                title="Product Mapped",
                message=f"'{original_code}' was mapped to '{target.product_code} - {target.product_name}'",
                notification_type="PRODUCT_MAPPED",
                resource_type="product",
                resource_id=target.id,
            ))

    db.commit()
    return {
        "status": "mapped",
        "original_code": original_code,
        "target_code": target.product_code,
        "target_name": target.product_name,
    }


@router.post("/product-requests/{request_id}/reject/")
def reject_product_request(
    request_id: str,
    data: RejectProductRequestBody,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Admin rejects a product request with a remark. Notifies client."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    if current_user.role not in ("ADMIN", "SUPER_ADMIN"):
        raise HTTPException(status_code=403, detail="Admin access required")
    from models import Notification

    pr = db.query(ProductRequest).filter(ProductRequest.id == request_id).first()
    if not pr or pr.status != "PENDING":
        raise HTTPException(status_code=404, detail="Pending product request not found")

    original_code = pr.product_code
    original_name = pr.product_name

    # Update request status
    pr.status = "REJECTED"
    pr.reject_remark = data.remark
    pr.reviewed_by = current_user.id
    pr.reviewed_at = datetime.utcnow()

    # Notify client
    if pr.client_id:
        db.add(Notification(
            user_id=pr.requested_by,
            client_id=pr.client_id,
            title="Product Request Rejected",
            message=f"'{original_code} - {original_name}' was rejected: {data.remark}",
            notification_type="PRODUCT_REJECTED",
            resource_type="product_request",
            resource_id=request_id,
        ))

    db.commit()
    return {"status": "rejected", "code": original_code, "remark": data.remark}


@router.get("/{product_id}/")
def get_product(product_id: str, db: Session = Depends(get_db)):
    """Get single product by ID"""
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    out = ProductOut.model_validate(product)
    # Fetch primary/first image as thumbnail
    img = db.query(ProductImage).filter(
        ProductImage.product_id == product_id
    ).order_by(ProductImage.is_primary.desc(), ProductImage.created_at).first()
    # If parent has no direct image, fall back to first child's image
    if not img and product.parent_id is None:
        child_ids = [c.id for c in db.query(Product.id).filter(
            Product.parent_id == product_id,
            Product.is_active == True,
            Product.deleted_at.is_(None),
        ).all()]
        if child_ids:
            img = db.query(ProductImage).filter(
                ProductImage.product_id.in_(child_ids)
            ).order_by(ProductImage.is_primary.desc(), ProductImage.created_at).first()
    if img:
        rel_path = img.thumbnail_path or img.image_path
        if rel_path and (UPLOAD_DIR / rel_path).exists():
            out.thumbnail_url = _prod_url(rel_path)
    return out


@router.get("/check-variants/{product_code}/")
def check_existing_variants(product_code: str, db: Session = Depends(get_db)):
    """Check if a product code already has existing variants.
    Used by ProductForm to decide whether to show the variant resolution dialog.
    """
    parent = db.query(Product).filter(
        Product.product_code == product_code,
        Product.parent_id.is_(None),
        Product.deleted_at.is_(None),
    ).first()

    if not parent:
        return {"parent_id": None, "variant_count": 0, "variants": []}

    children = db.query(Product).filter(
        Product.parent_id == parent.id,
        Product.is_active == True,
        Product.deleted_at.is_(None),
    ).all()

    variants = [{
        "id": c.id,
        "product_name": c.product_name,
        "material": c.material,
        "dimension": c.dimension,
        "part_type": c.part_type,
        "variant_note": c.variant_note,
        "is_default": c.is_default,
        "category": c.category or parent.category,
        "hs_code": c.hs_code or parent.hs_code,
        "brand": c.brand or parent.brand,
        "oem_reference": c.oem_reference or parent.oem_reference,
    } for c in children]

    return {
        "parent_id": parent.id,
        "variant_count": len(children),
        "parent_code": parent.product_code,
        "parent_category": parent.category,
        "parent_hs_code": parent.hs_code,
        "parent_brand": parent.brand,
        "variants": variants,
    }


# DEPRECATED: quick-add endpoint removed — Quick Add is now frontend-only.
# Product requests are created via POST /orders/client-inquiry/ with quick_add_items.
# Old /pending-review/ endpoint also removed — use /pending-review-list/ instead.


@router.post("/product-requests/{request_id}/approve/")
def approve_product_request(
    request_id: str,
    data: ApproveProductRequestBody,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Admin approves a product request. Creates real Product + OrderItem."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    if current_user.role not in ("ADMIN", "SUPER_ADMIN"):
        raise HTTPException(status_code=403, detail="Admin access required")
    from models import Notification, ClientProductAccess, ClientBrandAccess
    from enums import ApprovalStatus

    pr = db.query(ProductRequest).filter(ProductRequest.id == request_id).first()
    if not pr or pr.status != "PENDING":
        raise HTTPException(status_code=404, detail="Pending product request not found")

    # Check for duplicate product_code
    existing = db.query(Product).filter(
        Product.product_code == pr.product_code,
        Product.parent_id.isnot(None),
        Product.deleted_at.is_(None),
    ).first()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"Product code '{pr.product_code}' already exists. Use Map instead.",
        )

    # Create parent product
    parent = Product(
        product_code=pr.product_code,
        product_name=f"[{pr.product_code}]",
        approval_status=ApprovalStatus.APPROVED.value,
        category=data.category,
        brand=data.brand,
    )
    db.add(parent)
    db.flush()

    # Create child variant
    child = Product(
        parent_id=parent.id,
        product_code=pr.product_code,
        product_name=data.product_name or pr.product_name,
        product_name_chinese=data.product_name_chinese,
        category=data.category,
        material=data.material,
        dimension=data.dimension,
        part_type=data.part_type,
        brand=data.brand,
        hs_code=data.hs_code,
        unit_weight_kg=data.unit_weight_kg,
        approval_status=ApprovalStatus.APPROVED.value,
        is_default=True,
    )
    db.add(child)
    db.flush()

    # Update request status
    pr.status = "APPROVED"
    pr.created_product_id = child.id
    pr.reviewed_by = current_user.id
    pr.reviewed_at = datetime.utcnow()

    # Create OrderItem linking product to the order
    order_item = OrderItem(
        order_id=pr.order_id,
        product_id=child.id,
        quantity=pr.quantity,
        product_code_snapshot=child.product_code,
        product_name_snapshot=child.product_name,
        material_snapshot=child.material,
        category_snapshot=child.category,
        part_type_snapshot=child.part_type,
        dimension_snapshot=child.dimension,
        variant_note_snapshot=child.variant_note,
        status="ACTIVE",
    )
    db.add(order_item)

    # Grant client per-product access
    client_id = pr.client_id
    if client_id:
        brand_access = db.query(ClientBrandAccess).filter(
            ClientBrandAccess.client_id == client_id,
            ClientBrandAccess.brand == child.brand,
        ).first() if child.brand else None

        if not brand_access:
            existing_access = db.query(ClientProductAccess).filter(
                ClientProductAccess.client_id == client_id,
                ClientProductAccess.product_id == child.id,
            ).first()
            if not existing_access:
                db.add(ClientProductAccess(
                    client_id=client_id,
                    product_id=child.id,
                    added_via="APPROVED",
                    source_code=pr.product_code,
                ))
                db.add(ClientProductAccess(
                    client_id=client_id,
                    product_id=parent.id,
                    added_via="APPROVED",
                    source_code=pr.product_code,
                ))

    # Notify client
    existing_notif = db.query(Notification).filter(
        Notification.client_id == client_id,
        Notification.notification_type == "PRODUCT_APPROVED",
        Notification.is_read == False,
    ).first()

    if existing_notif:
        existing_notif.count = existing_notif.count + 1
        existing_notif.message = f"{existing_notif.count} products confirmed and added"
        existing_notif.updated_at = datetime.utcnow()
    else:
        db.add(Notification(
            user_id=pr.requested_by,
            client_id=client_id,
            title="Products Approved",
            message="1 product confirmed and added",
            notification_type="PRODUCT_APPROVED",
            resource_type="product",
            resource_id=child.id,
        ))

    db.commit()

    return {"id": child.id, "product_code": child.product_code, "approval_status": "APPROVED"}


@router.post("/")
def create_product(
    data: ProductCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Create new product variant (child). Auto-creates parent if needed.
    If replace_variant_id is set, updates that variant instead of creating new.
    """
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    # Find or create parent for this product_code
    parent = db.query(Product).filter(
        Product.product_code == data.product_code,
        Product.parent_id.is_(None),
        Product.deleted_at.is_(None),
    ).first()

    if not parent:
        # Auto-create parent product
        parent = Product(
            product_code=data.product_code,
            product_name=f"[{data.product_code}]",
            category=data.category,
            subcategory=data.subcategory,
            hs_code=data.hs_code,
            hs_code_description=data.hs_code_description,
            brand=data.brand,
            oem_reference=data.oem_reference,
            is_active=True,
            approval_status="APPROVED",
        )
        db.add(parent)
        db.flush()

    # --- REPLACE mode: update existing variant ---
    if data.replace_variant_id:
        existing_variant = db.query(Product).filter(
            Product.id == data.replace_variant_id,
            Product.parent_id == parent.id,
            Product.deleted_at.is_(None),
        ).first()
        if not existing_variant:
            raise HTTPException(status_code=404, detail="Variant to replace not found")

        # Check name uniqueness (exclude the variant being replaced)
        if data.product_name != existing_variant.product_name:
            name_conflict = db.query(Product).filter(
                Product.product_name == data.product_name,
                Product.id != existing_variant.id,
                Product.deleted_at.is_(None),
            ).first()
            if name_conflict:
                raise HTTPException(status_code=400, detail="Product name already exists")

        # Update variant fields
        update_data = data.model_dump(exclude={"replace_variant_id"})
        for key, value in update_data.items():
            setattr(existing_variant, key, value)
        existing_variant.parent_id = parent.id
        existing_variant.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(existing_variant)
        return ProductOut.model_validate(existing_variant)

    # --- ADD NEW mode: create new child variant ---
    # Check product_name uniqueness
    existing = db.query(Product).filter(
        Product.product_name == data.product_name,
        Product.deleted_at.is_(None),
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Product name already exists")

    create_data = data.model_dump(exclude={"replace_variant_id"})
    product = Product(**create_data, parent_id=parent.id)
    db.add(product)
    db.flush()

    # Auto-set is_default if this is the first/only child
    sibling_count = db.query(Product).filter(
        Product.parent_id == parent.id,
        Product.id != product.id,
        Product.deleted_at.is_(None),
    ).count()
    if sibling_count == 0:
        product.is_default = True

    db.commit()
    db.refresh(product)
    return ProductOut.model_validate(product)


@router.put("/{product_id}/")
def update_product(
    product_id: str,
    data: ProductCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Update existing product (variant or parent)"""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Check name uniqueness (exclude self)
    existing = db.query(Product).filter(
        Product.product_name == data.product_name,
        Product.id != product_id,
        Product.deleted_at.is_(None),
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Product name already exists")

    for key, value in data.model_dump().items():
        if key == "parent_id":
            continue  # Never overwrite parent_id via generic update
        setattr(product, key, value)

    db.commit()
    db.refresh(product)
    return ProductOut.model_validate(product)


@router.delete("/{product_id}/")
def delete_product(
    product_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Soft delete product"""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    product.deleted_at = datetime.utcnow()
    product.is_active = False

    # If this was the last active child, also soft-delete the orphan parent
    if product.parent_id:
        siblings = db.query(Product).filter(
            Product.parent_id == product.parent_id,
            Product.id != product_id,
            Product.is_active == True,
            Product.deleted_at.is_(None),
        ).count()
        if siblings == 0:
            parent = db.query(Product).filter(Product.id == product.parent_id).first()
            if parent:
                parent.deleted_at = datetime.utcnow()
                parent.is_active = False

    db.commit()
    return {"message": "Product deleted"}


@router.get("/{product_id}/images/")
def list_product_images(product_id: str, db: Session = Depends(get_db)):
    """Get all images for a product.

    For parent products, also includes images from child variants so
    the parent detail page shows all variant images.
    """
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Collect IDs to query: this product + children (if parent)
    query_ids = [product_id]
    if product.parent_id is None:
        child_ids = [c.id for c in db.query(Product.id).filter(
            Product.parent_id == product_id,
            Product.is_active == True,
            Product.deleted_at.is_(None),
        ).all()]
        query_ids.extend(child_ids)

    images = db.query(ProductImage).filter(
        ProductImage.product_id.in_(query_ids)
    ).order_by(ProductImage.is_primary.desc(), ProductImage.created_at).all()

    results = []
    orphaned_ids = []
    for img in images:
        # Check if image file still exists on disk
        img_file = UPLOAD_DIR / img.image_path if img.image_path else None
        if img_file and not img_file.exists():
            orphaned_ids.append(img.id)
            continue
        thumb_rel = img.thumbnail_path or img.image_path
        out = ProductImageOut(
            id=img.id,
            product_id=img.product_id,
            image_path=img.image_path,
            image_url=_prod_url(img.image_path),
            thumbnail_url=_prod_url(thumb_rel) if (UPLOAD_DIR / thumb_rel).exists() else _prod_url(img.image_path),
            source_type=img.source_type,
            source_order_id=img.source_order_id,
            width=img.width,
            height=img.height,
            file_size=img.file_size,
            is_primary=img.is_primary,
        )
        results.append(out)

    # Auto-cleanup orphaned DB records (files missing from disk)
    if orphaned_ids:
        db.query(ProductImage).filter(ProductImage.id.in_(orphaned_ids)).delete(synchronize_session=False)
        db.commit()

    return results


ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tiff"}
MAX_IMAGE_SIZE = 10 * 1024 * 1024  # 10 MB


@router.post("/{product_id}/images/upload/", status_code=201)
def upload_product_image(
    product_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Upload a single image for a product. Saves to disk and creates DB record."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Validate file extension
    _, ext = os.path.splitext(file.filename or "")
    if ext.lower() not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type '{ext}'. Allowed: {', '.join(sorted(ALLOWED_IMAGE_EXTENSIONS))}",
        )

    # Read file contents
    contents = file.file.read()
    if len(contents) > MAX_IMAGE_SIZE:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 10 MB.")
    if len(contents) == 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    # Compute hash for deduplication (not for security)
    file_hash = hashlib.md5(contents, usedforsecurity=False).hexdigest()

    # Check for duplicate image on this product
    existing = db.query(ProductImage).filter(
        ProductImage.product_id == product_id,
        ProductImage.image_hash == file_hash,
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="This image already exists for this product.")

    # Save to disk
    from models import gen_uuid
    image_id = gen_uuid()
    safe_code = product.product_code.replace("/", "_").replace("\\", "_")
    image_dir = UPLOAD_DIR / "products" / safe_code
    image_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{image_id}{ext.lower()}"
    file_path = image_dir / filename
    file_path.write_bytes(contents)

    # Get image dimensions if Pillow is available
    width, height = None, None
    try:
        from PIL import Image as PILImage
        from io import BytesIO
        img = PILImage.open(BytesIO(contents))
        width, height = img.size
    except Exception:
        pass

    # Determine if this is the first image (make it primary)
    existing_count = db.query(ProductImage).filter(
        ProductImage.product_id == product_id
    ).count()
    is_primary = existing_count == 0

    # Create DB record
    rel_path = f"products/{safe_code}/{filename}"
    db_image = ProductImage(
        id=image_id,
        product_id=product_id,
        image_path=rel_path,
        image_hash=file_hash,
        source_type="MANUAL_UPLOAD",
        width=width,
        height=height,
        file_size=len(contents),
        is_primary=is_primary,
    )
    db.add(db_image)
    db.commit()

    return {
        "id": image_id,
        "product_id": product_id,
        "image_url": _prod_url(rel_path),
        "is_primary": is_primary,
    }


@router.delete("/{product_id}/images/{image_id}/")
def delete_product_image(
    product_id: str,
    image_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Delete a single product image (DB record + file on disk)"""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Internal access only")
    image = db.query(ProductImage).filter(
        ProductImage.id == image_id,
        ProductImage.product_id == product_id,
    ).first()
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")

    # Delete file from disk
    file_path = os.path.join(str(UPLOAD_DIR), image.image_path)
    if os.path.exists(file_path):
        os.remove(file_path)

    # Delete thumbnail from disk
    if image.thumbnail_path:
        thumb_file = os.path.join(str(UPLOAD_DIR), image.thumbnail_path)
        if os.path.exists(thumb_file):
            os.remove(thumb_file)

    # Delete DB record
    db.delete(image)
    db.commit()
    return {"message": "Image deleted"}
