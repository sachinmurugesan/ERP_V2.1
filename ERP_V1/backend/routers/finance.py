"""
HarvestERP - Finance Router
Client payments (advance/balance) and factory payments with per-remittance tracking.
"""
import io
from rate_limiter import limiter
import os
import uuid as uuid_mod
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

import json

from database import get_db
from models import Payment, FactoryPayment, Order, ProformaInvoice, PIRevision, ExchangeRate, ClientCredit, FactoryCredit, Client, Factory, PaymentAuditLog
from enums import PaymentType, PaymentMethod, Currency, CreditStatus, PaymentVerificationStatus
from core.security import CurrentUser, get_current_user, get_scoped_query, verify_resource_access, require_factory_financial
from core.serializers import filter_for_role, filter_list_for_role
from core.file_upload import stream_upload_to_disk, sanitize_filename
from config import UPLOAD_DIR
from schemas.finance import (
    PaymentCreate, PaymentUpdate, PaymentOut, PaymentResponse,
    FactoryPaymentCreate, FactoryPaymentUpdate, FactoryPaymentOut, FactoryPaymentResponse,
    ClientCreditOut, ClientCreditResponse, ApplyCreditRequest,
    FactoryCreditOut, FactoryCreditResponse, ApplyFactoryCreditRequest,
    VerifyPaymentRequest,
)

router = APIRouter()


# ========================================
# Endpoints
# ========================================


# ========================================
# CLIENT PAYMENTS
# ========================================

@router.get("/orders/{order_id}/payments/")
def list_payments(order_id: str, db: Session = Depends(get_db), current_user: CurrentUser = Depends(get_current_user)):
    """List all client payments for an order + financial summary."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # RLS: CLIENT/FACTORY users can only see payments for their own orders
    if current_user.role in ("CLIENT", "FACTORY"):
        if not verify_resource_access(Order, order_id, db, current_user):
            raise HTTPException(status_code=403, detail="Access denied")

    payments = (
        db.query(Payment)
        .filter(Payment.order_id == order_id)
        .order_by(Payment.payment_date.desc(), Payment.created_at.desc())
        .all()
    )

    # Get PI total for financial summary
    pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order_id).first()
    pi_total_inr = pi.total_inr if pi else 0
    advance_percent = pi.advance_percent if pi else 30

    verified_payments = [p for p in payments if getattr(p, 'verification_status', 'VERIFIED') == 'VERIFIED']
    total_paid_inr = sum(p.amount_inr for p in verified_payments)
    balance_inr = pi_total_inr - total_paid_inr

    # Calculate revised totals (active items only, excludes UNLOADED/migrated)
    active_items = [i for i in order.items if i.status == "ACTIVE"]
    unloaded_count = sum(1 for i in order.items if i.status == "UNLOADED")

    exchange_rate = order.exchange_rate or 1.0

    # Detect transparency client — use client_factory_price * rate instead of selling_price_inr
    client = db.query(Client).filter(Client.id == order.client_id).first()
    is_transparency = client and client.client_type == "TRANSPARENCY"

    if is_transparency:
        revised_client_total = sum(
            round(float(i.client_factory_price or 0) * exchange_rate, 2) * i.quantity
            for i in active_items
        )
    else:
        revised_client_total = sum((i.selling_price_inr or 0) * i.quantity for i in active_items)

    revised_factory_cny = sum((i.factory_price or 0) * i.quantity for i in active_items)
    revised_factory_inr = round(revised_factory_cny * exchange_rate, 2)

    # Original factory total (all items including unloaded)
    all_items = [i for i in order.items if i.status in ("ACTIVE", "UNLOADED")]
    original_factory_cny = sum((i.factory_price or 0) * i.quantity for i in all_items)
    original_factory_inr = round(original_factory_cny * exchange_rate, 2)

    # Detect divergence: unloaded items OR after-sales price corrections changed the effective total
    has_divergence = unloaded_count > 0 or round(revised_client_total, 2) != round(pi_total_inr, 2)

    # Ensure credit is recalculated when totals have diverged
    if has_divergence:
        _recalculate_credit(order_id, db)

    # Build payment list with utilization tracking
    # Sort chronologically for utilization calculation (oldest first)
    effective_total = round(revised_client_total, 2) if has_divergence else pi_total_inr
    payments_chrono = sorted(payments, key=lambda p: (p.payment_date or p.created_at, p.created_at))
    running_total = 0.0
    utilization_map = {}  # payment_id -> {utilized_inr, surplus_inr}
    for p in payments_chrono:
        remaining = max(effective_total - running_total, 0)
        utilized = min(p.amount_inr, remaining)
        surplus = round(p.amount_inr - utilized, 2)
        utilization_map[p.id] = {
            "utilized_inr": round(utilized, 2),
            "surplus_inr": surplus if surplus > 0 else 0,
        }
        running_total += p.amount_inr

    payment_list = [
        {
            "id": p.id,
            "order_id": p.order_id,
            "payment_type": p.payment_type,
            "amount": p.amount,
            "currency": p.currency,
            "exchange_rate": p.exchange_rate,
            "amount_inr": p.amount_inr,
            "method": p.method,
            "reference": p.reference,
            "notes": p.notes,
            "payment_date": p.payment_date.isoformat() if p.payment_date else "",
            "created_at": p.created_at.isoformat() if p.created_at else "",
            "verification_status": getattr(p, 'verification_status', 'VERIFIED'),
            "proof_file_path": getattr(p, 'proof_file_path', None),
            "rejection_reason": getattr(p, 'rejection_reason', None),
            "submitted_by": getattr(p, 'submitted_by', None),
            **utilization_map.get(p.id, {"utilized_inr": p.amount_inr, "surplus_inr": 0}),
        }
        for p in payments  # Keep original desc order for display
    ]

    # Factory paid total (for revised section)
    factory_payments = db.query(FactoryPayment).filter(FactoryPayment.order_id == order_id).all()
    factory_paid_inr = sum(p.amount_inr for p in factory_payments)

    # D-010: OPERATIONS role cannot see factory cost or margin fields
    is_ops = current_user.role == "OPERATIONS"

    return {
        "payments": payment_list,
        "summary": {
            "pi_total_inr": round(pi_total_inr, 2),
            "advance_percent": advance_percent,
            "total_paid_inr": round(total_paid_inr, 2),
            "balance_inr": round(balance_inr, 2),
            "payment_count": len(payments),
            "paid_percent": round((total_paid_inr / pi_total_inr * 100) if pi_total_inr > 0 else 0, 1),
            # Revised PI data (for post-migration or after-sales price correction)
            "has_revisions": has_divergence,
            "unloaded_count": unloaded_count,
            "revised_client_total_inr": round(revised_client_total, 2),
            "revised_balance_inr": round(revised_client_total - total_paid_inr, 2),
            # Factory cost fields — stripped for OPERATIONS per D-010
            "original_factory_total_cny": None if is_ops else round(original_factory_cny, 2),
            "original_factory_total_inr": None if is_ops else original_factory_inr,
            "revised_factory_total_cny": None if is_ops else round(revised_factory_cny, 2),
            "revised_factory_total_inr": None if is_ops else revised_factory_inr,
            "factory_paid_inr": None if is_ops else round(factory_paid_inr, 2),
            "revised_factory_balance_inr": None if is_ops else round(revised_factory_inr - factory_paid_inr, 2),
        },
    }


def _recalculate_credit(order_id: str, db: Session):
    """Recalculate client credit for an order based on current payment totals.
    Surplus = real payments - effective PI total. The AVAILABLE credit = surplus - already-applied credits.
    CREDIT-method payments are excluded (they are internal reallocations, not real money).
    When items have been migrated (UNLOADED), use the revised total from active items only.
    """
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        return
    pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order_id).first()

    # Always recalculate effective total from active items (handles after-sales price corrections)
    active_items = [i for i in order.items if i.status == "ACTIVE"]
    has_unloaded = any(i.status == "UNLOADED" for i in order.items)
    recalc_total = round(sum((i.selling_price_inr or 0) * i.quantity for i in active_items), 2)
    stored_total = pi.total_inr if pi else 0
    has_divergence = has_unloaded or recalc_total != stored_total

    if has_divergence:
        effective_pi_total = recalc_total
    elif pi and stored_total > 0:
        effective_pi_total = stored_total
    else:
        return  # No PI and no divergence — nothing to recalculate

    # For revised PI (migrated/corrected items): count ALL payments including applied credits,
    # because the credit was allocated to cover the original higher PI.
    # For original PI: exclude CREDIT to avoid double-counting.
    if has_divergence:
        total_paid = db.query(func.coalesce(func.sum(Payment.amount_inr), 0)).filter(
            Payment.order_id == order_id,
            Payment.verification_status == 'VERIFIED',
        ).scalar()
    else:
        total_paid = db.query(func.coalesce(func.sum(Payment.amount_inr), 0)).filter(
            Payment.order_id == order_id,
            Payment.method != "CREDIT",
            Payment.verification_status == 'VERIFIED',
        ).scalar()

    existing_credit = db.query(ClientCredit).filter(
        ClientCredit.source_order_id == order_id,
        ClientCredit.status == CreditStatus.AVAILABLE.value,
    ).first()

    if total_paid > effective_pi_total:
        surplus = round(total_paid - effective_pi_total, 2)

        # Subtract credits already applied FROM this order to other orders
        applied_total = db.query(func.coalesce(func.sum(ClientCredit.amount), 0)).filter(
            ClientCredit.source_order_id == order_id,
            ClientCredit.status == CreditStatus.APPLIED.value,
        ).scalar()
        target_available = round(surplus - applied_total, 2)

        if target_available > 0:
            note = (
                f"Surplus after PI revision on {order.order_number or order_id}"
                if has_unloaded
                else f"Overpayment on {order.order_number or order_id}"
            )
            if existing_credit:
                existing_credit.amount = target_available
                existing_credit.notes = note
            else:
                db.add(ClientCredit(
                    client_id=order.client_id,
                    source_order_id=order_id,
                    amount=target_available,
                    notes=note,
                ))
        else:
            if existing_credit:
                db.delete(existing_credit)
    else:
        if existing_credit:
            db.delete(existing_credit)

    db.commit()


def _recalculate_factory_credit(order_id: str, db: Session):
    """Recalculate factory credit for an order based on current factory payment totals.
    Surplus = real payments - factory bill. The AVAILABLE credit = surplus - already-applied credits.
    CREDIT-method payments are excluded (they are internal reallocations, not real money).
    """
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        return

    # Compute factory bill from order items
    active_items = [i for i in order.items if i.status == "ACTIVE"]
    factory_total_cny = sum((i.factory_price or 0) * i.quantity for i in active_items)
    exchange_rate = order.exchange_rate or 1.0
    factory_total_inr = round(factory_total_cny * exchange_rate, 2)

    if factory_total_inr <= 0:
        return

    # Only count real payments (exclude CREDIT — internal reallocations)
    total_paid = db.query(func.coalesce(func.sum(FactoryPayment.amount_inr), 0)).filter(
        FactoryPayment.order_id == order_id,
        FactoryPayment.method != "CREDIT",
    ).scalar()

    existing_credit = db.query(FactoryCredit).filter(
        FactoryCredit.source_order_id == order_id,
        FactoryCredit.status == CreditStatus.AVAILABLE.value,
    ).first()

    if total_paid > factory_total_inr:
        surplus = round(total_paid - factory_total_inr, 2)

        # Subtract credits already applied FROM this order to other orders
        applied_total = db.query(func.coalesce(func.sum(FactoryCredit.amount), 0)).filter(
            FactoryCredit.source_order_id == order_id,
            FactoryCredit.status == CreditStatus.APPLIED.value,
        ).scalar()
        target_available = round(surplus - applied_total, 2)

        if target_available > 0:
            if existing_credit:
                existing_credit.amount = target_available
            else:
                db.add(FactoryCredit(
                    factory_id=order.factory_id,
                    source_order_id=order_id,
                    amount=target_available,
                    notes=f"Overpayment on {order.order_number or order_id}",
                ))
        else:
            if existing_credit:
                db.delete(existing_credit)
    else:
        if existing_credit:
            db.delete(existing_credit)

    db.commit()


# ========================================
# PAYMENT AUDIT HELPERS
# ========================================

def _payment_snapshot(payment) -> dict:
    """Serialize Payment or FactoryPayment for audit log."""
    snap = {
        "id": payment.id,
        "order_id": payment.order_id,
        "amount": payment.amount,
        "currency": payment.currency,
        "exchange_rate": payment.exchange_rate,
        "amount_inr": payment.amount_inr,
        "method": payment.method,
        "reference": payment.reference,
        "notes": payment.notes,
        "payment_date": payment.payment_date.isoformat() if payment.payment_date else None,
    }
    # Client payment has payment_type
    if hasattr(payment, "payment_type"):
        snap["payment_type"] = payment.payment_type
    return snap


def _log_payment_audit(db: Session, *, payment_id: str, table: str, order_id: str,
                       action: str, before: dict = None, after: dict = None,
                       changed: list = None, reason: str = None):
    """Append a payment audit log entry."""
    entry = PaymentAuditLog(
        payment_id=payment_id,
        payment_table=table,
        order_id=order_id,
        action=action,
        before_data=json.dumps(before) if before else None,
        after_data=json.dumps(after) if after else None,
        changed_fields=json.dumps(changed) if changed else None,
        reason=reason,
    )
    db.add(entry)


@router.post("/orders/{order_id}/payments/")
def create_payment(order_id: str, data: PaymentCreate, current_user: CurrentUser = Depends(get_current_user), db: Session = Depends(get_db)):
    """Record a client payment. amount_inr computed server-side."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    amount_inr = round(data.amount * data.exchange_rate, 2)

    payment = Payment(
        order_id=order_id,
        payment_type=data.payment_type,
        amount=data.amount,
        currency=data.currency,
        exchange_rate=data.exchange_rate,
        amount_inr=amount_inr,
        method=data.method,
        reference=data.reference,
        notes=data.notes,
        payment_date=date.fromisoformat(data.payment_date),
        verification_status=PaymentVerificationStatus.VERIFIED.value,
        submitted_by=None,
    )
    db.add(payment)
    db.commit()
    db.refresh(payment)

    # Audit log (domain-specific)
    _log_payment_audit(db, payment_id=payment.id, table="client", order_id=order_id,
                       action="CREATE", after=_payment_snapshot(payment))

    # Global audit trail
    from core.audit import log_audit_event
    log_audit_event(db, current_user, "PAYMENT_CREATE", "payment", payment.id,
                    new_values={"amount": data.amount, "currency": data.currency, "amount_inr": amount_inr, "method": data.method},
                    metadata={"order_id": order_id, "type": "client"})
    db.commit()

    # Recalculate credit (auto-creates if overpayment)
    _recalculate_credit(order_id, db)

    return PaymentOut(
        id=payment.id,
        order_id=payment.order_id,
        payment_type=payment.payment_type,
        amount=payment.amount,
        currency=payment.currency,
        exchange_rate=payment.exchange_rate,
        amount_inr=payment.amount_inr,
        method=payment.method,
        reference=payment.reference,
        notes=payment.notes,
        payment_date=payment.payment_date.isoformat(),
        created_at=payment.created_at.isoformat(),
    )


@router.delete("/orders/{order_id}/payments/{payment_id}/")
def delete_payment(order_id: str, payment_id: str, current_user: CurrentUser = Depends(get_current_user), db: Session = Depends(get_db)):
    """Delete a client payment."""
    payment = db.query(Payment).filter(
        Payment.id == payment_id,
        Payment.order_id == order_id,
    ).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")

    # Capture before snapshot for audit
    before_snap = _payment_snapshot(payment)

    # If this was a CREDIT payment, restore the original credit to AVAILABLE
    if payment.method == "CREDIT":
        applied_credit = db.query(ClientCredit).filter(
            ClientCredit.applied_to_order_id == order_id,
            ClientCredit.status == CreditStatus.APPLIED.value,
            ClientCredit.amount == payment.amount_inr,
        ).first()
        if applied_credit:
            applied_credit.status = CreditStatus.AVAILABLE.value
            applied_credit.applied_to_order_id = None
            applied_credit.applied_at = None

    # Audit log (before delete)
    _log_payment_audit(db, payment_id=payment_id, table="client", order_id=order_id,
                       action="DELETE", before=before_snap)

    # Global audit trail
    from core.audit import log_audit_event
    log_audit_event(db, current_user, "PAYMENT_DELETE", "payment", payment_id,
                    old_values=before_snap, metadata={"order_id": order_id, "type": "client"})

    db.delete(payment)
    db.commit()

    # Recalculate credit after payment deletion (handles overpayment surplus)
    _recalculate_credit(order_id, db)

    return {"deleted": True}


@router.put("/orders/{order_id}/payments/{payment_id}/")
def update_payment(order_id: str, payment_id: str, data: PaymentUpdate, db: Session = Depends(get_db), current_user: CurrentUser = Depends(get_current_user)):
    """Update a client payment record — admin/finance only."""
    if current_user.role not in ("ADMIN", "SUPER_ADMIN", "FINANCE"):
        raise HTTPException(status_code=403, detail="Only admin/finance can modify payments")
    payment = db.query(Payment).filter(
        Payment.id == payment_id,
        Payment.order_id == order_id,
    ).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")

    # Capture before snapshot for audit
    before_snap = _payment_snapshot(payment)

    # Update only provided fields
    if data.payment_type is not None:
        payment.payment_type = data.payment_type
    if data.amount is not None:
        payment.amount = data.amount
    if data.currency is not None:
        payment.currency = data.currency
    if data.exchange_rate is not None:
        payment.exchange_rate = data.exchange_rate
    if data.method is not None:
        payment.method = data.method
    if data.reference is not None:
        payment.reference = data.reference
    if data.notes is not None:
        payment.notes = data.notes
    if data.payment_date is not None:
        payment.payment_date = date.fromisoformat(data.payment_date)

    # Recompute amount_inr
    payment.amount_inr = round(payment.amount * payment.exchange_rate, 2)
    payment.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(payment)

    # Audit log — compute changed fields
    after_snap = _payment_snapshot(payment)
    changed = [k for k in before_snap if before_snap[k] != after_snap.get(k)]
    if changed:
        _log_payment_audit(db, payment_id=payment_id, table="client", order_id=order_id,
                           action="UPDATE", before=before_snap, after=after_snap, changed=changed)
        db.commit()

    # Recalculate client credit (overpayment tracking)
    _recalculate_credit(order_id, db)

    return PaymentOut(
        id=payment.id,
        order_id=payment.order_id,
        payment_type=payment.payment_type,
        amount=payment.amount,
        currency=payment.currency,
        exchange_rate=payment.exchange_rate,
        amount_inr=payment.amount_inr,
        method=payment.method,
        reference=payment.reference,
        notes=payment.notes,
        payment_date=payment.payment_date.isoformat() if payment.payment_date else "",
        created_at=payment.created_at.isoformat() if payment.created_at else "",
    )


# ========================================
# CLIENT PAYMENT SUBMISSION & VERIFICATION
# ========================================
# These endpoints use a separate router (client_router) registered with
# get_current_user auth only — not require_finance — so CLIENT users can access them.
# Each endpoint handles its own role checks.
client_router = APIRouter()


@client_router.post("/orders/{order_id}/submit-payment/")
@limiter.limit("10/hour")
async def submit_payment(
    request: Request,
    order_id: str,
    payment_type: str = Form(...),
    amount: float = Form(...),
    currency: str = Form("INR"),
    exchange_rate: float = Form(1.0),
    method: str = Form("BANK_TRANSFER"),
    reference: Optional[str] = Form(None),
    payment_date: str = Form(...),
    notes: Optional[str] = Form(None),
    proof_file: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Client submits a payment with proof file. Goes to PENDING_VERIFICATION."""
    if current_user.user_type != "CLIENT":
        raise HTTPException(status_code=403, detail="Only clients can submit payments")

    # Bounds validation
    if amount <= 0 or amount > 50_000_000:
        raise HTTPException(status_code=400, detail="Invalid payment amount (must be between 0 and 5 crore)")
    if exchange_rate <= 0 or exchange_rate > 200:
        raise HTTPException(status_code=400, detail="Invalid exchange rate")

    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if order.client_id != current_user.client_id:
        raise HTTPException(status_code=403, detail="Access denied")

    allowed_types = {"image/jpeg", "image/png", "application/pdf"}
    if proof_file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Only JPG, PNG, or PDF files are allowed")

    proof_dir = UPLOAD_DIR / "payments" / order_id
    proof_dir.mkdir(parents=True, exist_ok=True)
    safe_name = f"{uuid_mod.uuid4().hex[:8]}_{sanitize_filename(proof_file.filename)}"
    proof_path = proof_dir / safe_name
    max_proof_size = 5 * 1024 * 1024  # 5MB
    await stream_upload_to_disk(proof_file, proof_path, max_proof_size)
    relative_path = f"payments/{order_id}/{safe_name}"

    amount_inr = round(amount * exchange_rate, 2)
    payment = Payment(
        order_id=order_id,
        payment_type=payment_type,
        amount=amount,
        currency=currency,
        exchange_rate=exchange_rate,
        amount_inr=amount_inr,
        method=method,
        reference=reference,
        notes=notes,
        payment_date=date.fromisoformat(payment_date),
        proof_file_path=relative_path,
        verification_status=PaymentVerificationStatus.PENDING_VERIFICATION.value,
        submitted_by=current_user.id,
    )
    db.add(payment)

    # Notify admins about pending payment
    from models import Notification, AuditLog
    client_obj = db.query(Client).filter(Client.id == order.client_id).first()
    client_name = client_obj.company_name if client_obj else "Client"
    db.add(Notification(
        user_role="ADMIN",
        client_id=order.client_id,
        title="Payment Verification Required",
        message=f"{client_name} submitted ₹{amount_inr:,.2f} for {order.order_number or order_id}",
        notification_type="PAYMENT_SUBMITTED",
        resource_type="order",
        resource_id=order_id,
    ))

    # Flush to get payment.id before creating audit log
    db.flush()

    # Audit log for activity feed
    db.add(AuditLog(
        action="PAYMENT_SUBMITTED",
        resource_type="payment",
        resource_id=payment.id,
        user_id=current_user.id,
        user_email=current_user.email,
        new_values=json.dumps({
            "amount_inr": amount_inr,
            "method": method,
            "order_id": order_id,
            "payment_type": payment_type,
            "reference": reference,
        }),
        metadata_json=json.dumps({"order_id": order_id}),
    ))

    db.commit()
    db.refresh(payment)

    return {
        "id": payment.id,
        "verification_status": payment.verification_status,
        "amount_inr": payment.amount_inr,
        "message": "Payment submitted for verification",
    }


@router.post("/payments/{payment_id}/verify/")
def verify_payment(
    payment_id: str,
    data: VerifyPaymentRequest,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Admin approves or rejects a client-submitted payment."""
    if current_user.role not in ("ADMIN", "SUPER_ADMIN", "FINANCE"):
        raise HTTPException(status_code=403, detail="Not authorized")

    payment = db.query(Payment).filter(Payment.id == payment_id).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")

    if payment.verification_status != PaymentVerificationStatus.PENDING_VERIFICATION.value:
        raise HTTPException(status_code=400, detail="Payment is not pending verification")

    if data.action == "approve":
        payment.verification_status = PaymentVerificationStatus.VERIFIED.value
        _recalculate_credit(payment.order_id, db)
    elif data.action == "reject":
        if not data.reason:
            raise HTTPException(status_code=400, detail="Rejection reason is required")
        payment.verification_status = PaymentVerificationStatus.REJECTED.value
        payment.rejection_reason = data.reason
    else:
        raise HTTPException(status_code=400, detail="Invalid action. Use 'approve' or 'reject'")

    # Audit log for activity feed
    from models import AuditLog
    audit_action = "PAYMENT_APPROVED" if data.action == "approve" else "PAYMENT_REJECTED"
    db.add(AuditLog(
        action=audit_action,
        resource_type="payment",
        resource_id=payment.id,
        user_id=current_user.id,
        user_email=current_user.email,
        new_values=json.dumps({
            "amount_inr": payment.amount_inr,
            "method": payment.method,
            "order_id": payment.order_id,
            "rejection_reason": data.reason if data.action == "reject" else None,
        }),
        metadata_json=json.dumps({"order_id": payment.order_id}),
    ))

    db.commit()
    return {"status": payment.verification_status, "message": f"Payment {data.action}d"}


@client_router.get("/payments/{payment_id}/proof/")
def download_proof(
    payment_id: str,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Download payment proof file."""
    payment = db.query(Payment).filter(Payment.id == payment_id).first()
    if not payment or not payment.proof_file_path:
        raise HTTPException(status_code=404, detail="Proof not found")

    if current_user.user_type == "CLIENT":
        order = db.query(Order).filter(Order.id == payment.order_id).first()
        if not order or order.client_id != current_user.client_id:
            raise HTTPException(status_code=403, detail="Access denied")

    full_path = UPLOAD_DIR / payment.proof_file_path
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")

    import mimetypes
    mime_type = mimetypes.guess_type(str(full_path))[0] or "application/octet-stream"
    return FileResponse(
        path=str(full_path),
        filename=os.path.basename(payment.proof_file_path),
        media_type=mime_type,
    )


# ========================================
# FACTORY PAYMENTS
# ========================================

@router.get("/orders/{order_id}/factory-payments/")
def list_factory_payments(order_id: str, db: Session = Depends(get_db), current_user: CurrentUser = Depends(get_current_user), _: None = Depends(require_factory_financial)):
    """List all factory payments for an order + total summary."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # RLS: CLIENT users must not see factory payment details
    if current_user.role == "CLIENT":
        raise HTTPException(status_code=403, detail="Access denied")

    # RLS: FACTORY users can only see payments for their own orders
    if current_user.role == "FACTORY":
        if not verify_resource_access(Order, order_id, db, current_user):
            raise HTTPException(status_code=403, detail="Access denied")

    payments = (
        db.query(FactoryPayment)
        .filter(FactoryPayment.order_id == order_id)
        .order_by(FactoryPayment.payment_date.desc(), FactoryPayment.created_at.desc())
        .all()
    )

    # Fetch USD→INR rate for normalization
    usd_rate_obj = db.query(ExchangeRate).filter(ExchangeRate.from_currency == "USD").first()
    usd_inr_rate = usd_rate_obj.rate if usd_rate_obj else 84.0  # fallback

    total_inr = sum(p.amount_inr for p in payments)
    # Group totals by currency
    currency_totals = {}
    for p in payments:
        currency_totals[p.currency] = currency_totals.get(p.currency, 0) + p.amount

    payment_list = []
    total_usd = 0
    for p in payments:
        amount_usd = round(p.amount if p.currency == "USD" else (p.amount_inr / usd_inr_rate if usd_inr_rate else 0), 2)
        total_usd += amount_usd
        payment_list.append(FactoryPaymentOut(
            id=p.id,
            order_id=p.order_id,
            amount=p.amount,
            currency=p.currency,
            exchange_rate=p.exchange_rate,
            amount_inr=p.amount_inr,
            amount_usd=amount_usd,
            method=p.method,
            reference=p.reference,
            notes=p.notes,
            payment_date=p.payment_date.isoformat() if p.payment_date else "",
            created_at=p.created_at.isoformat() if p.created_at else "",
        ))

    avg_exchange_rate_usd = round(total_inr / total_usd, 2) if total_usd > 0 else 0

    # Compute factory order total (bill value) from order items
    active_items = [i for i in order.items if i.status == "ACTIVE"]
    factory_total_cny = sum((i.factory_price or 0) * i.quantity for i in active_items)
    exchange_rate = order.exchange_rate or 1.0
    factory_total_inr = round(factory_total_cny * exchange_rate, 2)
    balance_inr = round(factory_total_inr - total_inr, 2)
    paid_percent = round((total_inr / factory_total_inr) * 100, 1) if factory_total_inr > 0 else 0

    return {
        "payments": payment_list,
        "summary": {
            "factory_total_cny": round(factory_total_cny, 2),
            "factory_total_inr": factory_total_inr,
            "factory_currency": order.currency or "CNY",
            "total_inr": round(total_inr, 2),
            "total_usd": round(total_usd, 2),
            "avg_exchange_rate_usd": avg_exchange_rate_usd,
            "balance_inr": balance_inr,
            "paid_percent": min(paid_percent, 100),
            "currency_totals": {k: round(v, 2) for k, v in currency_totals.items()},
            "remittance_count": len(payments),
        },
    }


@router.post("/orders/{order_id}/factory-payments/")
def create_factory_payment(order_id: str, data: FactoryPaymentCreate, db: Session = Depends(get_db), _: None = Depends(require_factory_financial)):
    """Record a factory payment. Each remittance has its own method + exchange rate."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    amount_inr = round(data.amount * data.exchange_rate, 2)

    payment = FactoryPayment(
        order_id=order_id,
        amount=data.amount,
        currency=data.currency,
        exchange_rate=data.exchange_rate,
        amount_inr=amount_inr,
        method=data.method,
        reference=data.reference,
        notes=data.notes,
        payment_date=date.fromisoformat(data.payment_date),
    )
    db.add(payment)
    db.commit()
    db.refresh(payment)

    # Audit log
    _log_payment_audit(db, payment_id=payment.id, table="factory", order_id=order_id,
                       action="CREATE", after=_payment_snapshot(payment))
    db.commit()

    _recalculate_factory_credit(order_id, db)

    return FactoryPaymentOut(
        id=payment.id,
        order_id=payment.order_id,
        amount=payment.amount,
        currency=payment.currency,
        exchange_rate=payment.exchange_rate,
        amount_inr=payment.amount_inr,
        method=payment.method,
        reference=payment.reference,
        notes=payment.notes,
        payment_date=payment.payment_date.isoformat(),
        created_at=payment.created_at.isoformat(),
    )


@router.delete("/orders/{order_id}/factory-payments/{payment_id}/")
def delete_factory_payment(order_id: str, payment_id: str, db: Session = Depends(get_db), _: None = Depends(require_factory_financial)):
    """Delete a factory payment."""
    payment = db.query(FactoryPayment).filter(
        FactoryPayment.id == payment_id,
        FactoryPayment.order_id == order_id,
    ).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")

    # Capture before snapshot for audit
    before_snap = _payment_snapshot(payment)

    # If deleting a CREDIT-method payment, restore the factory credit
    if payment.method == "CREDIT":
        applied_credit = db.query(FactoryCredit).filter(
            FactoryCredit.applied_to_order_id == order_id,
            FactoryCredit.status == CreditStatus.APPLIED.value,
            FactoryCredit.amount == payment.amount_inr,
        ).first()
        if applied_credit:
            applied_credit.status = CreditStatus.AVAILABLE.value
            applied_credit.applied_to_order_id = None
            applied_credit.applied_at = None

    # Audit log (before delete)
    _log_payment_audit(db, payment_id=payment_id, table="factory", order_id=order_id,
                       action="DELETE", before=before_snap)

    db.delete(payment)
    db.commit()

    _recalculate_factory_credit(order_id, db)

    return {"deleted": True}


@router.put("/orders/{order_id}/factory-payments/{payment_id}/")
def update_factory_payment(order_id: str, payment_id: str, data: FactoryPaymentUpdate, db: Session = Depends(get_db), _: None = Depends(require_factory_financial)):
    """Update a factory payment record."""
    payment = db.query(FactoryPayment).filter(
        FactoryPayment.id == payment_id,
        FactoryPayment.order_id == order_id,
    ).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")

    # Capture before snapshot for audit
    before_snap = _payment_snapshot(payment)

    # Update only provided fields
    if data.amount is not None:
        payment.amount = data.amount
    if data.currency is not None:
        payment.currency = data.currency
    if data.exchange_rate is not None:
        payment.exchange_rate = data.exchange_rate
    if data.method is not None:
        payment.method = data.method
    if data.reference is not None:
        payment.reference = data.reference
    if data.notes is not None:
        payment.notes = data.notes
    if data.payment_date is not None:
        payment.payment_date = date.fromisoformat(data.payment_date)

    # Recompute amount_inr
    payment.amount_inr = round(payment.amount * payment.exchange_rate, 2)
    payment.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(payment)

    # Audit log — compute changed fields
    after_snap = _payment_snapshot(payment)
    changed = [k for k in before_snap if before_snap[k] != after_snap.get(k)]
    if changed:
        _log_payment_audit(db, payment_id=payment_id, table="factory", order_id=order_id,
                           action="UPDATE", before=before_snap, after=after_snap, changed=changed)
        db.commit()

    _recalculate_factory_credit(order_id, db)

    return FactoryPaymentOut(
        id=payment.id,
        order_id=payment.order_id,
        amount=payment.amount,
        currency=payment.currency,
        exchange_rate=payment.exchange_rate,
        amount_inr=payment.amount_inr,
        method=payment.method,
        reference=payment.reference,
        notes=payment.notes,
        payment_date=payment.payment_date.isoformat() if payment.payment_date else "",
        created_at=payment.created_at.isoformat() if payment.created_at else "",
    )


# ========================================
# PAYMENT AUDIT LOG
# ========================================

@router.get("/orders/{order_id}/payment-audit-log/")
def get_payment_audit_log(order_id: str, db: Session = Depends(get_db), _: None = Depends(require_factory_financial)):
    """Return all payment audit log entries for an order, newest first."""
    entries = (
        db.query(PaymentAuditLog)
        .filter(PaymentAuditLog.order_id == order_id)
        .order_by(PaymentAuditLog.created_at.desc())
        .all()
    )
    return [
        {
            "id": e.id,
            "payment_id": e.payment_id,
            "payment_table": e.payment_table,
            "order_id": e.order_id,
            "action": e.action,
            "before_data": json.loads(e.before_data) if e.before_data else None,
            "after_data": json.loads(e.after_data) if e.after_data else None,
            "changed_fields": json.loads(e.changed_fields) if e.changed_fields else None,
            "reason": e.reason,
            "created_at": e.created_at.isoformat() if e.created_at else None,
        }
        for e in entries
    ]


# ========================================
# PI REVISION HISTORY
# ========================================

@router.get("/orders/{order_id}/pi-history/")
def get_pi_history(order_id: str, db: Session = Depends(get_db)):
    """Return PI revision history for an order — all past PI values."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    current_pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order_id).first()
    revisions = db.query(PIRevision).filter(
        PIRevision.order_id == order_id
    ).order_by(PIRevision.revision_number.desc()).all()

    def _serialize_revision(r):
        return {
            "id": r.id,
            "revision_number": r.revision_number,
            "pi_number": r.pi_number,
            "total_cny": r.total_cny,
            "total_inr": r.total_inr,
            "exchange_rate": r.exchange_rate,
            "advance_percent": r.advance_percent,
            "advance_amount_inr": r.advance_amount_inr,
            "item_count": r.item_count,
            "trigger": r.trigger,
            "generated_at": r.generated_at.isoformat() if r.generated_at else None,
        }

    return {
        "current": {
            "pi_number": current_pi.pi_number,
            "total_cny": current_pi.total_cny,
            "total_inr": current_pi.total_inr,
            "exchange_rate": current_pi.exchange_rate,
            "advance_percent": current_pi.advance_percent,
            "advance_amount_inr": current_pi.advance_amount_inr,
            "generated_at": current_pi.generated_at.isoformat() if current_pi.generated_at else None,
            "updated_at": current_pi.updated_at.isoformat() if current_pi.updated_at else None,
        } if current_pi else None,
        "revisions": [_serialize_revision(r) for r in revisions],
        "total_revisions": len(revisions),
    }


# ========================================
# EXCHANGE RATES (convenience endpoint)
# ========================================

@router.get("/exchange-rates/")
def get_exchange_rates(db: Session = Depends(get_db)):
    """Return exchange rates as a map for quick form lookups."""
    rates = db.query(ExchangeRate).all()
    return {r.from_currency: r.rate for r in rates}


# ========================================
# CLIENT CREDIT SCHEMAS
# ========================================

# ========================================
# CLIENT CREDITS
# ========================================

@router.get("/clients/{client_id}/credits/")
def list_client_credits(client_id: str, db: Session = Depends(get_db)):
    """List available credits for a client."""
    credits = db.query(ClientCredit).filter(
        ClientCredit.client_id == client_id,
        ClientCredit.status == CreditStatus.AVAILABLE.value,
    ).order_by(ClientCredit.created_at.desc()).all()

    result = []
    for c in credits:
        source_order = db.query(Order).filter(Order.id == c.source_order_id).first()
        result.append(ClientCreditOut(
            id=c.id,
            client_id=c.client_id,
            source_order_id=c.source_order_id,
            amount=c.amount,
            status=c.status,
            applied_to_order_id=c.applied_to_order_id,
            applied_at=c.applied_at.isoformat() if c.applied_at else None,
            notes=c.notes,
            created_at=c.created_at.isoformat(),
            source_order_number=source_order.order_number if source_order else None,
        ))
    return result


@router.post("/orders/{order_id}/apply-credit/")
def apply_credit(order_id: str, data: ApplyCreditRequest, db: Session = Depends(get_db)):
    """Apply an available client credit as advance payment on an order."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Row-level lock to prevent credit double-spend from concurrent requests
    credit = db.query(ClientCredit).filter(
        ClientCredit.id == data.credit_id,
        ClientCredit.status == CreditStatus.AVAILABLE.value,
    ).with_for_update().first()
    if not credit:
        raise HTTPException(status_code=404, detail="Credit not found or already applied")

    # Verify credit belongs to the same client
    if credit.client_id != order.client_id:
        raise HTTPException(status_code=400, detail="Credit belongs to a different client")

    # Guard: cannot apply a credit sourced from the same order (prevents infinite loop)
    if credit.source_order_id == order_id:
        raise HTTPException(status_code=400, detail="Cannot apply a credit sourced from the same order")

    # Calculate how much is actually needed (order balance)
    # Use effective total from active items (handles after-sales price corrections & unloaded items)
    pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order_id).first()
    stored_pi_total = pi.total_inr if pi else 0
    active_items = [i for i in order.items if i.status == "ACTIVE"]
    recalc_total = round(sum((i.selling_price_inr or 0) * i.quantity for i in active_items), 2)
    pi_total = recalc_total if recalc_total != stored_pi_total else stored_pi_total
    total_paid = db.query(func.coalesce(func.sum(Payment.amount_inr), 0)).filter(
        Payment.order_id == order_id,
        Payment.verification_status == 'VERIFIED',
    ).scalar()
    balance = round(pi_total - total_paid, 2)

    if balance <= 0:
        raise HTTPException(status_code=400, detail="Order is already fully paid")

    # Apply only what's needed: min(credit, balance)
    apply_amount = round(min(credit.amount, balance), 2)

    # Create a payment record from credit
    payment = Payment(
        order_id=order_id,
        payment_type=PaymentType.CLIENT_ADVANCE.value,
        amount=apply_amount,
        currency=Currency.INR.value,
        exchange_rate=1.0,
        amount_inr=apply_amount,
        method="CREDIT",
        reference=f"Credit from {credit.notes or credit.source_order_id}",
        notes=f"Applied credit #{credit.id[:8]}",
        payment_date=date.today(),
    )
    db.add(payment)

    # Update credit: fully consumed or partially consumed
    if apply_amount >= credit.amount:
        # Full credit used
        credit.status = CreditStatus.APPLIED.value
        credit.applied_to_order_id = order_id
        credit.applied_at = datetime.utcnow()
    else:
        # Partial: reduce credit amount, keep AVAILABLE
        credit.amount = round(credit.amount - apply_amount, 2)

    db.commit()
    db.refresh(payment)

    return {
        "payment": PaymentOut(
            id=payment.id,
            order_id=payment.order_id,
            payment_type=payment.payment_type,
            amount=payment.amount,
            currency=payment.currency,
            exchange_rate=payment.exchange_rate,
            amount_inr=payment.amount_inr,
            method=payment.method,
            reference=payment.reference,
            notes=payment.notes,
            payment_date=payment.payment_date.isoformat(),
            created_at=payment.created_at.isoformat(),
        ),
        "credit_applied": True,
    }


# ========================================
# RECEIVABLES
# ========================================

@router.get("/receivables/")
def list_receivables(
    client_id: Optional[str] = None,
    status: str = "outstanding",
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """List all orders with outstanding balances."""
    # Query all orders that have a PI
    query = db.query(Order).join(
        ProformaInvoice, ProformaInvoice.order_id == Order.id
    ).filter(Order.deleted_at.is_(None))

    # RLS: CLIENT users only see their own receivables
    if current_user.role == "CLIENT" and current_user.client_id:
        query = query.filter(Order.client_id == current_user.client_id)
    elif client_id:
        query = query.filter(Order.client_id == client_id)

    orders = query.all()
    results = []

    total_outstanding = 0
    for order in orders:
        pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order.id).first()
        if not pi:
            continue

        total_paid = db.query(func.coalesce(func.sum(Payment.amount_inr), 0)).filter(
            Payment.order_id == order.id,
            Payment.verification_status == 'VERIFIED',
        ).scalar()

        outstanding = round(pi.total_inr - total_paid, 2)

        # Filter by status
        if status == "outstanding" and outstanding <= 0:
            continue
        if status == "settled" and outstanding > 0:
            continue

        # Get client and factory names
        client = db.query(Client).filter(Client.id == order.client_id).first()
        factory = None
        if order.factory_id:
            factory = db.query(Factory).filter(Factory.id == order.factory_id).first()

        # Last payment date
        last_payment = db.query(Payment).filter(
            Payment.order_id == order.id
        ).order_by(Payment.payment_date.desc()).first()

        # Days outstanding (from PI generation date)
        days_outstanding = 0
        if outstanding > 0 and pi.generated_at:
            days_outstanding = (datetime.utcnow() - pi.generated_at).days

        total_outstanding += max(outstanding, 0)

        results.append({
            "order_id": order.id,
            "order_number": order.order_number,
            "client_name": client.company_name if client else "Unknown",
            "client_id": order.client_id,
            "factory_name": factory.company_name if factory else "-",
            "pi_total_inr": round(pi.total_inr, 2),
            "total_paid_inr": round(total_paid, 2),
            "outstanding_inr": outstanding,
            "last_payment_date": last_payment.payment_date.isoformat() if last_payment else None,
            "days_outstanding": days_outstanding,
            "status": order.status,
            "paid_percent": round((total_paid / pi.total_inr * 100) if pi.total_inr > 0 else 0, 1),
        })

    # Sort by days outstanding descending
    results.sort(key=lambda x: x["days_outstanding"], reverse=True)

    return {
        "receivables": results,
        "summary": {
            "total_outstanding_inr": round(total_outstanding, 2),
            "count": len(results),
        }
    }


# ========================================
# CLIENT LEDGER
# ========================================

@router.get("/client-ledger/{client_id}/")
def get_client_ledger(
    client_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get full payment ledger for a client across all orders."""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    # Get all orders for this client
    order_ids = [o.id for o in db.query(Order).filter(
        Order.client_id == client_id,
        Order.deleted_at.is_(None)
    ).all()]

    if not order_ids:
        return {"entries": [], "summary": {"total_debit": 0, "total_credit": 0, "net_balance": 0}, "client_name": client.company_name}

    # Get all client payments for those orders
    # Exclude CREDIT-method payments — they are internal credit reallocations, not actual
    # money received from the client. The original overpayment is already in the ledger.
    query = db.query(Payment).filter(
        Payment.order_id.in_(order_ids),
        Payment.method != "CREDIT",
        Payment.verification_status == 'VERIFIED',
    )

    if start_date:
        query = query.filter(Payment.payment_date >= date.fromisoformat(start_date))
    if end_date:
        query = query.filter(Payment.payment_date <= date.fromisoformat(end_date))

    payments = query.order_by(Payment.payment_date.asc(), Payment.created_at.asc()).all()

    # Build order number lookup
    order_lookup = {}
    for oid in order_ids:
        o = db.query(Order).filter(Order.id == oid).first()
        if o:
            order_lookup[oid] = o.order_number or oid[:8]

    # Build ledger entries
    # For client ledger: PI total is "debit" (what client owes), payments are "credit" (what client paid)
    entries = []
    running_balance = 0

    # First add PI totals as debit entries — use EFFECTIVE total (excludes unloaded items)
    from core.finance_helpers import calc_effective_pi_total, calc_original_pi_total
    pi_entries = []
    for oid in order_ids:
        pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == oid).first()
        if pi:
            pi_date = pi.generated_at.date() if pi.generated_at else date.today()
            # Apply date filter
            if start_date and pi_date < date.fromisoformat(start_date):
                continue
            if end_date and pi_date > date.fromisoformat(end_date):
                continue

            # Use effective (revised) total instead of stale PI.total_inr
            o = db.query(Order).filter(Order.id == oid).first()
            effective_total = calc_effective_pi_total(o, db) if o else round(pi.total_inr, 2)
            original_total = round(pi.total_inr, 2)
            is_revised = abs(effective_total - original_total) > 0.01

            remark = f"PI raised for {order_lookup.get(oid, oid[:8])}"
            if is_revised:
                remark += f" (Revised from ₹{original_total:,.2f})"

            pi_entries.append({
                "date": pi_date.isoformat(),
                "order_number": order_lookup.get(oid, oid[:8]),
                "order_id": oid,
                "remark": remark,
                "debit": effective_total,
                "credit": 0,
                "method": "-",
                "reference": pi.pi_number,
                "sort_key": pi_date,
                "original_debit": original_total if is_revised else None,
            })

    # Payment entries as credits
    payment_entries = []
    for p in payments:
        remark = p.notes or f"{p.payment_type.replace('_', ' ').title()} for {order_lookup.get(p.order_id, p.order_id[:8])}"
        payment_entries.append({
            "date": p.payment_date.isoformat() if p.payment_date else "",
            "order_number": order_lookup.get(p.order_id, p.order_id[:8]),
            "order_id": p.order_id,
            "remark": remark,
            "debit": 0,
            "credit": round(p.amount_inr, 2),
            "method": p.method,
            "reference": p.reference or "-",
            "sort_key": p.payment_date,
        })

    # Merge and sort
    all_entries = pi_entries + payment_entries
    all_entries.sort(key=lambda x: x["sort_key"])

    # Calculate running balance
    total_debit = 0
    total_credit = 0
    for entry in all_entries:
        running_balance += entry["debit"] - entry["credit"]
        entry["running_balance"] = round(running_balance, 2)
        total_debit += entry["debit"]
        total_credit += entry["credit"]
        del entry["sort_key"]

    return {
        "entries": all_entries,
        "summary": {
            "total_debit": round(total_debit, 2),
            "total_credit": round(total_credit, 2),
            "net_balance": round(total_debit - total_credit, 2),
        },
        "client_name": client.company_name,
        "client_id": client_id,
    }


# ========================================
# FACTORY CREDITS
# ========================================

@router.get("/factories/{factory_id}/credits/")
def list_factory_credits(factory_id: str, db: Session = Depends(get_db), _: None = Depends(require_factory_financial)):
    """List available credits for a factory."""
    credits = db.query(FactoryCredit).filter(
        FactoryCredit.factory_id == factory_id,
        FactoryCredit.status == CreditStatus.AVAILABLE.value,
    ).all()

    result = []
    for c in credits:
        order = db.query(Order).filter(Order.id == c.source_order_id).first()
        result.append(FactoryCreditOut(
            id=c.id,
            factory_id=c.factory_id,
            source_order_id=c.source_order_id,
            amount=c.amount,
            status=c.status,
            applied_to_order_id=c.applied_to_order_id,
            applied_at=c.applied_at.isoformat() if c.applied_at else None,
            notes=c.notes,
            created_at=c.created_at.isoformat(),
            source_order_number=order.order_number if order else None,
        ))
    return result


@router.post("/orders/{order_id}/apply-factory-credit/")
def apply_factory_credit(order_id: str, data: ApplyFactoryCreditRequest, db: Session = Depends(get_db), _: None = Depends(require_factory_financial)):
    """Apply an available factory credit as payment on an order."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    credit = db.query(FactoryCredit).filter(
        FactoryCredit.id == data.credit_id,
        FactoryCredit.status == CreditStatus.AVAILABLE.value,
    ).first()
    if not credit:
        raise HTTPException(status_code=404, detail="Credit not found or already applied")

    if credit.factory_id != order.factory_id:
        raise HTTPException(status_code=400, detail="Credit belongs to a different factory")

    if credit.source_order_id == order_id:
        raise HTTPException(status_code=400, detail="Cannot apply a credit sourced from the same order")

    # Compute factory bill balance
    active_items = [i for i in order.items if i.status == "ACTIVE"]
    factory_total_cny = sum((i.factory_price or 0) * i.quantity for i in active_items)
    factory_total_inr = round(factory_total_cny * (order.exchange_rate or 1.0), 2)
    total_paid = db.query(func.coalesce(func.sum(FactoryPayment.amount_inr), 0)).filter(
        FactoryPayment.order_id == order_id
    ).scalar()
    balance = round(factory_total_inr - total_paid, 2)

    if balance <= 0:
        raise HTTPException(status_code=400, detail="Factory bill is already fully paid")

    # Partial: apply only what's needed
    apply_amount = round(min(credit.amount, balance), 2)

    payment = FactoryPayment(
        order_id=order_id,
        amount=apply_amount,
        currency=Currency.INR.value,
        exchange_rate=1.0,
        amount_inr=apply_amount,
        method="CREDIT",
        reference=f"Credit from {credit.notes or credit.source_order_id}",
        notes=f"Applied factory credit #{credit.id[:8]}",
        payment_date=date.today(),
    )
    db.add(payment)

    if apply_amount >= credit.amount:
        credit.status = CreditStatus.APPLIED.value
        credit.applied_to_order_id = order_id
        credit.applied_at = datetime.utcnow()
    else:
        credit.amount = round(credit.amount - apply_amount, 2)

    db.commit()
    db.refresh(payment)

    return {"payment_id": payment.id, "applied_amount": apply_amount, "credit_applied": True}


# ========================================
# FACTORY LEDGER
# ========================================

@router.get("/factory-ledger/{factory_id}/")
def get_factory_ledger(
    factory_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: None = Depends(require_factory_financial),
):
    """Get double-entry ledger for a factory: order values as debits, payments as credits."""
    factory = db.query(Factory).filter(Factory.id == factory_id).first()
    if not factory:
        raise HTTPException(status_code=404, detail="Factory not found")

    from enums import OrderStatus

    # Stages at or past FACTORY_ORDERED -- these orders have a factory commitment
    factory_committed_statuses = [
        OrderStatus.FACTORY_ORDERED.value,
        OrderStatus.PRODUCTION_60.value, OrderStatus.PRODUCTION_80.value,
        OrderStatus.PRODUCTION_90.value, OrderStatus.PRODUCTION_100.value,
        OrderStatus.BOOKED.value, OrderStatus.LOADED.value, OrderStatus.SAILED.value,
        OrderStatus.ARRIVED.value, OrderStatus.CUSTOMS_FILED.value,
        OrderStatus.CLEARED.value, OrderStatus.DELIVERED.value,
        OrderStatus.COMPLETED.value, OrderStatus.COMPLETED_EDITING.value,
    ]

    # Get orders for this factory that have reached FACTORY_ORDERED or later
    orders = db.query(Order).options(
        joinedload(Order.items)
    ).filter(
        Order.factory_id == factory_id,
        Order.deleted_at.is_(None),
        Order.status.in_(factory_committed_statuses),
    ).all()

    if not orders:
        return {
            "entries": [],
            "summary": {"total_debit": 0, "total_credit": 0, "net_balance": 0,
                        "total_debit_usd": 0, "total_credit_usd": 0, "net_balance_usd": 0},
            "factory_name": factory.company_name,
            "factory_id": factory_id,
        }

    # Fetch USD→INR rate for normalization
    usd_rate_obj = db.query(ExchangeRate).filter(ExchangeRate.from_currency == "USD").first()
    usd_inr_rate = usd_rate_obj.rate if usd_rate_obj else 84.0  # fallback

    order_ids = [o.id for o in orders]

    # Build all entries: debits (order values) + credits (payments)
    raw_entries = []

    for o in orders:
        active_items = [i for i in o.items if i.status == "ACTIVE"]
        total_cny = sum((i.factory_price or 0) * i.quantity for i in active_items)
        exchange_rate = o.exchange_rate or 1.0
        total_inr = round(total_cny * exchange_rate, 2)
        order_number = o.order_number or o.id[:8]

        # Debit entry: factory order value
        raw_entries.append({
            "date": o.created_at.strftime("%Y-%m-%d"),
            "order_number": order_number,
            "order_id": o.id,
            "remark": f"Factory order for {order_number}",
            "debit": total_inr,
            "credit": 0,
            "amount_foreign": round(total_cny, 2),
            "currency": o.currency or "CNY",
            "exchange_rate": exchange_rate,
            "amount_usd": round(total_inr / usd_inr_rate, 2) if usd_inr_rate else 0,
            "method": "-",
            "reference": order_number,
            "sort_key": (o.created_at.strftime("%Y-%m-%d"), o.created_at.isoformat(), "0"),
        })

    # Get all factory payments for these orders
    # Exclude CREDIT-method payments — they are internal credit reallocations, not actual
    # money sent to the factory. The original overpayment is already in the ledger.
    query = db.query(FactoryPayment).filter(
        FactoryPayment.order_id.in_(order_ids),
        FactoryPayment.method != "CREDIT",
    )
    if start_date:
        query = query.filter(FactoryPayment.payment_date >= date.fromisoformat(start_date))
    if end_date:
        query = query.filter(FactoryPayment.payment_date <= date.fromisoformat(end_date))

    payments = query.all()

    order_lookup = {o.id: (o.order_number or o.id[:8]) for o in orders}

    for p in payments:
        order_num = order_lookup.get(p.order_id, p.order_id[:8])
        raw_entries.append({
            "date": p.payment_date.isoformat() if p.payment_date else "",
            "order_number": order_num,
            "order_id": p.order_id,
            "remark": p.notes or f"Payment for {order_num}",
            "debit": 0,
            "credit": round(p.amount_inr, 2),
            "amount_foreign": round(p.amount, 2),
            "currency": p.currency,
            "exchange_rate": p.exchange_rate,
            "amount_usd": round(p.amount if p.currency == "USD" else (p.amount_inr / usd_inr_rate), 2) if usd_inr_rate else 0,
            "method": p.method,
            "reference": p.reference or "-",
            "sort_key": (p.payment_date.isoformat() if p.payment_date else "", p.created_at.isoformat(), "1"),
        })

    # Sort by date, then by created_at, then debits before credits
    raw_entries.sort(key=lambda e: e["sort_key"])

    # Compute running balance
    total_debit = 0
    total_credit = 0
    entries = []
    for e in raw_entries:
        total_debit += e["debit"]
        total_credit += e["credit"]
        running_balance = round(total_debit - total_credit, 2)
        entry = {k: v for k, v in e.items() if k != "sort_key"}
        entry["running_balance"] = running_balance
        entries.append(entry)

    return {
        "entries": entries,
        "summary": {
            "total_debit": round(total_debit, 2),
            "total_credit": round(total_credit, 2),
            "net_balance": round(total_debit - total_credit, 2),
            "total_debit_usd": round(total_debit / usd_inr_rate, 2) if usd_inr_rate else 0,
            "total_credit_usd": round(total_credit / usd_inr_rate, 2) if usd_inr_rate else 0,
            "net_balance_usd": round((total_debit - total_credit) / usd_inr_rate, 2) if usd_inr_rate else 0,
        },
        "factory_name": factory.company_name,
        "factory_id": factory_id,
    }


# ========================================
# CLIENT LEDGER DOWNLOAD (Excel / PDF)
# ========================================

@router.get("/client-ledger/{client_id}/download/")
def download_client_ledger(
    client_id: str,
    format: str = "xlsx",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Download client ledger as Excel or PDF."""
    # Reuse the existing ledger logic
    ledger_data = get_client_ledger(client_id, start_date, end_date, db)
    entries = ledger_data["entries"]
    summary = ledger_data["summary"]
    client_name = ledger_data["client_name"]

    if format == "xlsx":
        return _generate_client_ledger_excel(entries, summary, client_name, start_date, end_date)
    elif format == "pdf":
        return _generate_client_ledger_pdf(entries, summary, client_name, start_date, end_date)
    else:
        raise HTTPException(status_code=400, detail="Format must be xlsx or pdf")


def _generate_client_ledger_excel(entries, summary, client_name, start_date, end_date):
    """Generate Excel file for client ledger."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Ledger"

    # Styles
    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=11)
    col_header_font = Font(bold=True, size=10, color="FFFFFF")
    col_header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    summary_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = "HarvestERP - Client Ledger Statement"
    ws['A1'].font = header_font

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Client: {client_name}"
    ws['A2'].font = sub_header_font

    date_range = ""
    if start_date and end_date:
        date_range = f"Period: {start_date} to {end_date}"
    elif start_date:
        date_range = f"From: {start_date}"
    elif end_date:
        date_range = f"Until: {end_date}"
    else:
        date_range = "All Time"
    ws.merge_cells('A3:H3')
    ws['A3'] = date_range

    # Column headers
    headers = ["Date", "Order #", "Remark", "Debit (INR)", "Credit (INR)", "Balance (INR)", "Method", "Reference"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for i, entry in enumerate(entries, 6):
        ws.cell(row=i, column=1, value=entry["date"]).border = thin_border
        ws.cell(row=i, column=2, value=entry["order_number"]).border = thin_border
        ws.cell(row=i, column=3, value=entry["remark"]).border = thin_border
        cell_d = ws.cell(row=i, column=4, value=entry["debit"] if entry["debit"] > 0 else "")
        cell_d.border = thin_border
        cell_d.number_format = '#,##0.00'
        cell_c = ws.cell(row=i, column=5, value=entry["credit"] if entry["credit"] > 0 else "")
        cell_c.border = thin_border
        cell_c.number_format = '#,##0.00'
        cell_b = ws.cell(row=i, column=6, value=entry["running_balance"])
        cell_b.border = thin_border
        cell_b.number_format = '#,##0.00'
        ws.cell(row=i, column=7, value=entry["method"]).border = thin_border
        ws.cell(row=i, column=8, value=entry["reference"]).border = thin_border

    # Summary row
    summary_row = 6 + len(entries)
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(bold=True)
    ws.cell(row=summary_row, column=1).fill = summary_fill
    cell_td = ws.cell(row=summary_row, column=4, value=summary["total_debit"])
    cell_td.font = Font(bold=True)
    cell_td.fill = summary_fill
    cell_td.number_format = '#,##0.00'
    cell_tc = ws.cell(row=summary_row, column=5, value=summary["total_credit"])
    cell_tc.font = Font(bold=True)
    cell_tc.fill = summary_fill
    cell_tc.number_format = '#,##0.00'
    cell_nb = ws.cell(row=summary_row, column=6, value=summary["net_balance"])
    cell_nb.font = Font(bold=True)
    cell_nb.fill = summary_fill
    cell_nb.number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"client_ledger_{client_name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _generate_client_ledger_pdf(entries, summary, client_name, start_date, end_date):
    """Generate PDF for client ledger."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=6)
    elements.append(Paragraph("HarvestERP - Client Ledger Statement", title_style))

    # Client info
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, spaceAfter=3)
    elements.append(Paragraph(f"<b>Client:</b> {client_name}", info_style))
    date_range = "All Time"
    if start_date and end_date:
        date_range = f"{start_date} to {end_date}"
    elif start_date:
        date_range = f"From {start_date}"
    elif end_date:
        date_range = f"Until {end_date}"
    elements.append(Paragraph(f"<b>Period:</b> {date_range}", info_style))
    elements.append(Spacer(1, 10))

    # Table data
    table_data = [["Date", "Order #", "Remark", "Debit (INR)", "Credit (INR)", "Balance (INR)", "Method", "Reference"]]
    for entry in entries:
        table_data.append([
            entry["date"],
            entry["order_number"],
            entry["remark"][:40],  # Truncate for PDF
            f'{entry["debit"]:,.2f}' if entry["debit"] > 0 else "",
            f'{entry["credit"]:,.2f}' if entry["credit"] > 0 else "",
            f'{entry["running_balance"]:,.2f}',
            entry["method"],
            entry["reference"][:15],
        ])

    # Summary row
    table_data.append([
        "", "", "TOTALS",
        f'{summary["total_debit"]:,.2f}',
        f'{summary["total_credit"]:,.2f}',
        f'{summary["net_balance"]:,.2f}',
        "", ""
    ])

    col_widths = [60, 80, 150, 70, 70, 70, 70, 80]
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (3, 0), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8FAFC')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F0FDF4')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"client_ledger_{client_name.replace(' ', '_')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ========================================
# FACTORY LEDGER DOWNLOAD (Excel / PDF)
# ========================================

@router.get("/factory-ledger/{factory_id}/download/")
def download_factory_ledger(
    factory_id: str,
    format: str = "xlsx",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: None = Depends(require_factory_financial),
):
    """Download factory ledger as Excel or PDF."""
    ledger_data = get_factory_ledger(factory_id, start_date, end_date, db)
    entries = ledger_data["entries"]
    summary = ledger_data["summary"]
    factory_name = ledger_data["factory_name"]

    if format == "xlsx":
        return _generate_factory_ledger_excel(entries, summary, factory_name, start_date, end_date)
    elif format == "pdf":
        return _generate_factory_ledger_pdf(entries, summary, factory_name, start_date, end_date)
    else:
        raise HTTPException(status_code=400, detail="Format must be xlsx or pdf")


def _generate_factory_ledger_excel(entries, summary, factory_name, start_date, end_date):
    """Generate Excel file for factory ledger (double-entry format with currency & USD columns)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Factory Ledger"

    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=11)
    col_header_font = Font(bold=True, size=10, color="FFFFFF")
    col_header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    summary_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "HarvestERP - Factory Ledger Statement"
    ws['A1'].font = header_font

    ws.merge_cells('A2:K2')
    ws['A2'] = f"Factory: {factory_name}"
    ws['A2'].font = sub_header_font

    date_range = "All Time"
    if start_date and end_date:
        date_range = f"Period: {start_date} to {end_date}"
    elif start_date:
        date_range = f"From: {start_date}"
    elif end_date:
        date_range = f"Until: {end_date}"
    ws.merge_cells('A3:K3')
    ws['A3'] = date_range

    # Column headers (double-entry layout with currency & USD)
    headers = ["Date", "Order #", "Remark", "Currency", "Rate", "Debit (INR)", "Credit (INR)",
               "USD Amount", "Balance (INR)", "Method", "Reference"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for i, entry in enumerate(entries, 6):
        ws.cell(row=i, column=1, value=entry["date"]).border = thin_border
        ws.cell(row=i, column=2, value=entry["order_number"]).border = thin_border
        ws.cell(row=i, column=3, value=entry["remark"]).border = thin_border
        ws.cell(row=i, column=4, value=entry.get("currency", "")).border = thin_border
        cell_r = ws.cell(row=i, column=5, value=entry.get("exchange_rate", ""))
        cell_r.border = thin_border
        cell_r.number_format = '#,##0.00'
        cell_d = ws.cell(row=i, column=6, value=entry["debit"] if entry["debit"] > 0 else "")
        cell_d.border = thin_border
        cell_d.number_format = '#,##0.00'
        cell_c = ws.cell(row=i, column=7, value=entry["credit"] if entry["credit"] > 0 else "")
        cell_c.border = thin_border
        cell_c.number_format = '#,##0.00'
        cell_u = ws.cell(row=i, column=8, value=entry.get("amount_usd", ""))
        cell_u.border = thin_border
        cell_u.number_format = '#,##0.00'
        cell_b = ws.cell(row=i, column=9, value=entry["running_balance"])
        cell_b.border = thin_border
        cell_b.number_format = '#,##0.00'
        ws.cell(row=i, column=10, value=entry["method"]).border = thin_border
        ws.cell(row=i, column=11, value=entry["reference"]).border = thin_border

    # Summary row
    sr = 6 + len(entries)
    ws.merge_cells(f'A{sr}:E{sr}')
    ws.cell(row=sr, column=1, value="TOTALS").font = Font(bold=True)
    ws.cell(row=sr, column=1).fill = summary_fill
    cell_td = ws.cell(row=sr, column=6, value=summary["total_debit"])
    cell_td.font = Font(bold=True)
    cell_td.fill = summary_fill
    cell_td.number_format = '#,##0.00'
    cell_tc = ws.cell(row=sr, column=7, value=summary["total_credit"])
    cell_tc.font = Font(bold=True)
    cell_tc.fill = summary_fill
    cell_tc.number_format = '#,##0.00'
    cell_nu = ws.cell(row=sr, column=8, value=summary.get("net_balance_usd", ""))
    cell_nu.font = Font(bold=True)
    cell_nu.fill = summary_fill
    cell_nu.number_format = '#,##0.00'
    cell_nb = ws.cell(row=sr, column=9, value=summary["net_balance"])
    cell_nb.font = Font(bold=True)
    cell_nb.fill = summary_fill
    cell_nb.number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"factory_ledger_{factory_name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _generate_factory_ledger_pdf(entries, summary, factory_name, start_date, end_date):
    """Generate PDF for factory ledger (double-entry format with currency & USD columns)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=6)
    elements.append(Paragraph("HarvestERP - Factory Ledger Statement", title_style))

    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, spaceAfter=3)
    elements.append(Paragraph(f"<b>Factory:</b> {factory_name}", info_style))
    date_range = "All Time"
    if start_date and end_date:
        date_range = f"{start_date} to {end_date}"
    elif start_date:
        date_range = f"From {start_date}"
    elif end_date:
        date_range = f"Until {end_date}"
    elements.append(Paragraph(f"<b>Period:</b> {date_range}", info_style))
    elements.append(Spacer(1, 10))

    # Double-entry columns with currency & USD
    table_data = [["Date", "Order #", "Remark", "Currency", "Rate", "Debit (INR)", "Credit (INR)",
                   "USD Amt", "Balance (INR)", "Method", "Reference"]]
    for entry in entries:
        table_data.append([
            entry["date"],
            entry["order_number"],
            entry["remark"][:35],
            entry.get("currency", ""),
            f'{entry.get("exchange_rate", 0):,.2f}' if entry.get("exchange_rate") else "",
            f'{entry["debit"]:,.2f}' if entry["debit"] > 0 else "",
            f'{entry["credit"]:,.2f}' if entry["credit"] > 0 else "",
            f'{entry.get("amount_usd", 0):,.2f}' if entry.get("amount_usd") else "",
            f'{entry["running_balance"]:,.2f}',
            entry["method"],
            entry["reference"][:12],
        ])

    # Summary row
    table_data.append([
        "", "", "TOTALS", "", "",
        f'{summary["total_debit"]:,.2f}',
        f'{summary["total_credit"]:,.2f}',
        f'{summary.get("net_balance_usd", 0):,.2f}',
        f'{summary["net_balance"]:,.2f}',
        "", ""
    ])

    col_widths = [52, 65, 120, 42, 42, 62, 62, 55, 62, 55, 65]
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (3, 0), (8, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8FAFC')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EFF6FF')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"factory_ledger_{factory_name.replace(' ', '_')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
