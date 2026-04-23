"""
Order CRUD + Status Engine API endpoints (Level 2A + 2B)
The core entity — everything else connects to orders.
"""
from math import ceil
from typing import Optional, List
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, func, desc

import os
import shutil
import uuid
import openpyxl

from database import get_db
import json as json_mod

from models import Order, OrderItem, Product, Client, Factory, ExchangeRate, ProductImage, ProformaInvoice, Payment, FactoryPayment, PackingList, PackingListItem, UnloadedItem, StageOverride, ProductRequest, Notification
from enums import OrderStatus, Currency, OrderItemStatus, UnloadedItemStatus, STAGE_MAP, get_stage_info
from config import UPLOAD_DIR, MAX_UPLOAD_SIZE
from core.file_upload import stream_upload_to_disk
from core.security import CurrentUser, get_current_user, require_operations, get_scoped_query, verify_resource_access
from core.serializers import filter_for_role, filter_list_for_role
from decimal import Decimal, ROUND_HALF_UP
from services import stage_engine
import config as _config
from schemas.orders import (
    OrderCreate, OrderUpdate, OrderResponse,
    OrderItemResponse, OrderListResponse, OrderListItem,
    OrderAddItems, OrderUpdateItem, OrderRemoveItem,
    MigrateItemRequest, MigrateItemsRequest,
    BulkTextAddRequest, BulkTextApplyRequest,
    UndoMigrateRequest, ItemPriceUpdate,
    StageTransitionRequest, OrderReopenRequest,
    GoBackRequest, JumpToStageRequest,
    UpdatePackingItemRequest, SplitPackingItemRequest,
    ShippingDecisionRequest, ManualPackingItem,
    ManualPackingListRequest, ClientOrderItemCreate,
    QuickAddItemCreate, ClientOrderCreate,
    ApproveInquiryRequest, DeleteOrderRequest,
    ProductionDatesIn, OrderItemCreate,
)

router = APIRouter()


def _calc_client_factory_price(item: OrderItem, client: Client) -> None:
    """Auto-calculate client_factory_price for transparency clients.

    Sets item.client_factory_price using Decimal math.
    No-op for REGULAR clients or when feature flag is off.
    """
    if not _config.TRANSPARENCY_ENABLED:
        return
    if not client or client.client_type != "TRANSPARENCY":
        item.client_factory_price = None
        return
    if client.factory_markup_percent is None or item.factory_price is None:
        return
    markup = Decimal(str(client.factory_markup_percent))
    factory = Decimal(str(item.factory_price))
    item.client_factory_price = float(
        (factory * (1 + markup / 100)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    )


def _copy_image_to_order(order_id: str, source_rel_path: str) -> str | None:
    """Copy a product image into uploads/orders/{order_id}/ for permanent snapshot.
    Returns the new relative path (e.g. 'orders/{order_id}/img_xxx.jpg') or None on failure."""
    src = UPLOAD_DIR / source_rel_path
    if not src.exists():
        return None
    dest_dir = UPLOAD_DIR / "orders" / order_id
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / src.name
    # Avoid overwriting if same filename already exists (e.g. re-snapshot)
    if not dest.exists():
        shutil.copy2(str(src), str(dest))
    return f"orders/{order_id}/{src.name}"



# ========================================
# Stages where items can be added/updated/removed (up to Factory Ordered + Completed Editing)
ITEM_EDITABLE_STAGES = [
    OrderStatus.CLIENT_DRAFT.value,
    OrderStatus.DRAFT.value,
    OrderStatus.PENDING_PI.value,
    OrderStatus.PI_SENT.value,
    OrderStatus.ADVANCE_PENDING.value,
    OrderStatus.ADVANCE_RECEIVED.value,
    OrderStatus.FACTORY_ORDERED.value,
    OrderStatus.COMPLETED_EDITING.value,
]

# Stages where items can be ADDED (but not edited/removed) — mid-production
MID_ORDER_ADD_STAGES = [
    OrderStatus.FACTORY_ORDERED.value,
    OrderStatus.PRODUCTION_60.value,
    OrderStatus.PRODUCTION_80.value,
    OrderStatus.PRODUCTION_90.value,
]


# Helper: restore original UnloadedItem to PENDING when a carried item is removed
def _restore_carried_item_source(item: OrderItem, db: Session) -> None:
    """If this item was carried from a previous order, find the original UnloadedItem
    (now ADDED_TO_ORDER) and restore it to PENDING so it can be fetched again."""
    import re
    notes = item.notes or ""
    # Match "Carried from ORD-xxx" or "After-Sales (xxx) from ORD-xxx"
    carried_match = re.match(r'^Carried from (ORD-\S+)', notes)
    aftersales_match = re.match(r'^After-Sales \(.+?\) from (ORD-\S+)', notes)
    if not carried_match and not aftersales_match:
        return
    # Find the original UnloadedItem that was ADDED_TO_ORDER for this product
    source_ui = db.query(UnloadedItem).filter(
        UnloadedItem.product_id == item.product_id,
        UnloadedItem.status == UnloadedItemStatus.ADDED_TO_ORDER.value,
    ).first()
    if source_ui:
        source_ui.status = UnloadedItemStatus.PENDING.value


# Helper: generate order number  ORD-YYYYMM-NNN
def generate_order_number(db: Session) -> str:
    now = datetime.utcnow()
    prefix = f"ORD-{now.strftime('%Y%m')}-"

    # Count orders in this month
    count = db.query(func.count(Order.id)).filter(
        Order.order_number.like(f"{prefix}%")
    ).scalar()

    return f"{prefix}{(count + 1):03d}"


def _client_prefix(client_name: str) -> str:
    """Generate 3-letter prefix from client name.
    3+ words: first letter of each word (max 3) → 'Shree Supreme Performa' → SSP
    2 words: first letter of each → 'John Doe' → JD + first of first → JOD
    1 word: first 3 letters → 'Sachin' → SAC
    """
    if not client_name:
        return "CLI"
    words = [w for w in client_name.strip().split() if w]
    if len(words) >= 3:
        return (words[0][0] + words[1][0] + words[2][0]).upper()
    elif len(words) == 2:
        return (words[0][0] + words[1][0:2]).upper()[:3]
    else:
        return words[0][:3].upper()


def _indian_financial_year() -> str:
    """Get Indian FY string: Apr-Mar cycle. March 2026 → '25-26', April 2026 → '26-27'"""
    now = datetime.utcnow()
    year = now.year
    month = now.month
    if month >= 4:  # April onwards = start of new FY
        return f"{year % 100}-{(year + 1) % 100}"
    else:  # Jan-Mar = tail of previous FY
        return f"{(year - 1) % 100}-{year % 100}"


def generate_po_reference(db: Session, client_name: str) -> str:
    """Generate unique PO reference in Indian FY format.

    Format: SSP/25-26/03/0001
      SSP   = 3-letter client prefix
      25-26 = Indian Financial Year (Apr–Mar)
      03    = Month number
      0001  = Yearly sequential per client prefix

    Example: SSP/25-26/03/0001, SSP/25-26/03/0002, SSP/25-26/04/0003
    """
    prefix = _client_prefix(client_name)
    fy = _indian_financial_year()
    month = datetime.utcnow().strftime("%m")

    # Yearly sequential: count all POs for this client+FY
    fy_pattern = f"{prefix}/{fy}/%"
    count = db.query(func.count(Order.id)).filter(
        Order.po_reference.like(fy_pattern)
    ).scalar()

    return f"{prefix}/{fy}/{month}/{(count + 1):04d}"


# Helper: serialize order
def serialize_order(order: Order, db: Session) -> dict:
    stage_number, stage_name = get_stage_info(order.status)

    # Get active confirmed items count and total value (exclude PENDING/REJECTED)
    active_items = [i for i in order.items if i.status == "ACTIVE"]
    confirmed_items = [i for i in active_items if not i.pi_item_status or i.pi_item_status == "APPROVED"]
    item_count = len(confirmed_items)
    total_cny = sum((i.factory_price or 0) * i.quantity for i in confirmed_items)

    # PI staleness check
    pi_stale = False
    if order.items_modified_at:
        pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order.id).first()
        if pi and order.items_modified_at > pi.generated_at:
            pi_stale = True

    result = OrderResponse(
        id=order.id,
        order_number=order.order_number,
        client_id=order.client_id,
        factory_id=order.factory_id,
        status=order.status,
        currency=order.currency,
        exchange_rate=order.exchange_rate,
        exchange_rate_date=order.exchange_rate_date.isoformat() if order.exchange_rate_date else None,
        po_reference=order.po_reference,
        notes=order.notes,
        reopen_count=order.reopen_count,
        last_reopen_reason=order.last_reopen_reason,
        igst_credit_amount=order.igst_credit_amount,
        igst_credit_claimed=order.igst_credit_claimed,
        completed_at=order.completed_at.isoformat() if order.completed_at else None,
        created_at=order.created_at.isoformat() if order.created_at else None,
        updated_at=order.updated_at.isoformat() if order.updated_at else None,
        client_name=order.client.company_name if order.client else None,
        factory_name=order.factory.company_name if order.factory else None,
        item_count=item_count,
        total_value_cny=round(total_cny, 2),
        stage_number=stage_number,
        stage_name=stage_name,
        highest_unlocked_stage=order.highest_unlocked_stage,
        pi_stale=pi_stale,
        version=order.version or 1,
        client_reference=order.client_reference,
        deletion_reason=order.deletion_reason,
        deleted_by=order.deleted_by,
        deleted_at=order.deleted_at.isoformat() if order.deleted_at else None,
    ).model_dump()
    result["client_type"] = order.client.client_type if order.client else "REGULAR"
    return result


def serialize_order_item(item: OrderItem) -> dict:
    # Snapshot wins once populated (order is a frozen "xerox copy").
    # Live product data is only used as fallback when snapshots are empty (DRAFT stage).
    product = item.product

    # Build carry-forward label for client-safe display
    # (notes is stripped by role filter, so we need a separate field)
    carry_forward_label = None
    if item.notes:
        import re
        ul_match = re.match(r'^Carried from (ORD-\S+|previous order)', item.notes)
        as_match = re.match(r'^After-Sales \((.+?)\) from (ORD-\S+)', item.notes)
        if ul_match:
            carry_forward_label = f"Carried from {ul_match.group(1)}"
        elif as_match:
            resolution = as_match.group(1).lower()
            if 'replace' in resolution:
                carry_forward_label = f"After-Sales \u2022 Replace from {as_match.group(2)}"
            elif 'compensat' in resolution or 'reduct' in resolution or 'balance' in resolution:
                carry_forward_label = f"After-Sales \u2022 Reduct Balance from {as_match.group(2)}"
            else:
                carry_forward_label = f"After-Sales from {as_match.group(2)}"

    result = OrderItemResponse(
        id=item.id,
        order_id=item.order_id,
        product_id=item.product_id,
        quantity=item.quantity,
        factory_price=item.factory_price,
        client_factory_price=float(item.client_factory_price) if item.client_factory_price is not None else None,
        markup_percent=item.markup_percent,
        selling_price=item.selling_price,
        selling_price_inr=item.selling_price_inr,
        factory_image_path=item.image_path_snapshot or item.factory_image_path,
        status=item.status,
        pi_item_status=item.pi_item_status,
        pi_addition_lot=item.pi_addition_lot,
        notes=item.notes,
        cancel_note=item.cancel_note,
        product_code=item.product_code_snapshot or (product.product_code if product else None),
        product_name=item.product_name_snapshot or (product.product_name if product else None),
        material=item.material_snapshot if item.material_snapshot else (product.material if product else None),
        category=item.category_snapshot if item.category_snapshot else (product.category if product else None),
        part_type=item.part_type_snapshot if item.part_type_snapshot else (product.part_type if product else None),
        dimension=item.dimension_snapshot if item.dimension_snapshot else (product.dimension if product else None),
        variant_note=item.variant_note_snapshot if item.variant_note_snapshot else (product.variant_note if product else None),
    ).model_dump()
    result["carry_forward_label"] = carry_forward_label
    # Product image fallback: use ProductImage if no snapshot/factory image
    if not result.get("factory_image_path") and item.product_id:
        from models import ProductImage as _PI
        from database import SessionLocal as _SL
        _db = _SL()
        try:
            _img = _db.query(_PI).filter(_PI.product_id == item.product_id).first()
            if _img:
                result["factory_image_path"] = _img.image_path
        finally:
            _db.close()
    return result


# ========================================
# 2A: ORDER CRUD
# ========================================

@router.get("/")
def list_orders(
    search: Optional[str] = None,
    status: Optional[str] = None,
    client_id: Optional[str] = None,
    include_deleted: bool = False,
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """List orders with filters and pagination"""
    query = get_scoped_query(Order, db, current_user).options(
        joinedload(Order.client),
        joinedload(Order.factory),
        joinedload(Order.items).joinedload(OrderItem.product),
    )

    if not include_deleted:
        query = query.filter(Order.deleted_at.is_(None))

    if search:
        query = query.filter(
            or_(
                Order.order_number.ilike(f"%{search}%"),
                Order.po_reference.ilike(f"%{search}%"),
            )
        )

    if status:
        # Support comma-separated statuses
        statuses = [s.strip() for s in status.split(",")]
        query = query.filter(Order.status.in_(statuses))

    if client_id:
        query = query.filter(Order.client_id == client_id)

    total = query.count()
    orders = query.order_by(desc(Order.created_at)).offset(
        (page - 1) * per_page
    ).limit(per_page).all()

    return {
        "items": filter_list_for_role(
            [serialize_order(o, db) for o in orders],
            current_user.role
        ),
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": ceil(total / per_page) if per_page else 0,
    }


@router.get("/status-counts/")
def get_status_counts(db: Session = Depends(get_db)):
    """Get count of orders per status (for filter sidebar)"""
    counts = db.query(
        Order.status, func.count(Order.id)
    ).filter(Order.deleted_at.is_(None)).group_by(Order.status).all()

    result = {}
    for status, count in counts:
        stage_num, stage_name = get_stage_info(status)
        result[status] = {"count": count, "stage": stage_num, "name": stage_name}

    return result


# ── Client Ledger (MUST be before /{order_id}/ to avoid route collision) ────────

@router.get("/my-ledger/")
def client_ledger(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Client-facing statement of account: payments made vs order totals."""
    if not current_user.client_id:
        raise HTTPException(status_code=403, detail="Not a client user")

    client = db.query(Client).filter(Client.id == current_user.client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    perms = client.portal_permissions or {}
    if not perms.get("show_payments", False):
        raise HTTPException(status_code=403, detail="Payment access not enabled for your account")

    # Statuses where order is visible in ledger (PI_SENT+)
    _LEDGER_STATUSES = {
        OrderStatus.PI_SENT.value, OrderStatus.ADVANCE_PENDING.value,
        OrderStatus.ADVANCE_RECEIVED.value, OrderStatus.FACTORY_ORDERED.value,
        OrderStatus.PRODUCTION_60.value, OrderStatus.PRODUCTION_80.value,
        OrderStatus.PRODUCTION_90.value, OrderStatus.PLAN_PACKING.value,
        OrderStatus.FINAL_PI.value, OrderStatus.PRODUCTION_100.value,
        OrderStatus.BOOKED.value, OrderStatus.LOADED.value,
        OrderStatus.SAILED.value, OrderStatus.ARRIVED.value,
        OrderStatus.CUSTOMS_FILED.value, OrderStatus.CLEARED.value,
        OrderStatus.DELIVERED.value, OrderStatus.AFTER_SALES.value,
        OrderStatus.COMPLETED.value,
    }

    orders = db.query(Order).filter(
        Order.client_id == current_user.client_id,
        Order.deleted_at.is_(None),
        Order.status.in_(_LEDGER_STATUSES),
    ).order_by(desc(Order.created_at)).all()

    order_ids = [o.id for o in orders]
    all_payments = db.query(Payment).filter(
        Payment.order_id.in_(order_ids),
    ).order_by(desc(Payment.payment_date)).all() if order_ids else []

    # Only VERIFIED non-CREDIT payments shown in client ledger
    client_payments = [
        p for p in all_payments
        if (p.method or '').upper() != 'CREDIT'
        and getattr(p, 'verification_status', 'VERIFIED') != 'REJECTED'
    ]

    from core.finance_helpers import calc_order_finance_summary

    order_lines = []
    total_estimated = 0.0
    total_final = 0.0
    for o in orders:
        fin = calc_order_finance_summary(o, db)
        order_total = fin["effective_pi_total"]

        is_final = o.status == OrderStatus.COMPLETED.value
        if is_final:
            total_final += order_total
        else:
            total_estimated += order_total

        order_lines.append({
            "id": o.id, "order_number": o.order_number, "status": o.status,
            "is_final": is_final,
            "total_inr": order_total,
            "original_pi_total": fin["original_pi_total"],
            "is_revised": fin["is_revised"],
            "paid": fin["total_paid"],
            "balance": fin["balance"],
            "item_count": len([i for i in o.items if i.status == "ACTIVE" and (not i.pi_item_status or i.pi_item_status == "APPROVED")]),
            "created_at": o.created_at.isoformat() if o.created_at else None,
        })

    payment_lines = []
    total_paid = 0.0
    for p in client_payments:
        # Only VERIFIED payments count toward totals
        if getattr(p, 'verification_status', 'VERIFIED') == 'VERIFIED':
            total_paid += p.amount_inr or 0
        payment_lines.append({
            "id": p.id, "order_id": p.order_id,
            "payment_type": p.payment_type, "method": p.method,
            "reference": p.reference,
            "amount_inr": round(p.amount_inr or 0, 2),
            "payment_date": p.payment_date.isoformat() if p.payment_date else None,
            "verification_status": getattr(p, 'verification_status', 'VERIFIED'),
            "proof_file_path": getattr(p, 'proof_file_path', None),
            "rejection_reason": getattr(p, 'rejection_reason', None),
        })

    return {
        "client_name": client.company_name,
        "summary": {
            "total_paid": round(total_paid, 2),
            "total_estimated": round(total_estimated, 2),
            "total_final": round(total_final, 2),
            "total_orders": round(total_estimated + total_final, 2),
            "net_position": round(total_paid - (total_estimated + total_final), 2),
        },
        "orders": order_lines,
        "payments": payment_lines,
    }


# ── Order Reconciliation / Final Draft ─────────────────────────────────────────

@router.get("/reconciliation/{order_id}/")
def get_order_reconciliation(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Full reconciliation summary for a completed order.
    Shows: original PI, revised PI, payments, delivery, claims, and net position."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Client users can only see their own orders
    if current_user.user_type == "CLIENT" and order.client_id != current_user.client_id:
        raise HTTPException(status_code=403, detail="Access denied")

    is_client = current_user.user_type == "CLIENT"

    # ── A. Order Items ──
    all_items = db.query(OrderItem).filter(OrderItem.order_id == order_id).all()
    active_items = [i for i in all_items if i.status == "ACTIVE"]
    migrated_items = [i for i in all_items if i.status == "UNLOADED"]

    # For transparency clients: use client_factory_price * exchange_rate
    client = db.query(Client).filter(Client.id == order.client_id).first()
    is_transparency = client and client.client_type == "TRANSPARENCY"
    rate = order.exchange_rate or 1.0

    def _item_client_total(item):
        if is_transparency:
            return round(float(item.client_factory_price or 0) * rate, 2) * (item.quantity or 0)
        return (item.selling_price_inr or 0) * (item.quantity or 0)

    original_pi_total = sum(_item_client_total(i) for i in all_items if i.status in ("ACTIVE", "UNLOADED"))
    revised_pi_total = sum(_item_client_total(i) for i in active_items)
    migrated_value = sum(_item_client_total(i) for i in migrated_items)

    items_summary = {
        "original_count": len([i for i in all_items if i.status in ("ACTIVE", "UNLOADED")]),
        "original_qty": sum(i.quantity or 0 for i in all_items if i.status in ("ACTIVE", "UNLOADED")),
        "shipped_count": len(active_items),
        "shipped_qty": sum(i.quantity or 0 for i in active_items),
        "migrated_count": len(migrated_items),
        "migrated_qty": sum(i.quantity or 0 for i in migrated_items),
    }

    # ── B. Payments ──
    payments = db.query(Payment).filter(Payment.order_id == order_id).all()
    client_payments = [p for p in payments if (p.method or "").upper() != "CREDIT" and getattr(p, 'verification_status', 'VERIFIED') == 'VERIFIED']
    total_paid = sum(p.amount_inr or 0 for p in client_payments)

    payment_summary = {
        "total_paid": round(total_paid, 2),
        "payment_count": len(client_payments),
        "payments": [{
            "date": p.payment_date.isoformat() if p.payment_date else None,
            "method": p.method,
            "type": p.payment_type,
            "amount_inr": round(p.amount_inr or 0, 2),
            "reference": p.reference,
        } for p in client_payments],
    }

    # ── C. Delivery / After-Sales Claims ──
    from models import AfterSalesItem
    as_items = db.query(AfterSalesItem).filter(AfterSalesItem.order_id == order_id).all()
    claims_with_issue = [a for a in as_items if a.objection_type and a.objection_type.strip()]

    total_claim_qty = sum(a.affected_quantity or 0 for a in claims_with_issue)
    total_claim_value = sum((a.selling_price_inr or 0) * (a.affected_quantity or 0) for a in claims_with_issue)

    # Resolution breakdown
    replace_items = [a for a in claims_with_issue if a.resolution_type and "REPLACE" in a.resolution_type]
    compensate_items = [a for a in claims_with_issue if a.resolution_type and "COMPENSATE" in a.resolution_type]
    pending_resolution = [a for a in claims_with_issue if not a.resolution_type]

    replace_value = sum((a.selling_price_inr or 0) * (a.affected_quantity or 0) for a in replace_items)
    compensate_value = sum(a.compensation_amount or (a.selling_price_inr or 0) * (a.affected_quantity or 0) for a in compensate_items)

    claims_summary = {
        "total_claims": len(claims_with_issue),
        "total_claim_qty": total_claim_qty,
        "total_claim_value": round(total_claim_value, 2),
        "resolved": len([a for a in claims_with_issue if a.status == "RESOLVED"]),
        "pending": len(pending_resolution),
        "replace_count": len(replace_items),
        "replace_value": round(replace_value, 2),
        "compensate_count": len(compensate_items),
        "compensate_value": round(compensate_value, 2),
        "items": [{
            "product_code": a.product_code,
            "product_name": a.product_name,
            "sent_qty": a.sent_qty,
            "received_qty": a.received_qty,
            "affected_qty": a.affected_quantity,
            "issue_type": a.objection_type,
            "resolution": a.resolution_type,
            "claim_value": round((a.selling_price_inr or 0) * (a.affected_quantity or 0), 2),
            "status": a.status,
            "carry_forward": a.carry_forward_status,
        } for a in claims_with_issue],
    }

    # ── D. Final Reconciliation ──
    good_items_value = revised_pi_total - total_claim_value
    credit_balance = total_paid - revised_pi_total
    net_adjustment = compensate_value  # money back to client
    final_net = credit_balance + net_adjustment

    reconciliation = {
        "original_pi_total": round(original_pi_total, 2),
        "migrated_value": round(migrated_value, 2),
        "revised_pi_total": round(revised_pi_total, 2),
        "total_paid": round(total_paid, 2),
        "good_items_value": round(good_items_value, 2),
        "total_claim_value": round(total_claim_value, 2),
        "compensate_value": round(net_adjustment, 2),
        "replace_value": round(replace_value, 2),
        "credit_balance": round(credit_balance, 2),
        "final_net_position": round(final_net, 2),
        "pending_replacements": len([a for a in replace_items if a.carry_forward_status == "PENDING"]),
    }

    result = {
        "order_number": order.order_number,
        "order_id": order.id,
        "status": order.status,
        "client_name": order.client.company_name if order.client else None,
        "po_reference": order.po_reference,
        "created_at": order.created_at.isoformat() if order.created_at else None,
        "completed_at": order.completed_at.isoformat() if order.completed_at else None,
        "items": items_summary,
        "payments": payment_summary,
        "claims": claims_summary,
        "reconciliation": reconciliation,
    }

    # Admin-only: include factory costs
    if not is_client:
        factory_total_cny = sum((i.factory_price or 0) * (i.quantity or 0) for i in active_items)
        result["factory"] = {
            "name": order.factory.company_name if order.factory else None,
            "total_cny": round(factory_total_cny, 2),
            "exchange_rate": order.exchange_rate,
            "total_inr": round(factory_total_cny * (order.exchange_rate or 1), 2),
            "margin_inr": round(revised_pi_total - factory_total_cny * (order.exchange_rate or 1), 2),
        }

    return result


@router.get("/{order_id}/")
def get_order(order_id: str, db: Session = Depends(get_db), current_user: CurrentUser = Depends(get_current_user)):
    """Get full order with items"""
    order = db.query(Order).options(
        joinedload(Order.client),
        joinedload(Order.factory),
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()

    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # RLS: verify CLIENT/FACTORY users can only access their own orders
    if current_user.role in ("CLIENT", "FACTORY"):
        if not verify_resource_access(Order, order_id, db, current_user):
            raise HTTPException(status_code=403, detail="Access denied")

    data = serialize_order(order, db)
    data["items"] = [serialize_order_item(i) for i in order.items]

    # Apply transparency pricing mask BEFORE role filter
    from core.transparency import mask_transparency_pricing
    client_type = order.client.client_type if order.client else "REGULAR"
    data = mask_transparency_pricing(data, current_user.role, client_type)

    # Query counts for tab badge
    from models import ItemQuery
    open_q = db.query(ItemQuery).filter(ItemQuery.order_id == order.id, ItemQuery.status == "OPEN").count()
    replied_q = db.query(ItemQuery).filter(ItemQuery.order_id == order.id, ItemQuery.status == "REPLIED").count()
    data["query_counts"] = {"open": open_q, "replied": replied_q, "total": open_q + replied_q}

    return filter_for_role(data, current_user.role)


@router.post("/client-inquiry/")
def create_client_inquiry(
    data: ClientOrderCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Create a client-initiated order inquiry. CLIENT role only."""
    from core.serializers import filter_for_role

    # Guard: CLIENT role only
    if current_user.role != "CLIENT" or not current_user.client_id:
        raise HTTPException(status_code=403, detail="Only CLIENT users can create inquiries")

    # Validate client exists
    client = db.query(Client).filter(Client.id == current_user.client_id).first()
    if not client:
        raise HTTPException(status_code=400, detail="Client account not found")

    # Auto-generate PO reference
    po_ref = data.po_reference or generate_po_reference(db, client.company_name)

    # Create order with forced values
    order = Order(
        client_id=current_user.client_id,  # forced from JWT
        factory_id=None,
        status=OrderStatus.CLIENT_DRAFT.value,
        currency=Currency.CNY.value,
        po_reference=po_ref,
    )
    db.add(order)
    db.flush()

    if data.client_reference:
        existing_ref = db.query(Order).filter(
            Order.client_reference == data.client_reference,
            Order.deleted_at.is_(None),
            Order.id != order.id,
        ).first()
        if existing_ref:
            raise HTTPException(status_code=409, detail=f"Reference '{data.client_reference}' already exists")
        order.client_reference = data.client_reference

    # Create items with zero prices
    for item_data in data.items:
        product = db.query(Product).filter(Product.id == item_data.product_id).first()
        if not product:
            continue  # skip invalid products

        order_item = OrderItem(
            order_id=order.id,
            product_id=product.id,
            quantity=item_data.quantity,
            product_code_snapshot=product.product_code,
            product_name_snapshot=product.product_name,
            material_snapshot=product.material,
            category_snapshot=product.category,
            part_type_snapshot=getattr(product, 'part_type', None),
            dimension_snapshot=getattr(product, 'dimension', None),
            variant_note_snapshot=getattr(product, 'variant_note', None),
            status="ACTIVE",
        )
        db.add(order_item)

    # Create product requests for quick-add items
    _requester_name = None
    if data.quick_add_items:
        from models import User as _User
        _user = db.query(_User).filter(_User.id == current_user.id).first()
        _requester_name = _user.full_name if _user else current_user.email

    for qa_item in data.quick_add_items:
        pr = ProductRequest(
            product_code=qa_item.product_code.strip(),
            product_name=qa_item.product_name.strip(),
            quantity=qa_item.quantity,
            client_id=current_user.client_id,
            order_id=order.id,
            status="PENDING",
            requested_by=current_user.id,
            requested_by_name=_requester_name,
        )
        db.add(pr)

    # Notify admin about product requests (batched)
    if data.quick_add_items:
        count = len(data.quick_add_items)
        existing_notif = db.query(Notification).filter(
            Notification.user_role == "ADMIN",
            Notification.notification_type == "PRODUCT_REVIEW_REQUEST",
            Notification.is_read == False,
            Notification.client_id == current_user.client_id,
        ).first()
        if existing_notif:
            existing_notif.count = existing_notif.count + count
            existing_notif.message = f"{existing_notif.count} new products need review"
            existing_notif.updated_at = datetime.utcnow()
        else:
            db.add(Notification(
                user_role="ADMIN",
                client_id=current_user.client_id,
                title="New Product Request",
                message=f"{count} new product(s) need review",
                notification_type="PRODUCT_REVIEW_REQUEST",
                resource_type="product_request",
                resource_id=order.id,
            ))

    db.commit()
    db.refresh(order)

    # Serialize and filter for client
    result = serialize_order(order, db)
    return filter_for_role(result, current_user.role)


@router.get("/{order_id}/product-requests/")
def get_order_product_requests(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Get product requests (Quick Add items) linked to an order."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Row-level security: CLIENT can only see own orders
    if current_user.role == "CLIENT" and order.client_id != current_user.client_id:
        raise HTTPException(status_code=403, detail="Access denied")

    requests = db.query(ProductRequest).filter(
        ProductRequest.order_id == order_id,
    ).order_by(ProductRequest.created_at.desc()).all()

    return {
        "requests": [
            {
                "id": pr.id,
                "product_code": pr.product_code,
                "product_name": pr.product_name,
                "quantity": pr.quantity,
                "status": pr.status,
                "created_at": pr.created_at.isoformat() if pr.created_at else None,
            }
            for pr in requests
        ]
    }


@router.post("/{order_id}/approve-inquiry/")
def approve_inquiry(
    order_id: str,
    data: ApproveInquiryRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Admin approves a CLIENT_DRAFT inquiry, advancing it to DRAFT."""
    # Only internal users (ADMIN/OPERATIONS) can approve
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Only internal users can approve inquiries")

    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if order.status != OrderStatus.CLIENT_DRAFT.value:
        raise HTTPException(status_code=400, detail="Order is not in CLIENT_DRAFT status")

    # Assign factory if provided
    if data.factory_id:
        factory = db.query(Factory).filter(Factory.id == data.factory_id).first()
        if not factory:
            raise HTTPException(status_code=400, detail="Factory not found")
        order.factory_id = data.factory_id

    # Set dealing currency from request (defaults to USD)
    if data.currency:
        order.currency = data.currency

    # Snapshot exchange rate for the selected currency
    rate_row = db.query(ExchangeRate).filter(
        ExchangeRate.from_currency == order.currency,
        ExchangeRate.to_currency == "INR",
    ).first()
    if rate_row:
        order.exchange_rate = rate_row.rate
        order.exchange_rate_date = date.today()

    # Transition to DRAFT
    old_status = order.status
    order.status = OrderStatus.DRAFT.value

    # Generate order number if not set
    if not order.order_number:
        count = db.query(Order).filter(
            Order.order_number.isnot(None),
        ).count()
        order.order_number = f"ORD-{datetime.utcnow().strftime('%y%m')}-{count + 1:04d}"

    db.commit()
    db.refresh(order)

    # Carry-forward pending items now that factory is assigned
    carry_result = _add_pending_items_to_order(order, db)
    carried_count = len(carry_result.get("carried_aftersales", [])) + len(carry_result.get("carried_items", []))
    if carried_count > 0:
        db.commit()
        db.refresh(order)

    # Audit log
    try:
        from core.audit import log_audit_event
        log_audit_event(
            db, current_user, "INQUIRY_APPROVED", "order", order.id,
            old_values={"status": old_status},
            new_values={"status": order.status, "factory_id": order.factory_id, "order_number": order.order_number},
        )
    except Exception:
        pass

    result = serialize_order(order, db)
    result["carried_forward_count"] = carried_count
    return result


@router.post("/")
def create_order(data: OrderCreate, db: Session = Depends(get_db), current_user: CurrentUser = Depends(get_current_user)):
    """Create a new draft order — admin/operations only"""
    if current_user.role not in ("ADMIN", "SUPER_ADMIN", "OPERATIONS"):
        raise HTTPException(status_code=403, detail="Insufficient permissions to create orders")
    # Validate client exists
    client = db.query(Client).filter(
        Client.id == data.client_id, Client.is_active == True
    ).first()
    if not client:
        raise HTTPException(status_code=400, detail="Client not found")

    # Validate factory if provided
    if data.factory_id:
        factory = db.query(Factory).filter(
            Factory.id == data.factory_id, Factory.is_active == True
        ).first()
        if not factory:
            raise HTTPException(status_code=400, detail="Factory not found")

    # Snapshot exchange rate from settings
    rate_row = db.query(ExchangeRate).filter(
        ExchangeRate.from_currency == data.currency,
        ExchangeRate.to_currency == "INR",
    ).first()

    order = Order(
        client_id=data.client_id,
        factory_id=data.factory_id,
        status=OrderStatus.DRAFT.value,
        currency=data.currency,
        exchange_rate=rate_row.rate if rate_row else None,
        exchange_rate_date=date.today() if rate_row else None,
        po_reference=data.po_reference,
        notes=data.notes,
    )
    db.add(order)
    db.flush()  # Get order.id before adding items

    if data.client_reference:
        # Check uniqueness
        existing_ref = db.query(Order).filter(
            Order.client_reference == data.client_reference,
            Order.deleted_at.is_(None),
            Order.id != order.id,
        ).first()
        if existing_ref:
            raise HTTPException(status_code=409, detail=f"Client reference '{data.client_reference}' already exists")
        order.client_reference = data.client_reference

    # Add items
    for item_data in data.items:
        product = db.query(Product).filter(
            Product.id == item_data.product_id,
            Product.is_active == True,
        ).first()
        if not product:
            raise HTTPException(status_code=400, detail=f"Product {item_data.product_id} not found")

        order_item = OrderItem(
            order_id=order.id,
            product_id=item_data.product_id,
            quantity=item_data.quantity,
            notes=item_data.notes,
            product_code_snapshot=product.product_code,
            product_name_snapshot=product.product_name,
            material_snapshot=product.material,
            part_type_snapshot=product.part_type,
            dimension_snapshot=product.dimension,
            variant_note_snapshot=product.variant_note,
            category_snapshot=product.category,
        )
        db.add(order_item)

    # Auto-add pending carry-forward items (after-sales + unloaded) via shared helper
    pending_result = _add_pending_items_to_order(order, db)
    carried_aftersales = pending_result["carried_aftersales"]
    carried_items = pending_result["carried_items"]

    db.commit()

    # Re-fetch with relationships
    order = db.query(Order).options(
        joinedload(Order.client),
        joinedload(Order.factory),
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.id == order.id).first()

    result = serialize_order(order, db)
    if carried_items:
        result["carried_items"] = carried_items
        result["carried_count"] = len(carried_items)
    if carried_aftersales:
        result["carried_aftersales"] = carried_aftersales
        result["carried_aftersales_count"] = len(carried_aftersales)
    return result


@router.put("/{order_id}/")
def update_order(order_id: str, data: OrderUpdate, db: Session = Depends(get_db), current_user: CurrentUser = Depends(get_current_user)):
    """Update order header (only in DRAFT or COMPLETED_EDITING status) — admin/operations only"""
    if current_user.role not in ("ADMIN", "SUPER_ADMIN", "OPERATIONS"):
        raise HTTPException(status_code=403, detail="Insufficient permissions to modify orders")
    order = db.query(Order).filter(
        Order.id == order_id, Order.deleted_at.is_(None)
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if order.status not in [OrderStatus.CLIENT_DRAFT.value, OrderStatus.DRAFT.value, OrderStatus.COMPLETED_EDITING.value]:
        raise HTTPException(status_code=400, detail="Order can only be edited in CLIENT_DRAFT, DRAFT or COMPLETED_EDITING status")

    if data.client_id is not None:
        client = db.query(Client).filter(Client.id == data.client_id, Client.is_active == True).first()
        if not client:
            raise HTTPException(status_code=400, detail="Client not found")
        order.client_id = data.client_id

    if data.factory_id is not None:
        if data.factory_id:
            factory = db.query(Factory).filter(Factory.id == data.factory_id, Factory.is_active == True).first()
            if not factory:
                raise HTTPException(status_code=400, detail="Factory not found")
        order.factory_id = data.factory_id or None

    if data.currency is not None:
        order.currency = data.currency
        # Re-snapshot exchange rate
        rate_row = db.query(ExchangeRate).filter(
            ExchangeRate.from_currency == data.currency,
            ExchangeRate.to_currency == "INR",
        ).first()
        order.exchange_rate = rate_row.rate if rate_row else order.exchange_rate
        order.exchange_rate_date = date.today() if rate_row else order.exchange_rate_date

    if data.po_reference is not None:
        order.po_reference = data.po_reference
    if data.notes is not None:
        order.notes = data.notes

    db.commit()

    order = db.query(Order).options(
        joinedload(Order.client),
        joinedload(Order.factory),
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.id == order.id).first()

    return serialize_order(order, db)


@router.delete("/{order_id}/")
def delete_order(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Soft delete order with reason. Admin can delete any order."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Only ADMIN can delete non-DRAFT orders
    if order.status != OrderStatus.DRAFT.value and order.status != OrderStatus.CLIENT_DRAFT.value:
        if current_user.role != "ADMIN":
            raise HTTPException(status_code=403, detail="Only admins can delete non-draft orders")

    # Revert unloaded items linked to this order back to PENDING
    db.query(UnloadedItem).filter(
        UnloadedItem.added_to_order_id == order_id
    ).update({
        UnloadedItem.status: "PENDING",
        UnloadedItem.added_to_order_id: None,
    })

    # Revert after-sales carry-forward items
    from models import AfterSalesItem
    db.query(AfterSalesItem).filter(
        AfterSalesItem.added_to_order_id == order_id
    ).update({
        AfterSalesItem.carry_forward_status: "PENDING",
        AfterSalesItem.added_to_order_id: None,
    })

    # Clean up orphaned pending products from Quick Add
    pending_items = db.query(OrderItem).filter(
        OrderItem.order_id == order_id,
        OrderItem.product_id.isnot(None),
    ).all()

    for oi in pending_items:
        product = db.query(Product).filter(
            Product.id == oi.product_id,
            Product.approval_status == "PENDING_APPROVAL",
        ).first()
        if not product:
            continue

        # Check if any OTHER active order references this pending product
        other_refs = db.query(OrderItem).join(Order).filter(
            OrderItem.product_id == product.id,
            OrderItem.order_id != order_id,
            Order.deleted_at.is_(None),
        ).count()
        if other_refs > 0:
            continue

        # No other active order uses it — hard delete (same as reject-quick-add)
        db.query(OrderItem).filter(OrderItem.product_id == product.id).delete()

        parent_id = product.parent_id
        db.delete(product)

        # Delete orphaned parent if PENDING_APPROVAL with no other children
        if parent_id:
            parent = db.query(Product).filter(Product.id == parent_id).first()
            if parent and parent.approval_status == "PENDING_APPROVAL":
                other_children = db.query(Product).filter(
                    Product.parent_id == parent_id,
                    Product.id != product.id,
                ).count()
                if other_children == 0:
                    db.delete(parent)

    order.deleted_at = datetime.utcnow()
    order.deleted_by = current_user.id

    # Create notification for client
    if order.client_id:
        from models import Notification
        db.add(Notification(
            client_id=order.client_id,
            title="Order Cancelled",
            message=f"Order {order.order_number or 'Draft'} has been cancelled.",
            notification_type="ORDER_DELETED",
            resource_type="order",
            resource_id=order.id,
        ))

    db.commit()
    return {"message": "Order deleted"}


@router.put("/{order_id}/delete-reason/")
def set_deletion_reason(
    order_id: str,
    body: dict,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Set or update deletion reason on a soft-deleted order."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    reason = body.get("reason", "")
    order.deletion_reason = reason
    order.deleted_by = current_user.id

    # Update client notification with reason
    if order.client_id and reason:
        from models import Notification
        existing = db.query(Notification).filter(
            Notification.resource_id == order_id,
            Notification.notification_type == "ORDER_DELETED",
            Notification.is_read == False,
        ).first()
        if existing:
            existing.message = f"Order {order.order_number or 'Draft'} cancelled: {reason}"
        else:
            db.add(Notification(
                client_id=order.client_id,
                title="Order Cancelled",
                message=f"Order {order.order_number or 'Draft'} cancelled: {reason}",
                notification_type="ORDER_DELETED",
                resource_type="order",
                resource_id=order.id,
            ))

    db.commit()
    return {"message": "Deletion reason updated"}


# ========================================
# FETCH PENDING ITEMS (carry-forward + unloaded)
# ========================================

def _add_pending_items_to_order(order: Order, db: Session) -> dict:
    """Shared helper: find and add pending after-sales carry-forward items and
    pending unloaded items for the same client + factory to the given order.
    Returns counts of what was added."""
    from models import AfterSalesItem
    from enums import CarryForwardStatus, CarryForwardType, AfterSalesStatus

    carried_aftersales = []
    carried_items = []

    if not order.factory_id:
        return {"carried_aftersales": carried_aftersales, "carried_items": carried_items}

    client = db.query(Client).filter(Client.id == order.client_id).first()

    # Guard: check what's already carried into this order to prevent duplicates
    # This makes the function idempotent — safe to call from any path
    existing_as_sources = set(
        r[0] for r in db.query(AfterSalesItem.source_aftersales_id).filter(
            AfterSalesItem.order_id == order.id,
            AfterSalesItem.source_aftersales_id.isnot(None),
        ).all()
    )
    existing_ui_notes = set(
        r[0] for r in db.query(OrderItem.notes).filter(
            OrderItem.order_id == order.id,
            OrderItem.notes.like("Carried from %"),
        ).all()
    )

    # --- After-Sales carry-forward (HIGHER PRIORITY) ---
    pending_claims = db.query(AfterSalesItem).filter(
        AfterSalesItem.client_id == order.client_id,
        AfterSalesItem.factory_id == order.factory_id,
        AfterSalesItem.carry_forward_status == CarryForwardStatus.PENDING.value,
    ).all()

    for claim in pending_claims:
        # Skip if already carried into this order (idempotency guard)
        if claim.id in existing_as_sources:
            continue
        if claim.carry_forward_type == CarryForwardType.REPLACEMENT.value:
            sell_price = 0  # Free replacement
        elif claim.carry_forward_type == CarryForwardType.COMPENSATION.value:
            sell_price = -(claim.selling_price_inr or 0)  # Negate original per-unit price
        else:
            sell_price = 0

        prod = db.query(Product).filter(Product.id == claim.product_id).first()
        resolution_label = (claim.resolution_type or "").replace("_", " ").title()
        new_oi = OrderItem(
            order_id=order.id,
            product_id=claim.product_id,
            quantity=claim.affected_quantity or 1,
            factory_price=0,
            selling_price_inr=sell_price,
            notes=f"After-Sales ({resolution_label}) from {claim.order_number}",
            product_code_snapshot=prod.product_code if prod else claim.product_code,
            product_name_snapshot=prod.product_name if prod else claim.product_name,
            material_snapshot=prod.material if prod else None,
            part_type_snapshot=prod.part_type if prod else None,
            dimension_snapshot=prod.dimension if prod else None,
            variant_note_snapshot=prod.variant_note if prod else None,
            category_snapshot=prod.category if prod else None,
        )
        db.add(new_oi)
        db.flush()  # Get new_oi.id for the AfterSalesItem reference

        # Create AfterSalesItem in the new order for tracking fulfillment
        asi = AfterSalesItem(
            order_id=order.id,
            order_number=order.order_number or "",
            order_item_id=new_oi.id,
            product_id=claim.product_id,
            product_code=prod.product_code if prod else claim.product_code,
            product_name=prod.product_name if prod else claim.product_name,
            client_id=order.client_id,
            client_name=client.company_name if client else "",
            factory_id=order.factory_id,
            sent_qty=0,
            received_qty=0,
            ordered_quantity=claim.affected_quantity or 1,
            delivered_quantity=0,
            selling_price_inr=sell_price,
            affected_quantity=0,
            total_value_inr=0,
            objection_type="",
            description="",
            status=AfterSalesStatus.OPEN.value,
            carry_forward_type=claim.carry_forward_type,
            source_aftersales_id=claim.id,
        )
        db.add(asi)

        claim.carry_forward_status = CarryForwardStatus.ADDED_TO_ORDER.value
        claim.added_to_order_id = order.id
        carried_aftersales.append({
            "product_code": prod.product_code if prod else claim.product_code,
            "product_name": prod.product_name if prod else claim.product_name,
            "quantity": claim.affected_quantity or 1,
            "resolution_type": claim.resolution_type,
            "original_order_number": claim.order_number,
        })

    # --- Unloaded items for same client + factory ---
    source_order_ids = [o.id for o in db.query(Order.id).filter(
        Order.factory_id == order.factory_id
    ).all()]
    if source_order_ids:
        pending_unloaded = db.query(UnloadedItem).filter(
            UnloadedItem.client_id == order.client_id,
            UnloadedItem.original_order_id.in_(source_order_ids),
            UnloadedItem.status == UnloadedItemStatus.PENDING.value,
        ).all()

        for ui in pending_unloaded:
            # Skip carry-forwards from the current order (balance goes to next order)
            if ui.original_order_id == order.id:
                continue
            orig_order_num = db.query(Order.order_number).filter(Order.id == ui.original_order_id).scalar()
            # If product already exists as ACTIVE, add carry-forward qty to it
            existing_active = db.query(OrderItem).filter(
                OrderItem.order_id == order.id,
                OrderItem.product_id == ui.product_id,
                OrderItem.status == "ACTIVE",
            ).first()
            if existing_active:
                existing_active.quantity = (existing_active.quantity or 0) + ui.quantity
                existing_active.notes = f"Carried from {orig_order_num or 'previous order'} (+{ui.quantity} balance)"
                ui.status = UnloadedItemStatus.ADDED_TO_ORDER.value
                carried_items.append({
                    "product_code": existing_active.product_code_snapshot,
                    "product_name": existing_active.product_name_snapshot,
                    "quantity": ui.quantity,
                })
                continue
            # Skip if already carried into this order (idempotency guard)
            expected_note = f"Carried from {orig_order_num or 'previous order'}"
            if expected_note in existing_ui_notes:
                continue
            prod = db.query(Product).filter(Product.id == ui.product_id).first()
            # Calculate client_factory_price for transparency clients
            _cfp = None
            if _config.TRANSPARENCY_ENABLED and ui.factory_price:
                _cl = db.query(Client).filter(Client.id == order.client_id).first()
                if _cl and _cl.client_type == "TRANSPARENCY" and _cl.factory_markup_percent is not None:
                    _m = Decimal(str(_cl.factory_markup_percent))
                    _f = Decimal(str(ui.factory_price))
                    _cfp = float((_f * (1 + _m / 100)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            db.add(OrderItem(
                order_id=order.id,
                product_id=ui.product_id,
                quantity=ui.quantity,
                factory_price=ui.factory_price,
                client_factory_price=_cfp,
                notes=f"Carried from {orig_order_num or 'previous order'}",
                product_code_snapshot=prod.product_code if prod else None,
                product_name_snapshot=prod.product_name if prod else None,
                material_snapshot=prod.material if prod else None,
                part_type_snapshot=prod.part_type if prod else None,
                dimension_snapshot=prod.dimension if prod else None,
                variant_note_snapshot=prod.variant_note if prod else None,
                category_snapshot=prod.category if prod else None,
            ))
            ui.status = UnloadedItemStatus.ADDED_TO_ORDER.value
            ui.added_to_order_id = order.id
            carried_items.append({
                "product_code": prod.product_code if prod else None,
                "product_name": prod.product_name if prod else None,
                "quantity": ui.quantity,
                "original_order_number": orig_order_num,
            })

    return {"carried_aftersales": carried_aftersales, "carried_items": carried_items}


@router.post("/{order_id}/fetch-pending-items/")
def fetch_pending_items(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Manually fetch pending after-sales carry-forward and unloaded items into an existing order."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Not authorized")
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if order.status not in ITEM_EDITABLE_STAGES:
        raise HTTPException(status_code=400, detail="Items can only be added up to Factory Ordered stage or in Completed Editing")

    if not order.factory_id:
        raise HTTPException(status_code=400, detail="Order must have a factory assigned to fetch pending items")

    result = _add_pending_items_to_order(order, db)
    total_added = len(result["carried_aftersales"]) + len(result["carried_items"])

    # Mark PI as stale if items were added and PI exists
    if total_added > 0:
        pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order.id).first()
        if pi:
            order.items_modified_at = datetime.utcnow()

    db.commit()

    parts = []
    if result["carried_aftersales"]:
        parts.append(f"{len(result['carried_aftersales'])} after-sales item(s)")
    if result["carried_items"]:
        parts.append(f"{len(result['carried_items'])} unloaded item(s)")
    message = f"Added {', '.join(parts)}" if parts else "No pending items found"

    return {
        "carried_aftersales": result["carried_aftersales"],
        "carried_aftersales_count": len(result["carried_aftersales"]),
        "carried_items": result["carried_items"],
        "carried_count": len(result["carried_items"]),
        "total_added": total_added,
        "message": message,
    }


# ========================================
# BULK TEXT ADD (preview + apply)
# ========================================

@router.post("/{order_id}/bulk-text-add/")
def bulk_text_add_preview(
    order_id: str,
    data: BulkTextAddRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Parse pasted text lines of 'CODE [QTY]', validate each, and return preview results."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Not authorized")
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if order.status not in ITEM_EDITABLE_STAGES:
        raise HTTPException(status_code=400, detail="Items can only be modified up to Factory Ordered stage or in Completed Editing")

    # Get existing active product_ids in this order for duplicate check
    existing_product_ids = set(
        pid for (pid,) in db.query(OrderItem.product_id).filter(
            OrderItem.order_id == order.id,
            OrderItem.status == "ACTIVE",
        ).all()
    )

    results = []
    for line in data.lines:
        line = line.strip()
        if not line:
            continue

        # Parse: split by whitespace, last token is qty if numeric
        parts = line.split()
        qty = 1
        code = line
        if len(parts) >= 2:
            try:
                qty = int(parts[-1])
                code = " ".join(parts[:-1])
            except ValueError:
                code = line
                qty = 1

        # Lookup product by code
        matches = db.query(Product).filter(
            Product.product_code == code.strip(),
            Product.is_active == True,
            Product.deleted_at.is_(None),
        ).all()

        def _m(m):
            return {
                "id": m.id,
                "product_code": m.product_code,
                "product_name": m.product_name,
                "material": m.material,
                "part_type": m.part_type,
                "dimension": m.dimension,
                "variant_note": m.variant_note,
                "is_default": m.is_default,
            }

        if len(matches) == 0:
            results.append({"code": code, "status": "NOT_FOUND", "quantity": qty, "product_id": None, "product_name": None, "matches": []})
        elif len(matches) == 1:
            m = matches[0]
            already = m.id in existing_product_ids
            results.append({
                "code": code,
                "status": "ALREADY_IN_ORDER" if already else "FOUND",
                "quantity": qty,
                "product_id": m.id,
                "product_name": m.product_name,
                "matches": [_m(m)],
            })
        else:
            # Multiple matches — try to pick default
            default_match = next((m for m in matches if m.is_default), None)
            if default_match:
                already = default_match.id in existing_product_ids
                results.append({
                    "code": code,
                    "status": "ALREADY_IN_ORDER" if already else "FOUND",
                    "quantity": qty,
                    "product_id": default_match.id,
                    "product_name": default_match.product_name,
                    "matches": [_m(m) for m in matches],
                })
            else:
                results.append({
                    "code": code,
                    "status": "AMBIGUOUS",
                    "quantity": qty,
                    "product_id": None,
                    "product_name": None,
                    "matches": [_m(m) for m in matches],
                })

    counts = {"found": 0, "not_found": 0, "ambiguous": 0, "already_in_order": 0}
    for r in results:
        if r["status"] == "FOUND": counts["found"] += 1
        elif r["status"] == "NOT_FOUND": counts["not_found"] += 1
        elif r["status"] == "AMBIGUOUS": counts["ambiguous"] += 1
        elif r["status"] == "ALREADY_IN_ORDER": counts["already_in_order"] += 1

    return {"results": results, "counts": counts}


@router.post("/{order_id}/bulk-text-add/apply/")
def bulk_text_add_apply(
    order_id: str,
    data: BulkTextApplyRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Apply resolved bulk text items — delegates to add_order_items logic."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Not authorized")
    return add_order_items(order_id, OrderAddItems(items=data.items), db, current_user)


# ========================================
# ORDER ITEMS
# ========================================

@router.post("/{order_id}/items/")
def add_order_items(
    order_id: str,
    data: OrderAddItems,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Add items to an order (editable stages or mid-production)"""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    is_editable = order.status in ITEM_EDITABLE_STAGES
    is_mid_order = order.status in MID_ORDER_ADD_STAGES
    is_client = current_user.user_type == "CLIENT"
    needs_approval = is_mid_order or is_client
    if not is_editable and not is_mid_order:
        raise HTTPException(status_code=400, detail="Items cannot be added at this stage")

    added = []
    for item_data in data.items:
        product = db.query(Product).filter(
            Product.id == item_data.product_id, Product.is_active == True
        ).first()
        if not product:
            raise HTTPException(status_code=400, detail=f"Product {item_data.product_id} not found")

        # Check if this product was previously REMOVED from this order — reactivate it
        existing_removed = db.query(OrderItem).filter(
            OrderItem.order_id == order.id,
            OrderItem.product_id == item_data.product_id,
            OrderItem.status == "REMOVED",
        ).first()

        if existing_removed:
            # Reactivate: restore the old item with all its pricing data intact
            existing_removed.status = "ACTIVE"
            existing_removed.cancel_note = None
            if needs_approval:
                existing_removed.pi_item_status = "PENDING"
            # Remove corresponding UnloadedItem if it exists (no longer pending)
            db.query(UnloadedItem).filter(
                UnloadedItem.original_order_id == order.id,
                UnloadedItem.product_id == item_data.product_id,
                UnloadedItem.status == UnloadedItemStatus.PENDING.value,
            ).delete()
            # Update quantity if provided, otherwise keep old quantity
            if item_data.quantity and item_data.quantity > 0:
                existing_removed.quantity = item_data.quantity
            # Refresh snapshots with latest product data
            existing_removed.product_code_snapshot = product.product_code
            existing_removed.product_name_snapshot = product.product_name
            existing_removed.material_snapshot = product.material
            existing_removed.part_type_snapshot = product.part_type
            existing_removed.dimension_snapshot = product.dimension
            existing_removed.variant_note_snapshot = product.variant_note
            existing_removed.category_snapshot = product.category
            added.append(existing_removed)
        else:
            order_item = OrderItem(
                order_id=order.id,
                product_id=item_data.product_id,
                quantity=item_data.quantity,
                notes=item_data.notes,
                product_code_snapshot=product.product_code,
                product_name_snapshot=product.product_name,
                material_snapshot=product.material,
                part_type_snapshot=product.part_type,
                dimension_snapshot=product.dimension,
                variant_note_snapshot=product.variant_note,
                category_snapshot=product.category,
            )
            if needs_approval:
                order_item.pi_item_status = "PENDING"
            db.add(order_item)
            db.flush()
            added.append(order_item)

    # Mark PI as stale if PI already exists
    if added:
        pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order.id).first()
        if pi:
            order.items_modified_at = datetime.utcnow()

    # Notify admin when client adds items (needs approval)
    if is_client and added:
        db.add(Notification(
            user_role="ADMIN",
            title="Client Added Items",
            message=f"Client added {len(added)} items to {order.order_number} (Lot {next_lot}) — review and approve",
            notification_type="ITEMS_PENDING_APPROVAL",
            resource_type="order",
            resource_id=order_id,
        ))
    # Notify client when admin adds mid-order items
    elif is_mid_order and added and not is_client:
        db.add(Notification(
            user_role="CLIENT",
            client_id=order.client_id,
            title="New Items Added",
            message=f"New items added to {order.order_number} — review and confirm",
            notification_type="ITEMS_PENDING_CONFIRMATION",
            resource_type="order",
            resource_id=order_id,
        ))

    db.commit()

    # Audit log for mid-order additions
    if is_mid_order and added:
        from core.audit import log_audit_event
        codes = [a.product_code_snapshot for a in added]
        log_audit_event(
            db, current_user, "MID_ORDER_ITEMS_ADDED", "order", order_id,
            new_values={"count": len(added), "product_codes": codes},
        )
        db.commit()

    # Re-fetch items with product join
    result = []
    for item in added:
        item_with_product = db.query(OrderItem).options(
            joinedload(OrderItem.product)
        ).filter(OrderItem.id == item.id).first()
        result.append(serialize_order_item(item_with_product))

    return result


@router.post("/{order_id}/items/{item_id}/confirm/")
def confirm_order_item(
    order_id: str,
    item_id: str,
    action: str = Query(..., pattern="^(approve|reject)$"),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Client or admin confirms/rejects a pending mid-order item."""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Not authorized")
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # RLS: CLIENT can only confirm own orders
    if current_user.user_type == "CLIENT" and order.client_id != current_user.client_id:
        raise HTTPException(status_code=403, detail="Access denied")

    item = db.query(OrderItem).filter(
        OrderItem.id == item_id,
        OrderItem.order_id == order_id,
        OrderItem.status == "ACTIVE",
        OrderItem.pi_item_status == "PENDING",
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Pending item not found")

    if action == "approve":
        item.pi_item_status = "APPROVED"
        # Assign lot number if not already set
        if not item.pi_addition_lot:
            from sqlalchemy import func as _fn
            max_lot = db.query(_fn.max(OrderItem.pi_addition_lot)).filter(
                OrderItem.order_id == order.id
            ).scalar()
            item.pi_addition_lot = (max_lot or 0) + 1
        order.items_modified_at = datetime.utcnow()  # Mark PI stale
    else:
        item.pi_item_status = "REJECTED"

    # Notify admin when client confirms/rejects
    if current_user.user_type == "CLIENT" and action == "approve":
        db.add(Notification(
            user_role="ADMIN",
            title="Item Confirmed",
            message=f"Client confirmed {item.product_code_snapshot} for {order.order_number}",
            notification_type="ITEM_CONFIRMED",
            resource_type="order",
            resource_id=order_id,
        ))

    # Audit log
    from core.audit import log_audit_event
    audit_action = "MID_ORDER_ITEMS_APPROVED" if action == "approve" else "MID_ORDER_ITEMS_REJECTED"
    log_audit_event(
        db, current_user, audit_action, "order", order_id,
        new_values={"product_code": item.product_code_snapshot, "item_id": item_id},
    )

    db.commit()
    return {"status": item.pi_item_status, "item_id": item_id}


@router.post("/{order_id}/items/bulk-confirm/")
def bulk_confirm_order_items(
    order_id: str,
    action: str = Query(..., pattern="^(approve|reject)$"),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Bulk approve/reject all PENDING items in an order. Assigns same lot number to all approved items."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if current_user.user_type == "CLIENT" and order.client_id != current_user.client_id:
        raise HTTPException(status_code=403, detail="Access denied")

    is_client = current_user.user_type == "CLIENT"

    # Client confirms PENDING items → CONFIRMED (price accepted)
    # Admin approves CONFIRMED items → APPROVED (merged to main list)
    if is_client:
        target_status = "PENDING"
    else:
        target_status = "CONFIRMED"  # Admin approves client-confirmed items

    items = db.query(OrderItem).filter(
        OrderItem.order_id == order_id,
        OrderItem.status == "ACTIVE",
        OrderItem.pi_item_status == target_status,
    ).all()

    if not items:
        # Fallback: admin can also approve PENDING items directly
        if not is_client:
            items = db.query(OrderItem).filter(
                OrderItem.order_id == order_id,
                OrderItem.status == "ACTIVE",
                OrderItem.pi_item_status == "PENDING",
            ).all()
        if not items:
            raise HTTPException(status_code=404, detail="No items to process")

    from core.audit import log_audit_event
    next_lot = None

    if action == "approve":
        if is_client:
            # Client confirms prices → CONFIRMED
            for item in items:
                item.pi_item_status = "CONFIRMED"
            log_audit_event(
                db, current_user, "MID_ORDER_ITEMS_CLIENT_CONFIRMED", "order", order_id,
                new_values={"count": len(items)},
            )
            db.add(Notification(
                user_role="ADMIN",
                title="Client Confirmed Prices",
                message=f"Client accepted prices for {len(items)} items in {order.order_number} — approve to add to order",
                notification_type="ITEMS_CLIENT_CONFIRMED",
                resource_type="order",
                resource_id=order_id,
            ))
        else:
            # Admin final approval → APPROVED with lot number
            from sqlalchemy import func as _fn
            max_lot = db.query(_fn.max(OrderItem.pi_addition_lot)).filter(
                OrderItem.order_id == order.id
            ).scalar()
            next_lot = (max_lot or 0) + 1
            codes = []
            for item in items:
                item.pi_item_status = "APPROVED"
                item.pi_addition_lot = next_lot
                codes.append(item.product_code_snapshot)
            order.items_modified_at = datetime.utcnow()
            log_audit_event(
                db, current_user, "MID_ORDER_ITEMS_APPROVED", "order", order_id,
                new_values={"count": len(items), "lot": next_lot, "product_codes": codes},
            )
            db.add(Notification(
                user_role="CLIENT",
                client_id=order.client_id,
                title="Items Added to Order",
                message=f"{len(items)} items approved and added to {order.order_number} (Lot {next_lot})",
                notification_type="ITEMS_APPROVED",
                resource_type="order",
                resource_id=order_id,
            ))
    else:
        for item in items:
            item.pi_item_status = "REJECTED"
        log_audit_event(
            db, current_user, "MID_ORDER_ITEMS_REJECTED", "order", order_id,
            new_values={"count": len(items)},
        )

    db.commit()
    return {"action": action, "count": len(items), "lot": next_lot}


@router.post("/{order_id}/items/send-prices/")
def send_pending_prices(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Admin sends pricing notification to client for pending items."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    pending = db.query(OrderItem).filter(
        OrderItem.order_id == order_id,
        OrderItem.status == "ACTIVE",
        OrderItem.pi_item_status == "PENDING",
    ).all()
    priced = [i for i in pending if i.factory_price or i.selling_price_inr or i.client_factory_price]
    if not priced:
        raise HTTPException(status_code=400, detail="No priced pending items to send")

    db.add(Notification(
        user_role="CLIENT",
        client_id=order.client_id,
        title="Prices Ready for Review",
        message=f"{len(priced)} new items in {order.order_number} have been priced — review and confirm",
        notification_type="PRICES_SENT_FOR_REVIEW",
        resource_type="order",
        resource_id=order_id,
    ))

    from core.audit import log_audit_event
    log_audit_event(
        db, current_user, "PRICES_SENT_TO_CLIENT", "order", order_id,
        new_values={"count": len(priced)},
    )

    db.commit()
    return {"sent": len(priced)}


@router.put("/{order_id}/items/{item_id}/")
def update_order_item(
    order_id: str,
    item_id: str,
    data: OrderUpdateItem,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Update an order item (quantity, notes)"""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Not authorized")
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    item = db.query(OrderItem).filter(
        OrderItem.id == item_id, OrderItem.order_id == order_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Order item not found")

    # Allow editing in editable stages, OR pending items in mid-order stages
    is_pending_mid_order = item.pi_item_status == "PENDING" and order.status in MID_ORDER_ADD_STAGES
    if order.status not in ITEM_EDITABLE_STAGES and not is_pending_mid_order:
        raise HTTPException(status_code=400, detail="Items can only be modified up to Factory Ordered stage or in Completed Editing")

    if data.quantity is not None:
        if data.quantity < 1:
            raise HTTPException(status_code=400, detail="Quantity must be at least 1")
        item.quantity = data.quantity
    if data.notes is not None:
        item.notes = data.notes

    # Mark PI as stale if PI already exists
    pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order.id).first()
    if pi:
        order.items_modified_at = datetime.utcnow()

    db.commit()

    item = db.query(OrderItem).options(joinedload(OrderItem.product)).filter(OrderItem.id == item.id).first()
    return serialize_order_item(item)


@router.put("/{order_id}/items/{item_id}/remove/")
def remove_order_item_put(
    order_id: str,
    item_id: str,
    data: OrderRemoveItem = None,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Remove item from order (soft delete — mark as REMOVED) with optional cancel_note"""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Not authorized")
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    item = db.query(OrderItem).filter(
        OrderItem.id == item_id, OrderItem.order_id == order_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Order item not found")

    # Allow removal in editable stages, OR pending items in mid-order stages
    is_pending_mid_order = item.pi_item_status == "PENDING" and order.status in MID_ORDER_ADD_STAGES
    if order.status not in ITEM_EDITABLE_STAGES and not is_pending_mid_order:
        raise HTTPException(status_code=400, detail="Items can only be modified up to Factory Ordered stage or in Completed Editing")

    item.status = "REMOVED"
    if data and data.cancel_note:
        item.cancel_note = data.cancel_note

    # If this item was carried from a previous order, restore the original UnloadedItem to PENDING
    _restore_carried_item_source(item, db)

    # Create UnloadedItem record for carry-forward to next order
    # Only if the order is beyond DRAFT (has been committed to) and item was confirmed
    if order.status != OrderStatus.DRAFT.value and not is_pending_mid_order:
        existing_ui = db.query(UnloadedItem).filter(
            UnloadedItem.original_order_id == order.id,
            UnloadedItem.product_id == item.product_id,
            UnloadedItem.status == UnloadedItemStatus.PENDING.value,
        ).first()
        if not existing_ui:
            db.add(UnloadedItem(
                original_order_id=order.id,
                client_id=order.client_id,
                product_id=item.product_id,
                quantity=item.quantity,
                amount_paid_inr=0,
                status=UnloadedItemStatus.PENDING.value,
                reason=data.cancel_note if data and data.cancel_note else "REMOVED",
                factory_price=item.factory_price,
            ))

    # Mark PI as stale if PI already exists
    pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order.id).first()
    if pi:
        order.items_modified_at = datetime.utcnow()

    db.commit()
    return {"message": "Item removed"}


@router.delete("/{order_id}/items/{item_id}/")
def remove_order_item(
    order_id: str,
    item_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Remove item from order (soft delete — mark as REMOVED) — backward compatible DELETE"""
    if current_user.user_type != "INTERNAL":
        raise HTTPException(status_code=403, detail="Not authorized")
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    item = db.query(OrderItem).filter(
        OrderItem.id == item_id, OrderItem.order_id == order_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Order item not found")

    # Allow removal in editable stages, OR pending items in mid-order stages
    is_pending_mid_order = item.pi_item_status == "PENDING" and order.status in MID_ORDER_ADD_STAGES
    if order.status not in ITEM_EDITABLE_STAGES and not is_pending_mid_order:
        raise HTTPException(status_code=400, detail="Items can only be modified up to Factory Ordered stage or in Completed Editing")

    item.status = "REMOVED"

    # If this item was carried from a previous order, restore the original UnloadedItem to PENDING
    _restore_carried_item_source(item, db)

    # Create UnloadedItem record for carry-forward to next order (skip for pending mid-order items)
    if order.status != OrderStatus.DRAFT.value and not is_pending_mid_order:
        existing_ui = db.query(UnloadedItem).filter(
            UnloadedItem.original_order_id == order.id,
            UnloadedItem.product_id == item.product_id,
            UnloadedItem.status == UnloadedItemStatus.PENDING.value,
        ).first()
        if not existing_ui:
            db.add(UnloadedItem(
                original_order_id=order.id,
                client_id=order.client_id,
                product_id=item.product_id,
                quantity=item.quantity,
                amount_paid_inr=0,
                status=UnloadedItemStatus.PENDING.value,
                reason="REMOVED",
                factory_price=item.factory_price,
            ))

    # Mark PI as stale if PI already exists
    pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order.id).first()
    if pi:
        order.items_modified_at = datetime.utcnow()

    db.commit()
    return {"message": "Item removed"}


# ========================================
# ITEM PRICING (Stage 2 — PENDING_PI)
# ========================================

@router.put("/{order_id}/items/{item_id}/prices/")
def update_item_prices(
    order_id: str, item_id: str, data: ItemPriceUpdate, db: Session = Depends(get_db)
):
    """Update factory price, markup, and/or selling price for an order item.
    Allowed at PENDING_PI or COMPLETED_EDITING stage, or for PENDING mid-order items."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    item = db.query(OrderItem).filter(
        OrderItem.id == item_id, OrderItem.order_id == order_id, OrderItem.status == "ACTIVE"
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Order item not found")

    is_pending_mid_order = item.pi_item_status == "PENDING" and order.status in MID_ORDER_ADD_STAGES
    if order.status not in ITEM_EDITABLE_STAGES and not is_pending_mid_order:
        raise HTTPException(
            status_code=400,
            detail=f"Prices can only be modified up to Factory Ordered stage or in Completed Editing (current: {order.status})"
        )

    if data.factory_price is not None:
        item.factory_price = data.factory_price
    if data.markup_percent is not None:
        item.markup_percent = data.markup_percent
    if data.selling_price_inr is not None:
        item.selling_price_inr = data.selling_price_inr

    # Auto-calculate selling price in CNY if factory price + markup provided
    if item.factory_price and item.markup_percent is not None:
        item.selling_price = round(item.factory_price * (1 + item.markup_percent / 100), 2)

    # Auto-calculate selling price in INR if selling_price + exchange_rate available
    if item.selling_price and order.exchange_rate and data.selling_price_inr is None:
        item.selling_price_inr = round(item.selling_price * order.exchange_rate, 2)

    # Auto-calculate client_factory_price for transparency clients
    client = db.query(Client).filter(Client.id == order.client_id).first()
    _calc_client_factory_price(item, client)

    # Mark PI as stale if PI already exists
    pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order.id).first()
    if pi:
        order.items_modified_at = datetime.utcnow()

    # Audit log for pricing pending mid-order items — show client-facing price, not real factory
    if is_pending_mid_order:
        from core.audit import log_audit_event
        client_price = float(item.client_factory_price or item.factory_price or 0)
        log_audit_event(
            db, None, "MID_ORDER_ITEM_PRICED", "order", order_id,
            new_values={"product_code": item.product_code_snapshot, "client_price": client_price},
        )

    db.commit()
    item = db.query(OrderItem).options(joinedload(OrderItem.product)).filter(OrderItem.id == item.id).first()
    return serialize_order_item(item)


@router.post("/{order_id}/recalculate-prices/")
def recalculate_prices(
    order_id: str,
    refresh_rate: bool = Query(False, description="Fetch latest exchange rate from DB before recalculating"),
    db: Session = Depends(get_db),
):
    """Recalculate all derived prices for an order.

    For ALL orders: recalculates selling_price and selling_price_inr from
    factory_price, markup_percent, and exchange_rate.
    For TRANSPARENCY orders: also recalculates client_factory_price.
    If refresh_rate=true, fetches the latest exchange rate from the DB first.
    """
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Optionally refresh exchange rate from the latest DB value
    old_rate = order.exchange_rate
    if refresh_rate:
        rate_row = db.query(ExchangeRate).filter(
            ExchangeRate.from_currency == (order.currency or "USD"),
            ExchangeRate.to_currency == "INR",
        ).first()
        if rate_row:
            order.exchange_rate = rate_row.rate
            order.exchange_rate_date = date.today()

    client = db.query(Client).filter(Client.id == order.client_id).first()
    is_transparency = client and client.client_type == "TRANSPARENCY"
    rate = order.exchange_rate

    items = db.query(OrderItem).filter(
        OrderItem.order_id == order_id, OrderItem.status == "ACTIVE"
    ).all()

    updated = 0
    cfp_updated = 0
    for item in items:
        changed = False
        # Skip after-sales items with negative/zero pricing
        if item.selling_price_inr is not None and item.selling_price_inr < 0:
            continue

        # Recalculate selling prices from factory + markup
        if item.factory_price and item.markup_percent is not None:
            fp = Decimal(str(item.factory_price))
            mp = Decimal(str(item.markup_percent))
            new_cny = float((fp * (1 + mp / 100)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
            if item.selling_price != new_cny:
                item.selling_price = new_cny
                changed = True
            if rate and item.selling_price:
                new_inr = float((Decimal(str(item.selling_price)) * Decimal(str(rate))).quantize(
                    Decimal('0.01'), rounding=ROUND_HALF_UP))
                if item.selling_price_inr != new_inr:
                    item.selling_price_inr = new_inr
                    changed = True

        # Recalculate client_factory_price for transparency clients
        if is_transparency:
            old_cfp = item.client_factory_price
            _calc_client_factory_price(item, client)
            if item.client_factory_price != old_cfp:
                cfp_updated += 1
                changed = True

        if changed:
            updated += 1

    db.commit()
    result = {
        "recalculated": updated,
        "cfp_updated": cfp_updated,
        "total_items": len(items),
    }
    if refresh_rate:
        result["exchange_rate"] = order.exchange_rate
        result["exchange_rate_date"] = order.exchange_rate_date.isoformat() if order.exchange_rate_date else None
        result["rate_changed"] = old_rate != order.exchange_rate
    return result


@router.post("/{order_id}/copy-previous-prices/")
def copy_previous_prices(order_id: str, db: Session = Depends(get_db)):
    """Copy factory prices from past same-factory orders, per-part.
    For each item, finds its most recent price from ANY past order with the same factory.
    Items with no past order history are flagged as 'not_found'."""
    order = db.query(Order).options(
        joinedload(Order.items)
    ).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if order.status not in ITEM_EDITABLE_STAGES:
        raise HTTPException(status_code=400, detail="Prices can only be modified up to Factory Ordered stage or in Completed Editing")

    if not order.factory_id:
        raise HTTPException(status_code=400, detail="Factory must be set to look up previous prices")

    active_items = [i for i in order.items if i.status == "ACTIVE"]
    if not active_items:
        return {"copied": 0, "total": 0, "not_found": [], "price_sources": {}, "message": "No active items"}

    # Get IDs of all past same-factory orders (not current, not draft, not deleted)
    past_order_ids = [
        row[0] for row in db.query(Order.id).filter(
            Order.factory_id == order.factory_id,
            Order.id != order.id,
            Order.deleted_at.is_(None),
            Order.status != OrderStatus.DRAFT.value,
        ).all()
    ]

    if not past_order_ids:
        not_found_codes = [i.product_code_snapshot or "?" for i in active_items
                           if not i.factory_price and not (i.notes and i.notes.startswith("After-Sales"))]
        return {
            "copied": 0, "total": len(active_items),
            "not_found": not_found_codes, "price_sources": {},
            "message": "No previous orders found for this factory",
        }

    # For each product_id, find the most recent OrderItem with a factory price
    # across ALL same-factory orders
    product_ids = [i.product_id for i in active_items if i.product_id]

    # Single query: get the most recent priced item per product across all past orders
    # Use a window function approach: rank by order created_at desc per product_id
    past_items = db.query(OrderItem, Order.order_number, Order.created_at).join(
        Order, OrderItem.order_id == Order.id
    ).filter(
        OrderItem.order_id.in_(past_order_ids),
        OrderItem.product_id.in_(product_ids),
        OrderItem.factory_price.isnot(None),
        OrderItem.factory_price > 0,  # Exclude after-sales items with factory_price = 0
    ).order_by(Order.created_at.desc()).all()

    # Build lookup: product_id -> (most recent item, order_number)
    # First match per product_id wins (ordered by most recent)
    best_price_map = {}  # product_id -> (OrderItem, order_number)
    for prev_item, order_num, _ in past_items:
        if prev_item.product_id not in best_price_map:
            best_price_map[prev_item.product_id] = (prev_item, order_num)

    copied = 0
    not_found = []
    price_sources = {}

    for item in active_items:
        # Skip after-sales carry-forward items — they have fixed pricing (0 or negative)
        if item.notes and item.notes.startswith("After-Sales"):
            continue
        # Only copy if current item has no factory price yet
        if item.factory_price:
            continue

        best = best_price_map.get(item.product_id)
        if best:
            prev_item, source_order = best
            item.factory_price = prev_item.factory_price
            item.markup_percent = prev_item.markup_percent
            item.selling_price = prev_item.selling_price
            item.selling_price_inr = prev_item.selling_price_inr
            # Auto-calc client_factory_price for transparency clients
            _client = db.query(Client).filter(Client.id == order.client_id).first()
            _calc_client_factory_price(item, _client)
            copied += 1
            code = item.product_code_snapshot or item.product_id[:8]
            price_sources[code] = source_order or "?"
        else:
            code = item.product_code_snapshot or item.product_id[:8]
            not_found.append(code)

    db.commit()

    return {
        "copied": copied,
        "total": len(active_items),
        "not_found": not_found,
        "price_sources": price_sources,
        "message": f"Copied prices for {copied} of {len(active_items)} items" + (
            f". {len(not_found)} items have no past order history." if not_found else ""
        ),
    }


@router.post("/{order_id}/reset-aftersales-prices/")
def reset_aftersales_prices(order_id: str, db: Session = Depends(get_db)):
    """Reset after-sales carry-forward items to their correct fixed prices.
    Replace → factory=0, selling=0; Compensate → factory=0, selling=-(original selling_price_inr)."""
    from models import AfterSalesItem
    order = db.query(Order).options(joinedload(Order.items)).filter(
        Order.id == order_id, Order.deleted_at.is_(None)
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    fixed = 0
    for item in order.items:
        if item.status != "ACTIVE" or not item.notes or not item.notes.startswith("After-Sales"):
            continue
        # Look up the original AfterSalesItem claim
        claim = db.query(AfterSalesItem).filter(
            AfterSalesItem.added_to_order_id == order.id,
            AfterSalesItem.product_id == item.product_id,
        ).first()
        if not claim:
            # Fallback: just reset to 0
            item.factory_price = 0
            item.markup_percent = None
            item.selling_price = None
            item.selling_price_inr = 0
            fixed += 1
            continue

        from enums import CarryForwardType
        if claim.carry_forward_type == CarryForwardType.REPLACEMENT.value:
            item.factory_price = 0
            item.selling_price_inr = 0
        else:
            item.factory_price = 0
            item.selling_price_inr = -(claim.selling_price_inr or 0)
        item.markup_percent = None
        item.selling_price = None
        fixed += 1

    db.commit()
    return {"fixed": fixed, "message": f"Reset {fixed} after-sales item(s) to correct pricing"}


@router.post("/{order_id}/parse-price-excel/")
async def parse_price_excel(
    order_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Parse an Excel file to extract product codes + prices.
    Returns matched/unmatched lists for the frontend to preview before applying."""
    import openpyxl
    from io import BytesIO

    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Stream to temp file, enforce size limit (NamedTemporaryFile avoids mktemp race)
    import tempfile
    from pathlib import Path as _Path
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_fh:
        tmp_path = _Path(tmp_fh.name)
    try:
        await stream_upload_to_disk(file, tmp_path, MAX_UPLOAD_SIZE)
        raw = tmp_path.read_bytes()
    finally:
        tmp_path.unlink(missing_ok=True)

    try:
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read Excel file")
    ws = wb.active

    # Extract rows: look for columns with part code + price
    # Strategy: scan header row to find code and price columns
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty Excel file")

    # Auto-detect columns: find header row with "part" or "code" and "price" or "unit"
    header_row_idx = 0
    code_col = None
    price_col = None
    part_no_cols = []  # Collect all part number columns
    price_cols = []
    for ri, row in enumerate(rows[:5]):  # Check first 5 rows for header
        for ci, cell in enumerate(row or []):
            val = str(cell or "").strip().lower()
            if any(kw in val for kw in ["part no", "part code", "manufacturer", "code", "mfr"]):
                part_no_cols.append((ci, val, ri))
            if any(kw in val for kw in ["price", "unit price", "单价"]):
                price_cols.append(ci)
                header_row_idx = ri

    # Prefer MFR Part No. column; otherwise use second Part No column (factory: first=barcode, second=MFR)
    if part_no_cols:
        mfr_col = next((c for c in part_no_cols if "mfr" in c[1] or "manufacturer" in c[1]), None)
        if mfr_col:
            code_col, _, header_row_idx = mfr_col
        elif len(part_no_cols) >= 2:
            code_col, _, header_row_idx = part_no_cols[1]
        else:
            code_col, _, header_row_idx = part_no_cols[0]
    # Use LAST price column (factory format has summary price early, actual price later)
    if price_cols:
        price_col = price_cols[-1]

    # Fallback: if factory format (col 3 = MFR Part No, col 8 = UNIT PRICE)
    if code_col is None and len(rows[0] or []) >= 8:
        code_col = 3  # MFR Part No. column
        price_col = 8  # UNIT PRICE column
        header_row_idx = 0

    if code_col is None or price_col is None:
        raise HTTPException(
            status_code=400,
            detail="Could not find part code and price columns. Expected headers like 'Part No.' and 'UNIT PRICE'.",
        )

    # Parse data rows
    price_entries = []
    for row in rows[header_row_idx + 1:]:
        if not row or len(row) <= max(code_col, price_col):
            continue
        code = str(row[code_col] or "").strip()
        price_raw = row[price_col]
        if not code:
            continue
        try:
            price = float(price_raw) if price_raw else None
        except (ValueError, TypeError):
            price = None
        if price is not None and price > 0:
            price_entries.append({"code": code, "price": price})

    return {"entries": price_entries, "total_parsed": len(price_entries)}


# ========================================
# 2B: STATUS ENGINE — Stage Transitions
# (Business logic lives in services/stage_engine.py)
# ========================================



@router.get("/{order_id}/next-stages/")
def get_next_stages(order_id: str, db: Session = Depends(get_db)):
    """Get available next and previous stages for an order."""
    order = db.query(Order).filter(
        Order.id == order_id, Order.deleted_at.is_(None)
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return stage_engine.get_next_stages_info(order)


@router.put("/{order_id}/transition/")
def transition_order(
    order_id: str,
    target_status: str,
    req: StageTransitionRequest,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Move order to next stage."""
    order = db.query(Order).options(
        joinedload(Order.items)
    ).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    try:
        result = stage_engine.execute_transition(
            db, order, target_status, current_user,
            notes=req.transition_reason,
            acknowledge_warnings=req.acknowledge_warnings,
            transition_reason=req.transition_reason,
            generate_order_number_fn=generate_order_number,
            copy_image_fn=_copy_image_to_order,
        )
    except ValueError as e:
        detail = e.args[0] if e.args else str(e)
        raise HTTPException(status_code=400, detail=detail)

    # Warnings need user acknowledgement — return as-is
    if result.get("status") == "warnings":
        return JSONResponse(status_code=200, content=result)

    # Re-fetch with relationships for serialization
    order = db.query(Order).options(
        joinedload(Order.client),
        joinedload(Order.factory),
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.id == order.id).first()

    return serialize_order(order, db)


@router.put("/{order_id}/reopen/")
def reopen_order(
    order_id: str,
    req: OrderReopenRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Re-open a completed order for editing — admin only."""
    if current_user.role not in ("ADMIN", "SUPER_ADMIN"):
        raise HTTPException(status_code=403, detail="Only admin can reopen orders")
    order = db.query(Order).filter(
        Order.id == order_id, Order.deleted_at.is_(None)
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    try:
        stage_engine.execute_reopen(db, order, req.reason)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Re-fetch with relationships for serialization
    order = db.query(Order).options(
        joinedload(Order.client),
        joinedload(Order.factory),
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.id == order.id).first()

    return serialize_order(order, db)


@router.put("/{order_id}/go-back/")
def go_back_order(
    order_id: str,
    req: GoBackRequest = None,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Move order back to its previous status."""
    order = db.query(Order).filter(
        Order.id == order_id, Order.deleted_at.is_(None)
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    try:
        stage_engine.execute_go_back(
            db, order, current_user,
            reason=req.reason if req and hasattr(req, 'reason') else None,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Re-fetch with relationships for serialization
    order = db.query(Order).options(
        joinedload(Order.client),
        joinedload(Order.factory),
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.id == order.id).first()

    return serialize_order(order, db)


@router.put("/{order_id}/jump-to-stage/")
def jump_to_stage(
    order_id: str,
    req: JumpToStageRequest,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Jump order directly to a reachable stage."""
    order = db.query(Order).filter(
        Order.id == order_id, Order.deleted_at.is_(None)
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    try:
        stage_engine.execute_jump(
            db, order, req.target_status, current_user,
            reason=req.reason if hasattr(req, 'reason') else None,
        )
    except ValueError as e:
        detail = e.args[0] if e.args else str(e)
        raise HTTPException(status_code=400, detail=detail)

    # Re-fetch with relationships for serialization
    order = db.query(Order).options(
        joinedload(Order.client),
        joinedload(Order.factory),
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.id == order.id).first()

    return serialize_order(order, db)


@router.get("/{order_id}/activity-feed/")
def get_activity_feed(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Client-friendly activity feed built from audit logs.
    Translates raw audit actions into human-readable timeline events."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    is_client = current_user.user_type == "CLIENT"

    from models import AuditLog

    # Fetch audit logs for this order
    logs = db.query(AuditLog).filter(
        AuditLog.resource_id == order_id,
    ).order_by(AuditLog.timestamp.asc()).all()

    # Also fetch payment logs
    payment_logs = db.query(AuditLog).filter(
        AuditLog.resource_type == "payment",
        AuditLog.metadata_json.contains(order_id),
    ).order_by(AuditLog.timestamp.asc()).all()

    import json

    # Translation map: audit action → client-friendly message
    def translate_event(log):
        action = log.action or ""
        actor = "Our team" if log.user_email and "admin" in log.user_email.lower() else "You"
        if is_client and "admin" in (log.user_email or "").lower():
            actor = "Our team"
        elif is_client:
            actor = "You"
        else:
            actor = log.user_email or "System"

        new_vals = {}
        try:
            new_vals = json.loads(log.new_values) if log.new_values else {}
        except (json.JSONDecodeError, TypeError):
            pass

        old_vals = {}
        try:
            old_vals = json.loads(log.old_values) if log.old_values else {}
        except (json.JSONDecodeError, TypeError):
            pass

        icon = "pi-info-circle"
        color = "slate"

        if action == "INQUIRY_APPROVED":
            icon = "pi-check-circle"
            color = "emerald"
            return icon, color, f"{actor} approved your inquiry and created order {order.order_number}"

        elif action == "ORDER_STAGE_CHANGE":
            new_status = new_vals.get("status", "")
            old_status = old_vals.get("status", "")
            stage_labels = {
                "DRAFT": "started processing your order",
                "PENDING_PI": "began preparing your quotation",
                "PI_SENT": f"sent you the Proforma Invoice",
                "ADVANCE_PENDING": "is awaiting your advance payment",
                "ADVANCE_RECEIVED": "confirmed your advance payment",
                "FACTORY_ORDERED": "placed your order with the manufacturer",
                "PRODUCTION_60": "reports manufacturing is 60% complete",
                "PRODUCTION_80": "reports manufacturing is 80% complete",
                "PRODUCTION_90": "reports manufacturing is 90% complete",
                "PLAN_PACKING": "is preparing your shipment packing",
                "FINAL_PI": "issued the final revised invoice",
                "PRODUCTION_100": "reports manufacturing is complete",
                "BOOKED": "booked a shipping container for your order",
                "LOADED": "loaded your goods onto the vessel",
                "SAILED": "confirms your shipment has departed",
                "ARRIVED": "confirms your shipment has arrived at port",
                "CUSTOMS_FILED": "filed customs declaration for your shipment",
                "CLEARED": "cleared your shipment through customs",
                "DELIVERED": "delivered your order",
                "AFTER_SALES": "opened after-sales review for your order",
                "COMPLETED": "marked your order as completed",
            }
            msg = stage_labels.get(new_status, f"moved order to {new_status}")
            if new_status in ("PI_SENT", "FINAL_PI"):
                icon = "pi-file"
                color = "indigo"
            elif new_status in ("ADVANCE_RECEIVED",):
                icon = "pi-wallet"
                color = "emerald"
            elif new_status in ("FACTORY_ORDERED",):
                icon = "pi-building"
                color = "violet"
            elif "PRODUCTION" in new_status:
                icon = "pi-cog"
                color = "orange"
            elif new_status in ("BOOKED", "LOADED", "SAILED", "ARRIVED"):
                icon = "pi-send"
                color = "blue"
            elif new_status in ("CUSTOMS_FILED", "CLEARED"):
                icon = "pi-file-check"
                color = "amber"
            elif new_status == "DELIVERED":
                icon = "pi-check-circle"
                color = "emerald"
            elif new_status == "COMPLETED":
                icon = "pi-star"
                color = "emerald"
            else:
                icon = "pi-arrow-right"
                color = "slate"
            return icon, color, f"{actor} {msg}"

        elif action == "PAYMENT_SUBMITTED" and log.resource_type == "payment":
            amount = new_vals.get("amount_inr", "")
            method = (new_vals.get("method", "") or "").replace("_", " ")
            icon = "pi-upload"
            color = "amber"
            amt_str = f" of ₹{float(amount):,.2f}" if amount else ""
            return icon, color, f"{actor} submitted a payment proof{amt_str} via {method}" if method else f"{actor} submitted a payment proof{amt_str}"

        elif action == "PAYMENT_APPROVED" and log.resource_type == "payment":
            amount = new_vals.get("amount_inr", "")
            icon = "pi-check-circle"
            color = "emerald"
            amt_str = f" of ₹{float(amount):,.2f}" if amount else ""
            return icon, color, f"{actor} verified your payment{amt_str}"

        elif action == "PAYMENT_REJECTED" and log.resource_type == "payment":
            amount = new_vals.get("amount_inr", "")
            reason = new_vals.get("rejection_reason", "")
            icon = "pi-times-circle"
            color = "red"
            amt_str = f" of ₹{float(amount):,.2f}" if amount else ""
            reason_str = f" — {reason}" if reason else ""
            return icon, color, f"{actor} rejected your payment{amt_str}{reason_str}"

        elif action == "CREATE" and log.resource_type == "payment":
            amount = new_vals.get("amount_inr", "")
            method = new_vals.get("method", "")
            icon = "pi-wallet"
            color = "emerald"
            amt_str = f" of ₹{float(amount):,.2f}" if amount else ""
            return icon, color, f"{actor} recorded a payment{amt_str} via {method}" if method else f"{actor} recorded a payment{amt_str}"

        elif action == "PACKING_MIGRATION":
            count = new_vals.get("count", "")
            icon = "pi-arrow-right"
            color = "amber"
            return icon, color, f"{count or 'Some'} items were migrated to next order (carry-forward)"

        elif action == "AFTER_SALES_SUBMIT":
            count = new_vals.get("claims_count", "")
            icon = "pi-exclamation-triangle"
            color = "amber"
            return icon, color, f"{actor} reported {count or ''} quality issues on delivered goods"

        elif action == "AFTER_SALES_RESOLVE":
            resolution = new_vals.get("resolution_type", "")
            icon = "pi-check"
            color = "emerald"
            return icon, color, f"{actor} resolved claim: {resolution.replace('_', ' ').title()}" if resolution else f"{actor} resolved a claim"

        elif "PI_GENERATED" in action or "PI_REGENERATED" in action:
            icon = "pi-file"
            color = "indigo"
            return icon, color, f"{actor} generated the Proforma Invoice"

        elif action == "MID_ORDER_ITEMS_ADDED":
            count = new_vals.get("count", "")
            codes = new_vals.get("product_codes", [])
            icon = "pi-plus-circle"
            color = "amber"
            codes_str = ", ".join(codes[:3]) + ("..." if len(codes) > 3 else "") if codes else ""
            return icon, color, f"{actor} added {count or 'new'} items mid-order ({codes_str})" if codes_str else f"{actor} added {count or 'new'} items mid-order"

        elif action == "MID_ORDER_ITEM_PRICED":
            code = new_vals.get("product_code", "")
            icon = "pi-tag"
            color = "indigo"
            return icon, color, f"{actor} set pricing for {code}"

        elif action == "PRICES_SENT_TO_CLIENT":
            count = new_vals.get("count", "")
            icon = "pi-send"
            color = "indigo"
            return icon, color, f"{actor} sent prices for {count} new items — awaiting your confirmation"

        elif action == "MID_ORDER_ITEMS_CLIENT_CONFIRMED":
            count = new_vals.get("count", "")
            icon = "pi-thumbs-up"
            color = "indigo"
            return icon, color, f"{actor} accepted prices for {count} items — awaiting final approval"

        elif action == "MID_ORDER_ITEMS_APPROVED":
            count = new_vals.get("count", "")
            lot = new_vals.get("lot", "")
            codes = new_vals.get("product_codes", [])
            icon = "pi-check-circle"
            color = "emerald"
            lot_str = f" (Lot {lot})" if lot else ""
            return icon, color, f"{actor} approved {count} items{lot_str} — added to order"

        elif action == "MID_ORDER_ITEMS_REJECTED":
            code = new_vals.get("product_code", "")
            icon = "pi-times-circle"
            color = "red"
            return icon, color, f"{actor} rejected mid-order item {code}"

        else:
            # Generic fallback
            readable = action.replace("_", " ").title()
            return icon, color, f"{readable}"

    events = []
    seen_actions = set()

    for log in logs + payment_logs:
        # Deduplicate
        key = f"{log.action}_{log.timestamp}"
        if key in seen_actions:
            continue
        seen_actions.add(key)

        icon, color, message = translate_event(log)
        events.append({
            "timestamp": log.timestamp.isoformat() if log.timestamp else None,
            "message": message,
            "icon": icon,
            "color": color,
            "action": log.action,
        })

    # Sort by timestamp descending (newest first)
    events.sort(key=lambda e: e["timestamp"] or "", reverse=True)

    return {"events": events, "total": len(events)}


@router.get("/{order_id}/timeline/")
def get_order_timeline(order_id: str, db: Session = Depends(get_db)):
    """Get order timeline — all status changes would be tracked here.
    For now, return computed stage progression based on current status."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    current_stage, current_name = get_stage_info(order.status)
    highest_stage = get_stage_info(order.highest_unlocked_stage)[0] if order.highest_unlocked_stage else current_stage

    # Build stage list showing which stages are complete vs current vs unlocked vs pending
    # Must match STAGE_MAP numbering (17 stages, some share a number)
    stages = [
        (1, "Draft"), (2, "Pending PI"), (3, "PI Sent"),
        (4, "Payment"), (5, "Factory Ordered"),
        (6, "Production 60%"), (7, "Production 80%"),
        (8, "Production 90%"), (9, "Plan Packing"),
        (10, "Final PI"), (11, "Production 100%"),
        (12, "Booked"), (13, "Sailing"),
        (14, "Customs"), (15, "Delivered"),
        (16, "After-Sales"), (17, "Completed"),
    ]

    timeline = []
    for num, name in stages:
        if num < current_stage:
            status = "completed"
        elif num == current_stage:
            status = "current"
        elif num <= highest_stage:
            status = "unlocked"
        else:
            status = "pending"
        timeline.append({"stage": num, "name": name, "status": status})

    # Fetch stage overrides for this order
    overrides = db.query(StageOverride).filter(
        StageOverride.order_id == order_id
    ).order_by(StageOverride.created_at).all()

    override_list = []
    for ov in overrides:
        to_num, to_name = get_stage_info(ov.to_stage)
        override_list.append({
            "id": ov.id,
            "from_stage": ov.from_stage,
            "to_stage": ov.to_stage,
            "to_stage_number": to_num,
            "to_stage_name": to_name,
            "reason": ov.reason,
            "warnings": json_mod.loads(ov.warnings_json) if ov.warnings_json else [],
            "created_at": ov.created_at.isoformat() if ov.created_at else None,
        })

    return {
        "current_status": order.status,
        "current_stage": current_stage,
        "current_name": current_name,
        "timeline": timeline,
        "overrides": override_list,
        "created_at": order.created_at.isoformat() if order.created_at else None,
        "completed_at": order.completed_at.isoformat() if order.completed_at else None,
    }


# ========================================
# Packing List Management (Level 5B)
# ========================================

@router.post("/{order_id}/packing-list/upload/")
async def upload_packing_list(
    order_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Upload and parse a packing list Excel for the order."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if order.status != OrderStatus.PLAN_PACKING.value:
        raise HTTPException(status_code=400, detail="Packing list can only be uploaded at Plan Packing stage")

    # Save file (shared utility handles chunking + size validation + cleanup)
    temp_filename = f"packing_{uuid.uuid4().hex[:8]}.xlsx"
    save_path = UPLOAD_DIR / "orders" / order_id / temp_filename
    await stream_upload_to_disk(file, save_path, MAX_UPLOAD_SIZE)

    # Parse Excel
    wb = openpyxl.load_workbook(str(save_path), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not all_rows:
        raise HTTPException(status_code=400, detail="Empty Excel file")

    # Auto-detect columns from header
    header = all_rows[0]
    col_part_code = None
    col_qty = None
    col_package = None
    col_description = None

    for ci, cell in enumerate(header or []):
        val = str(cell or "").strip().lower()
        if "part no" in val or "part code" in val or "mfr" in val or "manufacturer" in val:
            col_part_code = ci
        elif val in ("quantity", "qty"):
            col_qty = ci
        elif any(kw in val for kw in ["pallet", "package", "packing", "pkg", "carton", "箱号"]):
            col_package = ci
        elif "description" in val or "goods" in val:
            col_description = ci

    if col_part_code is None:
        raise HTTPException(status_code=400, detail="Could not detect part code column in header row")

    # Get order items for matching
    order_items = db.query(OrderItem).filter(
        OrderItem.order_id == order_id,
        OrderItem.status == OrderItemStatus.ACTIVE.value,
    ).all()

    # Build lookup by product code (case-insensitive) → list of OIs
    # Multiple order items can share a code (e.g. compensation + regular)
    items_by_code: dict[str, list] = {}
    for oi in order_items:
        code = oi.product_code_snapshot
        if not code:
            product = db.query(Product).filter(Product.id == oi.product_id).first()
            code = product.product_code if product else None
        if code:
            items_by_code.setdefault(code.strip().upper(), []).append(oi)

    # Delete existing packing list if re-uploading
    existing_pl = db.query(PackingList).filter(PackingList.order_id == order_id).first()
    if existing_pl:
        db.query(PackingListItem).filter(PackingListItem.packing_list_id == existing_pl.id).delete()
        db.delete(existing_pl)
        db.flush()

    packing_list = PackingList(
        order_id=order_id,
        uploaded_date=date.today(),
        file_path=f"orders/{order_id}/{temp_filename}",
    )
    db.add(packing_list)
    db.flush()

    # Parse data rows
    matched = 0
    unmatched = 0
    total_packages = set()
    warnings = []
    matched_oi_ids = set()  # Track which OIs got at least one packing row

    for row_data in all_rows[1:]:
        if not row_data or not row_data[col_part_code]:
            continue

        part_code = str(row_data[col_part_code]).strip().upper()
        qty = 0
        if col_qty is not None and row_data[col_qty]:
            try:
                qty = int(float(str(row_data[col_qty])))
            except (ValueError, TypeError):
                qty = 0

        package_raw = ""
        if col_package is not None and row_data[col_package]:
            package_raw = str(row_data[col_package]).strip()

        # Determine packing status
        package_number = None
        packing_status = "NOT_READY"
        if package_raw:
            lower = package_raw.lower()
            if lower in ("bulk", "loose", "散装"):
                package_number = "BULK"
                packing_status = "LOOSE"
            else:
                package_number = package_raw
                packing_status = "PALLETED"
                total_packages.add(package_raw)

        # Match to order item — prefer an OI not yet matched (handles
        # multiple OIs with the same code, e.g. compensation + regular).
        # If all OIs for this code are already matched, reuse the first
        # one (valid for split-packing across pallets).
        candidates = items_by_code.get(part_code, [])
        oi = None
        for c in candidates:
            if c.id not in matched_oi_ids:
                oi = c
                break
        if oi is None and candidates:
            oi = candidates[0]  # split-packing: reuse first OI

        if oi:
            # Validate: factory qty should not exceed ordered qty
            if qty > oi.quantity:
                warnings.append(f"{part_code}: factory qty ({qty}) exceeds ordered qty ({oi.quantity})")

            matched_oi_ids.add(oi.id)
            db.add(PackingListItem(
                packing_list_id=packing_list.id,
                order_item_id=oi.id,
                product_id=oi.product_id,
                ordered_qty=oi.quantity,
                factory_ready_qty=qty,
                loaded_qty=min(qty, oi.quantity) if packing_status != "NOT_READY" else 0,
                package_number=package_number,
                packing_status=packing_status,
            ))
            matched += 1
        else:
            unmatched += 1

    # Create NOT_READY placeholders for ACTIVE order items not in factory Excel
    not_in_excel = 0
    for oi in order_items:
        if oi.id not in matched_oi_ids:
            code = oi.product_code_snapshot or "?"
            db.add(PackingListItem(
                packing_list_id=packing_list.id,
                order_item_id=oi.id,
                product_id=oi.product_id,
                ordered_qty=oi.quantity,
                factory_ready_qty=0,
                loaded_qty=0,
                package_number=None,
                packing_status="NOT_READY",
            ))
            not_in_excel += 1
            warnings.append(f"{code}: not in factory Excel — added as NOT_READY")

    packing_list.total_packages = len(total_packages)
    db.commit()

    return {
        "packing_list_id": packing_list.id,
        "matched_items": matched,
        "unmatched_items": unmatched,
        "not_in_excel": not_in_excel,
        "total_packages": len(total_packages),
        "warnings": warnings,
    }


def _classify_packing_item(oi) -> str:
    """Classify a packing list item for frontend color coding."""
    if not oi:
        return "regular"
    if (oi.selling_price_inr or 0) < 0:
        return "compensation"
    if oi.notes and oi.notes.startswith("After-Sales") and (oi.selling_price_inr is None or oi.selling_price_inr == 0):
        return "aftersales_replacement"
    if oi.notes and oi.notes.startswith("Carried from"):
        return "carried_forward"
    return "regular"


@router.post("/{order_id}/packing-list/manual/")
def create_manual_packing_list(
    order_id: str,
    data: ManualPackingListRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Create packing list manually (alternative to Excel upload)."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.status != OrderStatus.PLAN_PACKING.value:
        raise HTTPException(status_code=400, detail="Order must be at Plan Packing stage")

    # Delete existing packing list if present (re-create)
    existing_pl = db.query(PackingList).filter(PackingList.order_id == order_id).first()
    if existing_pl:
        db.query(PackingListItem).filter(PackingListItem.packing_list_id == existing_pl.id).delete()
        db.delete(existing_pl)
        db.flush()

    # Create new PackingList
    pl = PackingList(
        order_id=order_id,
        uploaded_date=date.today(),
        file_path=None,  # Manual — no file
    )
    db.add(pl)
    db.flush()

    # Build lookup of ACTIVE order items
    active_ois = db.query(OrderItem).filter(
        OrderItem.order_id == order_id,
        OrderItem.status == OrderItemStatus.ACTIVE.value,
    ).all()
    oi_map = {oi.id: oi for oi in active_ois}
    covered_oi_ids = set()

    unique_packages = set()
    created_count = 0

    for mi in data.items:
        oi = oi_map.get(mi.order_item_id)
        if not oi:
            continue  # Skip invalid items

        covered_oi_ids.add(oi.id)

        # Determine packing status from package_number
        pkg = (mi.package_number or "").strip()
        if not pkg:
            packing_status = "NOT_READY"
            loaded_qty = 0
        elif pkg.upper() in ("BULK", "LOOSE", "散装"):
            packing_status = "LOOSE"
            pkg = "BULK"
            loaded_qty = min(mi.factory_ready_qty, oi.quantity)
        else:
            packing_status = "PALLETED"
            loaded_qty = min(mi.factory_ready_qty, oi.quantity)
            unique_packages.add(pkg)

        pli = PackingListItem(
            packing_list_id=pl.id,
            order_item_id=oi.id,
            product_id=oi.product_id,
            ordered_qty=oi.quantity,
            factory_ready_qty=mi.factory_ready_qty,
            loaded_qty=loaded_qty,
            unloaded_qty=0,
            package_number=pkg or None,
            packing_status=packing_status,
        )
        db.add(pli)
        created_count += 1

    # Create NOT_READY placeholders for ACTIVE items not in request
    for oi in active_ois:
        if oi.id not in covered_oi_ids:
            pli = PackingListItem(
                packing_list_id=pl.id,
                order_item_id=oi.id,
                product_id=oi.product_id,
                ordered_qty=oi.quantity,
                factory_ready_qty=0,
                loaded_qty=0,
                unloaded_qty=0,
                packing_status="NOT_READY",
            )
            db.add(pli)

    pl.total_packages = len(unique_packages)
    db.commit()

    return {
        "packing_list_id": pl.id,
        "created_items": created_count,
        "total_packages": len(unique_packages),
    }


@router.get("/{order_id}/packing-list/client-summary/")
def get_packing_client_summary(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Client-safe read-only packing summary. No pallet numbers, no factory data."""
    if not current_user.client_id:
        raise HTTPException(status_code=403, detail="Not a client user")

    # Check permission
    client = db.query(Client).filter(Client.id == current_user.client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    perms = client.portal_permissions or {}
    if not perms.get("show_packing", False):
        raise HTTPException(status_code=403, detail="Packing access not enabled")

    # Verify order belongs to this client
    order = db.query(Order).filter(
        Order.id == order_id, Order.deleted_at.is_(None),
        Order.client_id == current_user.client_id,
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    pl = db.query(PackingList).filter(PackingList.order_id == order_id).first()
    if not pl:
        return {"summary": None, "items": [], "carried_forward": []}

    plis = db.query(PackingListItem).filter(PackingListItem.packing_list_id == pl.id).all()

    items_out = []
    carried_forward = []
    total_ordered = 0
    total_ready = 0
    loaded_count = 0
    not_ready_count = 0
    migrated_count = 0

    for pli in plis:
        oi = db.query(OrderItem).filter(OrderItem.id == pli.order_item_id).first()
        if not oi:
            continue

        # Skip balance-only items (compensation — internal)
        if (oi.selling_price_inr or 0) < 0:
            continue

        code = oi.product_code_snapshot or ""
        name = oi.product_name_snapshot or ""

        if oi.status == OrderItemStatus.UNLOADED.value:
            migrated_count += 1
            carried_forward.append({
                "product_code": code,
                "product_name": name,
                "ordered_qty": pli.ordered_qty,
                "reason": pli.shortage_reason or "NOT_PRODUCED",
            })
        else:
            total_ordered += pli.ordered_qty or 0
            total_ready += pli.factory_ready_qty or 0

            if pli.packing_status in ("PALLETED", "LOOSE"):
                status = "LOADED"
                loaded_count += 1
            elif pli.packing_status == "NOT_READY":
                status = "NOT_READY"
                not_ready_count += 1
            else:
                status = "NOT_READY"
                not_ready_count += 1

            items_out.append({
                "product_code": code,
                "product_name": name,
                "ordered_qty": pli.ordered_qty,
                "ready_qty": pli.factory_ready_qty,
                "status": status,
                "carry_forward_reason": None,
            })

    # Include pending carry-forward items (from SHIP_CARRY_FORWARD decisions)
    pending_carry = db.query(UnloadedItem).filter(
        UnloadedItem.original_order_id == order_id,
        UnloadedItem.status == UnloadedItemStatus.PENDING.value,
    ).all()
    for ui in pending_carry:
        # Skip if already shown via UNLOADED order item
        already = any(cf["product_code"] == (db.query(Product).filter(Product.id == ui.product_id).first().product_code if db.query(Product).filter(Product.id == ui.product_id).first() else "") for cf in carried_forward)
        if already:
            continue
        product = db.query(Product).filter(Product.id == ui.product_id).first()
        carried_forward.append({
            "product_code": product.product_code if product else "",
            "product_name": product.product_name if product else "",
            "ordered_qty": ui.quantity,
            "reason": ui.reason or "NOT_PRODUCED",
        })
        migrated_count += 1

    produced_pct = round((total_ready / total_ordered * 100)) if total_ordered > 0 else 0

    return {
        "summary": {
            "total_items": len(items_out),
            "produced_count": sum(1 for i in items_out if (i["ready_qty"] or 0) > 0),
            "loaded_count": loaded_count,
            "not_ready_count": not_ready_count,
            "migrated_count": migrated_count,
            "total_ordered_qty": total_ordered,
            "total_ready_qty": total_ready,
            "produced_percent": produced_pct,
        },
        "items": items_out,
        "carried_forward": carried_forward,
    }


@router.get("/{order_id}/packing-list/")
def get_packing_list(order_id: str, db: Session = Depends(get_db)):
    """Get packing list with all items for an order."""
    pl = db.query(PackingList).filter(PackingList.order_id == order_id).first()
    if not pl:
        return {"packing_list": None, "items": []}

    items = db.query(PackingListItem).filter(PackingListItem.packing_list_id == pl.id).all()
    result_items = []
    for item in items:
        product = db.query(Product).filter(Product.id == item.product_id).first()
        oi = db.query(OrderItem).filter(OrderItem.id == item.order_item_id).first()
        result_items.append({
            "id": item.id,
            "order_item_id": item.order_item_id,
            "product_id": item.product_id,
            "product_code": oi.product_code_snapshot or (product.product_code if product else ""),
            "product_name": oi.product_name_snapshot or (product.product_name if product else ""),
            "ordered_qty": item.ordered_qty,
            "factory_ready_qty": item.factory_ready_qty,
            "loaded_qty": item.loaded_qty,
            "unloaded_qty": item.unloaded_qty,
            "package_number": item.package_number,
            "packing_status": item.packing_status,
            "shortage_reason": item.shortage_reason,
            "factory_price": oi.factory_price if oi else None,
            "order_item_status": oi.status if oi else None,
            "cancel_note": oi.cancel_note if oi else None,
            "selling_price_inr": oi.selling_price_inr if oi else None,
            "is_balance_only": (oi.selling_price_inr or 0) < 0 if oi else False,
            "notes": oi.notes if oi else None,
            "item_type": _classify_packing_item(oi),
            "parent_packing_item_id": item.parent_packing_item_id,
            "is_split": item.is_split or False,
            "split_qty": item.split_qty,
            "shipping_decision": item.shipping_decision,
            "cancel_reason": item.cancel_reason,
        })

    # Include pending carry-forward items (from SHIP_CARRY_FORWARD decisions)
    pending_carry = db.query(UnloadedItem).filter(
        UnloadedItem.original_order_id == order_id,
        UnloadedItem.status == UnloadedItemStatus.PENDING.value,
    ).all()
    carry_forward_items = []
    for ui in pending_carry:
        # Skip if already represented by an UNLOADED order item in the packing list
        already_shown = any(
            ri["product_id"] == ui.product_id and ri["order_item_status"] == "UNLOADED"
            for ri in result_items
        )
        if already_shown:
            continue
        product = db.query(Product).filter(Product.id == ui.product_id).first()
        carry_forward_items.append({
            "id": ui.id,
            "product_id": ui.product_id,
            "product_code": product.product_code if product else "",
            "product_name": product.product_name if product else "",
            "quantity": ui.quantity,
            "reason": ui.reason,
            "status": ui.status,
        })

    return {
        "packing_list": {
            "id": pl.id,
            "uploaded_date": str(pl.uploaded_date),
            "total_packages": pl.total_packages,
            "total_gross_weight": pl.total_gross_weight,
            "total_net_weight": pl.total_net_weight,
            "total_cbm": pl.total_cbm,
        },
        "items": result_items,
        "carry_forward_items": carry_forward_items,
    }


@router.delete("/{order_id}/packing-list/")
def delete_packing_list(order_id: str, db: Session = Depends(get_db)):
    """Delete the packing list and all its items for re-upload."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Stage guard: cannot delete packing list after BOOKED (shipping depends on it)
    _blocked_stages = [
        OrderStatus.BOOKED.value, OrderStatus.LOADED.value, OrderStatus.SAILED.value,
        OrderStatus.ARRIVED.value, OrderStatus.CUSTOMS_FILED.value, OrderStatus.CLEARED.value,
        OrderStatus.DELIVERED.value, OrderStatus.AFTER_SALES.value, OrderStatus.COMPLETED.value,
    ]
    if order.status in _blocked_stages:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete packing list at stage '{order.status}'. Shipping allocations depend on it."
        )

    pl = db.query(PackingList).filter(PackingList.order_id == order_id).first()
    if not pl:
        raise HTTPException(status_code=404, detail="No packing list found")
    db.query(PackingListItem).filter(PackingListItem.packing_list_id == pl.id).delete()
    db.delete(pl)
    db.commit()
    return {"deleted": True}


@router.patch("/{order_id}/packing-list/items/{item_id}/")
def update_packing_item(order_id: str, item_id: str, data: UpdatePackingItemRequest, db: Session = Depends(get_db)):
    """Update a single packing list item (e.g., pallet assignment)."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.status != OrderStatus.PLAN_PACKING.value:
        raise HTTPException(status_code=400, detail="Packing edits only allowed at Plan Packing stage")

    pl = db.query(PackingList).filter(PackingList.order_id == order_id).first()
    if not pl:
        raise HTTPException(status_code=404, detail="No packing list found")

    pli = db.query(PackingListItem).filter(
        PackingListItem.id == item_id,
        PackingListItem.packing_list_id == pl.id,
    ).first()
    if not pli:
        raise HTTPException(status_code=404, detail="Packing list item not found")

    if data.package_number is not None:
        raw = data.package_number.strip()
        if not raw:
            pli.package_number = None
            pli.packing_status = "NOT_READY"
        elif raw.upper() in ("BULK", "LOOSE"):
            pli.package_number = "BULK"
            pli.packing_status = "LOOSE"
        else:
            pli.package_number = raw
            pli.packing_status = "PALLETED"

    if data.factory_ready_qty is not None:
        pli.factory_ready_qty = data.factory_ready_qty
        # Auto-recalculate loaded_qty based on packing status (capped at ordered qty)
        if pli.packing_status in ("PALLETED", "LOOSE"):
            pli.loaded_qty = min(data.factory_ready_qty, pli.ordered_qty)
        else:
            pli.loaded_qty = 0

    # Recalculate total_packages
    all_items = db.query(PackingListItem).filter(PackingListItem.packing_list_id == pl.id).all()
    unique_packages = set()
    for item in all_items:
        if item.package_number and item.package_number not in ("BULK", None):
            unique_packages.add(item.package_number)
    pl.total_packages = len(unique_packages)

    db.commit()
    return {
        "updated": True,
        "package_number": pli.package_number,
        "packing_status": pli.packing_status,
        "factory_ready_qty": pli.factory_ready_qty,
        "loaded_qty": pli.loaded_qty,
        "total_packages": pl.total_packages,
    }


@router.post("/{order_id}/packing-list/items/{item_id}/split/")
def split_packing_item(
    order_id: str,
    item_id: str,
    data: SplitPackingItemRequest,
    db: Session = Depends(get_db),
):
    """Split a packing list item into multiple sub-rows with different pallet assignments."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order or order.status != OrderStatus.PLAN_PACKING.value:
        raise HTTPException(400, "Can only split items during Plan Packing stage")

    parent = db.query(PackingListItem).filter(PackingListItem.id == item_id).first()
    if not parent:
        raise HTTPException(404, "Packing list item not found")
    if parent.is_split:
        raise HTTPException(400, "Item is already a sub-row. Split the parent instead.")

    # Validate total split qty matches factory_ready_qty
    total_split = sum(s["qty"] for s in data.splits)
    if total_split != parent.factory_ready_qty:
        raise HTTPException(400, f"Split quantities ({total_split}) must equal factory ready qty ({parent.factory_ready_qty})")

    if len(data.splits) < 2:
        raise HTTPException(400, "Must split into at least 2 rows")

    # Create sub-rows
    sub_items = []
    for s in data.splits:
        pkg = (s.get("package_number") or "").strip()
        status = "PALLETED" if pkg and pkg != "BULK" else ("LOOSE" if pkg == "BULK" else "NOT_READY")
        sub = PackingListItem(
            packing_list_id=parent.packing_list_id,
            order_item_id=parent.order_item_id,
            product_id=parent.product_id,
            ordered_qty=parent.ordered_qty,  # preserve original ordered qty
            factory_ready_qty=s["qty"],
            loaded_qty=s["qty"] if status in ("PALLETED", "LOOSE") else 0,
            package_number=pkg or None,
            packing_status=status,
            parent_packing_item_id=parent.id,
            is_split=True,
            split_qty=s["qty"],
        )
        db.add(sub)
        sub_items.append(sub)

    # Mark parent as split (hide from active view, keep for reference)
    parent.packing_status = "SPLIT"
    parent.loaded_qty = 0  # children own the loaded qty now

    db.commit()
    return {"split_count": len(sub_items), "parent_id": parent.id}


@router.post("/{order_id}/packing-list/items/{item_id}/unsplit/")
def unsplit_packing_item(
    order_id: str,
    item_id: str,
    db: Session = Depends(get_db),
):
    """Undo a split — delete sub-rows, restore parent."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order or order.status != OrderStatus.PLAN_PACKING.value:
        raise HTTPException(400, "Can only unsplit during Plan Packing stage")

    parent = db.query(PackingListItem).filter(PackingListItem.id == item_id).first()
    if not parent or parent.packing_status != "SPLIT":
        raise HTTPException(400, "Item is not split")

    # Delete sub-rows
    db.query(PackingListItem).filter(
        PackingListItem.parent_packing_item_id == parent.id
    ).delete()

    # Restore parent
    parent.packing_status = "NOT_READY"
    parent.loaded_qty = 0
    db.commit()
    return {"status": "unsplit", "parent_id": parent.id}


@router.post("/{order_id}/packing-list/items/{item_id}/decision/")
def set_shipping_decision(
    order_id: str,
    item_id: str,
    data: ShippingDecisionRequest,
    db: Session = Depends(get_db),
):
    """Set shipping decision for a partially ready item."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order or order.status != OrderStatus.PLAN_PACKING.value:
        raise HTTPException(400, "Decisions only allowed during Plan Packing")

    item = db.query(PackingListItem).filter(PackingListItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "Item not found")

    if data.decision not in ("SHIP_CARRY_FORWARD", "SHIP_CANCEL_BALANCE", "WAIT"):
        raise HTTPException(400, "Invalid decision")

    if data.decision == "SHIP_CANCEL_BALANCE" and not data.cancel_reason:
        raise HTTPException(400, "Reason required when cancelling balance")

    # Validate item is actually partial (factory_ready_qty < ordered_qty)
    if item.factory_ready_qty >= item.ordered_qty:
        raise HTTPException(400, "Item is fully ready — no decision needed")

    item.shipping_decision = data.decision
    item.cancel_reason = data.cancel_reason

    if data.decision == "SHIP_CARRY_FORWARD":
        # Ship what's ready, create UnloadedItem for balance
        balance = item.ordered_qty - item.factory_ready_qty
        item.loaded_qty = item.factory_ready_qty

        oi = db.query(OrderItem).filter(OrderItem.id == item.order_item_id).first()
        existing = db.query(UnloadedItem).filter(
            UnloadedItem.original_order_id == order_id,
            UnloadedItem.product_id == item.product_id,
            UnloadedItem.reason == "NOT_PRODUCED",
        ).first()
        if not existing:
            ui = UnloadedItem(
                original_order_id=order_id,
                client_id=order.client_id,
                product_id=item.product_id,
                quantity=balance,
                amount_paid_inr=(oi.selling_price_inr or 0) * balance if oi else 0,
                status=UnloadedItemStatus.PENDING.value,
                reason="NOT_PRODUCED",
                factory_price=oi.factory_price if oi else None,
            )
            db.add(ui)

    elif data.decision == "SHIP_CANCEL_BALANCE":
        # Ship what's ready, reduce order item quantity
        item.loaded_qty = item.factory_ready_qty
        oi = db.query(OrderItem).filter(OrderItem.id == item.order_item_id).first()
        if oi:
            oi.quantity = item.factory_ready_qty
            oi.notes = f"Qty reduced from {item.ordered_qty} to {item.factory_ready_qty}: {data.cancel_reason}"

    elif data.decision == "WAIT":
        # Don't load — entire item waits
        item.loaded_qty = 0

    # Notify client about partial shipping decision
    if data.decision in ("SHIP_CARRY_FORWARD", "SHIP_CANCEL_BALANCE"):
        oi_for_notif = db.query(OrderItem).filter(OrderItem.id == item.order_item_id).first()
        product_label = oi_for_notif.product_code_snapshot if oi_for_notif else "Item"
        balance = item.ordered_qty - item.factory_ready_qty
        if data.decision == "SHIP_CARRY_FORWARD":
            msg = f"Order {order.order_number}: {product_label} — {balance} units carried forward to next order"
        else:
            msg = f"Order {order.order_number}: {product_label} — {balance} units cancelled ({data.cancel_reason})"

        notif = Notification(
            user_role="CLIENT",
            client_id=order.client_id,
            title="Packing Update",
            notification_type="PACKING_DECISION",
            message=msg,
            resource_type="order",
            resource_id=order_id,
        )
        db.add(notif)

    db.commit()
    return {"decision": data.decision, "loaded_qty": item.loaded_qty}


# ========================================
# 5D: PACKING LIST EXPORT (Excel + PDF)
# ========================================

def _get_packing_data(order_id: str, db: Session):
    """Shared helper: load packing list items split into active + migrated."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    pl = db.query(PackingList).filter(PackingList.order_id == order_id).first()
    if not pl:
        raise HTTPException(status_code=404, detail="No packing list found")

    items = db.query(PackingListItem).filter(PackingListItem.packing_list_id == pl.id).all()
    active, migrated = [], []
    for item in items:
        oi = db.query(OrderItem).filter(OrderItem.id == item.order_item_id).first()
        product = db.query(Product).filter(Product.id == item.product_id).first()

        # Resolve image: snapshot → product primary image → None
        img_path = None
        if oi and oi.image_path_snapshot:
            candidate = os.path.join(str(UPLOAD_DIR), oi.image_path_snapshot)
            if os.path.exists(candidate):
                img_path = candidate
        if not img_path and oi and hasattr(oi, 'factory_image_path') and oi.factory_image_path:
            candidate = os.path.join(str(UPLOAD_DIR), oi.factory_image_path)
            if os.path.exists(candidate):
                img_path = candidate
        if not img_path and product:
            prod_img = db.query(ProductImage).filter(
                ProductImage.product_id == product.id
            ).order_by(ProductImage.is_primary.desc(), ProductImage.created_at).first()
            if prod_img:
                candidate = os.path.join(str(UPLOAD_DIR), prod_img.image_path)
                if os.path.exists(candidate):
                    img_path = candidate

        row = {
            "product_code": oi.product_code_snapshot or (product.product_code if product else ""),
            "product_name": oi.product_name_snapshot or (product.product_name if product else ""),
            "ordered_qty": item.ordered_qty,
            "factory_ready_qty": item.factory_ready_qty,
            "package_number": item.package_number or "",
            "packing_status": item.packing_status or "",
            "order_item_status": oi.status if oi else "",
            "cancel_note": oi.cancel_note if oi else "",
            "image_path": img_path,
            "item_type": _classify_packing_item(oi),
        }
        if oi and oi.status == "UNLOADED":
            migrated.append(row)
        else:
            active.append(row)
    return order, pl, active, migrated


@router.get("/{order_id}/packing-list/download-excel/")
def download_packing_excel(order_id: str, db: Session = Depends(get_db)):
    """Download packing list as formatted Excel with product images."""
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from fastapi.responses import StreamingResponse
    import io

    order, pl, active, migrated = _get_packing_data(order_id, db)

    IMG_PX = 55  # Display size for product thumbnails

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing List"

    # Styles
    title_font = Font(name='Calibri', bold=True, size=14)
    info_font = Font(name='Calibri', bold=True, size=10)
    info_val_font = Font(name='Calibri', size=10)
    header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    normal_font = Font(name='Calibri', size=10)
    header_fill = PatternFill(start_color='6B21A8', end_color='6B21A8', fill_type='solid')
    migrated_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
    alt_fill = PatternFill(start_color='F5F3FF', end_color='F5F3FF', fill_type='solid')
    migrated_alt_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Type label mapping
    TYPE_LABELS = {
        "carried_forward": "Carried Fwd",
        "aftersales_replacement": "Replacement",
        "compensation": "Compensation",
        "regular": "",
    }
    type_amber_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    type_teal_fill = PatternFill(start_color='CCFBF1', end_color='CCFBF1', fill_type='solid')
    type_amber_font = Font(name='Calibri', size=9, color='92400E')
    type_teal_font = Font(name='Calibri', size=9, color='134E4A')

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 16

    # Title
    ws.merge_cells('A1:I1')
    cell = ws['A1']
    cell.value = 'PACKING LIST'
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')

    # Info
    ws['A3'] = 'Order:'
    ws['A3'].font = info_font
    ws['B3'] = order.order_number or order_id
    ws['B3'].font = info_val_font
    ws['D3'] = 'Date:'
    ws['D3'].font = info_font
    ws['E3'] = str(pl.uploaded_date)
    ws['E3'].font = info_val_font
    ws['A4'] = 'Packages:'
    ws['A4'].font = info_font
    ws['B4'] = pl.total_packages or 0
    ws['B4'].font = info_val_font

    # Active items header
    row = 6
    ws.merge_cells(f'A{row}:I{row}')
    ws[f'A{row}'] = f'PACKED ITEMS ({len(active)})'
    ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11, color='6B21A8')

    row = 7
    headers = ['#', 'Image', 'Part Code', 'Product Name', 'Ordered', 'Ready', 'Pallet #', 'Status', 'Type']
    center_cols = {1, 2, 5, 6, 7, 8, 9}
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    for idx, item in enumerate(active, 1):
        row += 1
        ws.row_dimensions[row].height = 50
        row_fill = alt_fill if idx % 2 == 0 else None

        type_label = TYPE_LABELS.get(item.get('item_type', ''), '')
        vals = [idx, '', item['product_code'], item['product_name'],
                item['ordered_qty'], item['factory_ready_qty'],
                item['package_number'], item['packing_status'], type_label]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx in center_cols else left_align
            if row_fill:
                cell.fill = row_fill
        # Color the Type cell based on item_type
        type_cell = ws.cell(row=row, column=9)
        itype = item.get('item_type', '')
        if itype == 'carried_forward':
            type_cell.fill = type_amber_fill
            type_cell.font = type_amber_font
        elif itype in ('aftersales_replacement', 'compensation'):
            type_cell.fill = type_teal_fill
            type_cell.font = type_teal_font

        # Embed image in column B
        img_file = item.get('image_path')
        if img_file and os.path.exists(img_file):
            try:
                xl_img = XlImage(img_file)
                scale = min(IMG_PX / xl_img.width, IMG_PX / xl_img.height)
                xl_img.width = int(xl_img.width * scale)
                xl_img.height = int(xl_img.height * scale)
                x_off = max(0, (84 - xl_img.width) // 2)
                y_off = max(0, (67 - xl_img.height) // 2)
                xl_img.anchor = OneCellAnchor(
                    _from=AnchorMarker(col=1, colOff=pixels_to_EMU(x_off),
                                       row=row - 1, rowOff=pixels_to_EMU(y_off)),
                    ext=XDRPositiveSize2D(pixels_to_EMU(xl_img.width),
                                          pixels_to_EMU(xl_img.height)),
                )
                ws.add_image(xl_img)
            except Exception:
                pass

    # Migrated items section
    if migrated:
        row += 2
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'] = f'MIGRATED / UNLOADED ITEMS ({len(migrated)})'
        ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=11, color='DC2626')

        row += 1
        mig_headers = ['#', 'Image', 'Part Code', 'Product Name', 'Qty', 'Type', 'Reason', '', '']
        for col_idx, h in enumerate(mig_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = migrated_fill
            cell.border = thin_border
            cell.alignment = center_align

        for idx, item in enumerate(migrated, 1):
            row += 1
            ws.row_dimensions[row].height = 50
            row_fill = migrated_alt_fill if idx % 2 == 0 else None
            reason = (item['cancel_note'] or '').replace('Migrated: ', '')
            type_label = TYPE_LABELS.get(item.get('item_type', ''), '')
            vals = [idx, '', item['product_code'], item['product_name'],
                    item['ordered_qty'], type_label, reason, '', '']
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = center_align if col_idx in {1, 2, 5, 6} else left_align
                if row_fill:
                    cell.fill = row_fill
            # Color the Type cell
            type_cell = ws.cell(row=row, column=6)
            itype = item.get('item_type', '')
            if itype == 'carried_forward':
                type_cell.fill = type_amber_fill
                type_cell.font = type_amber_font
            elif itype in ('aftersales_replacement', 'compensation'):
                type_cell.fill = type_teal_fill
                type_cell.font = type_teal_font

            img_file = item.get('image_path')
            if img_file and os.path.exists(img_file):
                try:
                    xl_img = XlImage(img_file)
                    scale = min(IMG_PX / xl_img.width, IMG_PX / xl_img.height)
                    xl_img.width = int(xl_img.width * scale)
                    xl_img.height = int(xl_img.height * scale)
                    x_off = max(0, (84 - xl_img.width) // 2)
                    y_off = max(0, (67 - xl_img.height) // 2)
                    xl_img.anchor = OneCellAnchor(
                        _from=AnchorMarker(col=1, colOff=pixels_to_EMU(x_off),
                                           row=row - 1, rowOff=pixels_to_EMU(y_off)),
                        ext=XDRPositiveSize2D(pixels_to_EMU(xl_img.width),
                                              pixels_to_EMU(xl_img.height)),
                    )
                    ws.add_image(xl_img)
                except Exception:
                    pass

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"PackingList_{order.order_number or order_id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{order_id}/packing-list/download-pdf/")
def download_packing_pdf(order_id: str, db: Session = Depends(get_db)):
    """Download packing list as landscape PDF with product images."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from fastapi.responses import StreamingResponse
    import io

    order, pl, active, migrated = _get_packing_data(order_id, db)

    from PIL import Image as PILImage

    IMG_W, IMG_H = 42, 42
    THUMB_PX = 120

    def _rl_image(path):
        if not path or not os.path.exists(path):
            return ''
        try:
            pil_img = PILImage.open(path)
            pil_img.thumbnail((THUMB_PX, THUMB_PX), PILImage.LANCZOS)
            # Convert RGBA/P to RGB (JPEG doesn't support transparency)
            if pil_img.mode in ('RGBA', 'P', 'LA'):
                bg = PILImage.new('RGB', pil_img.size, (255, 255, 255))
                bg.paste(pil_img, mask=pil_img.split()[-1] if pil_img.mode == 'RGBA' else None)
                pil_img = bg
            elif pil_img.mode != 'RGB':
                pil_img = pil_img.convert('RGB')
            thumb_buf = io.BytesIO()
            pil_img.save(thumb_buf, format='JPEG', quality=80)
            thumb_buf.seek(0)
            return RLImage(thumb_buf, width=IMG_W, height=IMG_H, kind='proportional')
        except Exception:
            return ''

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm,
    )
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=4)
    elements.append(Paragraph(f"Packing List &mdash; {order.order_number or order_id}", title_style))

    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, spaceAfter=2)
    elements.append(Paragraph(f"<b>Date:</b> {pl.uploaded_date} &nbsp;&nbsp;&nbsp; <b>Total Packages:</b> {pl.total_packages or 0}", info_style))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph(f"<b>Packed Items ({len(active)})</b>", ParagraphStyle('SH', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#6B21A8'), spaceAfter=4)))

    TYPE_LABELS = {
        "carried_forward": "Carried Fwd",
        "aftersales_replacement": "Replacement",
        "compensation": "Compensation",
        "regular": "",
    }

    # Paragraph styles for wrapping text in cells
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10)
    cell_bold = ParagraphStyle('CellBold', parent=cell_style, fontName='Helvetica-Bold')

    table_data = [['#', 'Image', 'Part Code', 'Product Name', 'Ordered', 'Ready', 'Pallet #', 'Status', 'Type']]
    row_heights = [None]
    for idx, item in enumerate(active, 1):
        table_data.append([
            str(idx),
            _rl_image(item.get('image_path')),
            Paragraph(item['product_code'], cell_bold),
            Paragraph(item['product_name'][:60], cell_style),
            str(item['ordered_qty']),
            str(item['factory_ready_qty']),
            item['package_number'],
            item['packing_status'],
            TYPE_LABELS.get(item.get('item_type', ''), ''),
        ])
        row_heights.append(52)

    col_widths = [22, 48, 95, 190, 42, 42, 44, 56, 60]
    table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B21A8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('ALIGN', (4, 0), (8, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F3FF')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    # Color-code Type cells per row
    for ridx, item in enumerate(active, 1):
        itype = item.get('item_type', '')
        if itype == 'carried_forward':
            style_cmds.append(('BACKGROUND', (8, ridx), (8, ridx), colors.HexColor('#FEF3C7')))
            style_cmds.append(('TEXTCOLOR', (8, ridx), (8, ridx), colors.HexColor('#92400E')))
        elif itype in ('aftersales_replacement', 'compensation'):
            style_cmds.append(('BACKGROUND', (8, ridx), (8, ridx), colors.HexColor('#CCFBF1')))
            style_cmds.append(('TEXTCOLOR', (8, ridx), (8, ridx), colors.HexColor('#134E4A')))
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    if migrated:
        elements.append(Spacer(1, 14))
        elements.append(Paragraph(f"<b>Migrated / Unloaded Items ({len(migrated)})</b>", ParagraphStyle('SH2', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#DC2626'), spaceAfter=4)))

        mig_data = [['#', 'Image', 'Part Code', 'Product Name', 'Qty', 'Type', 'Reason']]
        mig_heights = [None]
        for idx, item in enumerate(migrated, 1):
            reason = (item['cancel_note'] or '').replace('Migrated: ', '')
            mig_data.append([
                str(idx),
                _rl_image(item.get('image_path')),
                Paragraph(item['product_code'], cell_bold),
                Paragraph(item['product_name'][:60], cell_style),
                str(item['ordered_qty']),
                TYPE_LABELS.get(item.get('item_type', ''), ''),
                Paragraph(reason, cell_style),
            ])
            mig_heights.append(52)

        mig_col_widths = [22, 48, 95, 200, 42, 68, 120]
        mig_table = Table(mig_data, colWidths=mig_col_widths, rowHeights=mig_heights)
        mig_style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (4, 0), (5, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FEF2F2')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        for ridx, item in enumerate(migrated, 1):
            itype = item.get('item_type', '')
            if itype == 'carried_forward':
                mig_style_cmds.append(('BACKGROUND', (5, ridx), (5, ridx), colors.HexColor('#FEF3C7')))
                mig_style_cmds.append(('TEXTCOLOR', (5, ridx), (5, ridx), colors.HexColor('#92400E')))
            elif itype in ('aftersales_replacement', 'compensation'):
                mig_style_cmds.append(('BACKGROUND', (5, ridx), (5, ridx), colors.HexColor('#CCFBF1')))
                mig_style_cmds.append(('TEXTCOLOR', (5, ridx), (5, ridx), colors.HexColor('#134E4A')))
        mig_table.setStyle(TableStyle(mig_style_cmds))
        elements.append(mig_table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"PackingList_{order.order_number or order_id}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ========================================
# Item Migration (Level 5C)
# ========================================

@router.post("/{order_id}/migrate-items/")
def migrate_items(order_id: str, data: MigrateItemsRequest, db: Session = Depends(get_db)):
    """Migrate selected items to unloaded status for carry-forward to next order."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if order.status != OrderStatus.PLAN_PACKING.value:
        raise HTTPException(status_code=400, detail="Migration only allowed at Plan Packing stage")

    migrated = 0
    for item_req in data.items:
        oi = db.query(OrderItem).filter(
            OrderItem.id == item_req.order_item_id,
            OrderItem.order_id == order_id,
            OrderItem.status == OrderItemStatus.ACTIVE.value,
        ).first()
        if not oi:
            continue

        # Block compensation balance items from migration (they're ledger-only)
        if (oi.selling_price_inr or 0) < 0:
            continue

        # Lookup packing list item first — we need factory_ready_qty for accurate carry-forward
        pli = db.query(PackingListItem).filter(
            PackingListItem.order_item_id == oi.id,
        ).first()

        # Create UnloadedItem record — use split_qty for sub-rows, factory_ready_qty otherwise
        if pli and pli.is_split:
            carry_qty = pli.split_qty or pli.factory_ready_qty
        elif pli:
            carry_qty = pli.factory_ready_qty
        else:
            carry_qty = oi.quantity
        db.add(UnloadedItem(
            original_order_id=order_id,
            client_id=order.client_id,
            product_id=oi.product_id,
            quantity=carry_qty,
            amount_paid_inr=0,
            reason=item_req.reason,
            factory_price=oi.factory_price,
            status=UnloadedItemStatus.PENDING.value,
        ))

        # Mark order item as UNLOADED
        oi.status = OrderItemStatus.UNLOADED.value
        oi.cancel_note = f"Migrated: {item_req.reason}"

        # Update packing list item quantities
        if pli:
            pli.unloaded_qty = pli.factory_ready_qty
            pli.loaded_qty = 0

        # If this is a carry-forward REPLACEMENT item, revert original claim to PENDING
        if (oi.selling_price_inr is not None and oi.selling_price_inr == 0
                and oi.notes and "After-Sales" in oi.notes):
            # Find the AfterSalesItem linked to this order item
            from models import AfterSalesItem
            from enums import CarryForwardStatus
            linked_asi = db.query(AfterSalesItem).filter(
                AfterSalesItem.order_id == order_id,
                AfterSalesItem.order_item_id == oi.id,
            ).first()
            if linked_asi and linked_asi.source_aftersales_id:
                # Revert the ORIGINAL claim back to PENDING
                original_claim = db.query(AfterSalesItem).filter(
                    AfterSalesItem.id == linked_asi.source_aftersales_id
                ).first()
                if original_claim and original_claim.carry_forward_status == CarryForwardStatus.ADDED_TO_ORDER.value:
                    original_claim.carry_forward_status = CarryForwardStatus.PENDING.value
                    original_claim.added_to_order_id = None

        migrated += 1

    # Notify client about carried-forward items
    if migrated > 0 and order.client_id:
        not_produced = sum(1 for i in data.items if i.reason == "NOT_PRODUCED")
        no_space = sum(1 for i in data.items if i.reason == "NO_SPACE")
        msg_parts = []
        if not_produced:
            msg_parts.append(f"{not_produced} not produced")
        if no_space:
            msg_parts.append(f"{no_space} no space in container")
        reason_text = f" ({', '.join(msg_parts)})" if msg_parts else ""

        db.add(Notification(
            client_id=order.client_id,
            user_role="CLIENT",
            title="Items Carried Forward",
            message=f"{migrated} item(s) from {order.order_number} will be in your next order{reason_text}",
            notification_type="ITEMS_MIGRATED",
            resource_type="order",
            resource_id=order.id,
        ))

    # Update effective PI total after migration
    from core.finance_helpers import update_effective_pi_total
    update_effective_pi_total(order, db)

    db.commit()
    return {"migrated": migrated}


@router.post("/{order_id}/undo-migrate/")
def undo_migrate_items(order_id: str, data: UndoMigrateRequest, db: Session = Depends(get_db)):
    """Undo migration — restore items back to ACTIVE status."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.status != OrderStatus.PLAN_PACKING.value:
        raise HTTPException(status_code=400, detail="Undo migration only allowed at Plan Packing stage")

    restored = 0
    for oi_id in data.order_item_ids:
        oi = db.query(OrderItem).filter(
            OrderItem.id == oi_id,
            OrderItem.order_id == order_id,
            OrderItem.status == OrderItemStatus.UNLOADED.value,
        ).first()
        if not oi:
            continue

        # Restore order item
        oi.status = OrderItemStatus.ACTIVE.value
        oi.cancel_note = None

        # Delete ONE matching UnloadedItem record (not all for same product)
        ui = db.query(UnloadedItem).filter(
            UnloadedItem.original_order_id == order_id,
            UnloadedItem.product_id == oi.product_id,
            UnloadedItem.status == UnloadedItemStatus.PENDING.value,
        ).first()
        if ui:
            db.delete(ui)

        # Restore packing list item quantities
        pli = db.query(PackingListItem).filter(
            PackingListItem.order_item_id == oi.id,
        ).first()
        if pli:
            pli.loaded_qty = pli.factory_ready_qty
            pli.unloaded_qty = 0

        # If this was a carry-forward REPLACEMENT item, re-link original claim
        if (oi.selling_price_inr is not None and oi.selling_price_inr == 0
                and oi.notes and "After-Sales" in oi.notes):
            from models import AfterSalesItem
            from enums import CarryForwardStatus
            linked_asi = db.query(AfterSalesItem).filter(
                AfterSalesItem.order_id == order_id,
                AfterSalesItem.order_item_id == oi.id,
            ).first()
            if linked_asi and linked_asi.source_aftersales_id:
                original_claim = db.query(AfterSalesItem).filter(
                    AfterSalesItem.id == linked_asi.source_aftersales_id
                ).first()
                if original_claim and original_claim.carry_forward_status == CarryForwardStatus.PENDING.value:
                    original_claim.carry_forward_status = CarryForwardStatus.ADDED_TO_ORDER.value
                    original_claim.added_to_order_id = order_id

        restored += 1

    # Update effective PI total after restoring items
    from core.finance_helpers import update_effective_pi_total
    update_effective_pi_total(order, db)

    db.commit()
    return {"restored": restored}


# ========================================
# Production Progress Dates
# ========================================
@router.get("/{order_id}/production-progress/")
def get_production_progress(order_id: str, db: Session = Depends(get_db)):
    """Get production start date, target date, percent complete, and days remaining."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    started = order.production_started_at
    target = order.production_target_date
    today = date.today()

    percent = 0
    days_remaining = None
    is_overdue = False
    overdue_days = 0

    if started and target:
        total_span = (target - started).days
        elapsed = (today - started).days
        if total_span > 0:
            percent = round(elapsed / total_span * 100)
        elif total_span == 0:
            percent = 100 if elapsed >= 0 else 0

        days_remaining = (target - today).days
        if days_remaining < 0:
            is_overdue = True
            overdue_days = abs(days_remaining)
            days_remaining = 0

    return {
        "started_at": started.isoformat() if started else None,
        "target_date": target.isoformat() if target else None,
        "percent": percent,
        "days_remaining": days_remaining,
        "is_overdue": is_overdue,
        "overdue_days": overdue_days,
    }

@router.put("/{order_id}/production-dates/")
def set_production_dates(order_id: str, data: ProductionDatesIn, db: Session = Depends(get_db)):
    """Set production start date and/or target completion date."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if data.started_at:
        order.production_started_at = date.fromisoformat(data.started_at)
    if data.target_date:
        order.production_target_date = date.fromisoformat(data.target_date)
    db.commit()
    return {
        "started_at": order.production_started_at.isoformat() if order.production_started_at else None,
        "target_date": order.production_target_date.isoformat() if order.production_target_date else None,
    }


# ── Client-accessible PI download ──────────────────────────────────────────────

from fastapi.responses import FileResponse

# Statuses where PI is available for client download
_PI_AVAILABLE_STATUSES = {
    OrderStatus.PI_SENT.value, OrderStatus.ADVANCE_PENDING.value,
    OrderStatus.ADVANCE_RECEIVED.value, OrderStatus.FACTORY_ORDERED.value,
    OrderStatus.PRODUCTION_60.value, OrderStatus.PRODUCTION_80.value,
    OrderStatus.PRODUCTION_90.value, OrderStatus.PLAN_PACKING.value,
    OrderStatus.FINAL_PI.value, OrderStatus.PRODUCTION_100.value,
    OrderStatus.BOOKED.value, OrderStatus.LOADED.value,
    OrderStatus.SAILED.value, OrderStatus.ARRIVED.value,
    OrderStatus.CUSTOMS_FILED.value, OrderStatus.CLEARED.value,
    OrderStatus.DELIVERED.value, OrderStatus.AFTER_SALES.value,
    OrderStatus.COMPLETED.value,
}


def _get_order_for_pi(order_id: str, current_user: CurrentUser, db: Session) -> Order:
    """Fetch order with row-level security for PI download."""
    order = db.query(Order).filter(
        Order.id == order_id, Order.deleted_at.is_(None),
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Client users can only download their own orders
    if current_user.user_type == "CLIENT" and order.client_id != current_user.client_id:
        raise HTTPException(status_code=403, detail="Access denied")

    if order.status not in _PI_AVAILABLE_STATUSES:
        raise HTTPException(status_code=400, detail="PI is not yet available for this order")

    return order


@router.get("/{order_id}/download-pi/")
def client_download_pi(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Download the PI Excel for an order. Accessible by both admin and client users."""
    order = _get_order_for_pi(order_id, current_user, db)

    pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order_id).first()
    if not pi or not pi.file_path or not os.path.exists(pi.file_path):
        raise HTTPException(status_code=404, detail="PI file has not been generated yet. Please contact admin.")

    return FileResponse(
        path=pi.file_path,
        filename=os.path.basename(pi.file_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/{order_id}/download-pi-with-images/")
def client_download_pi_with_images(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Generate and download PI with images. Accessible by both admin and client users."""
    order = _get_order_for_pi(order_id, current_user, db)

    # Delegate to the excel module's generation logic
    from routers.excel import download_pi_with_images as _gen_pi_images
    return _gen_pi_images(order_id, db)


# ── Client Ledger / Statement of Account ───────────────────────────────────────

