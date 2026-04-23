"""
Excel Upload & Processing API (Level 3)
Handles large file uploads (up to 500MB) with background processing.
"""
import json
from math import ceil
import os
import shutil
import uuid as uuid_mod
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks, Query
from fastapi.responses import FileResponse
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pydantic import BaseModel, Field, model_validator
from fastapi import Request as FastAPIRequest
from rate_limiter import limiter
from sqlalchemy import desc, or_
from sqlalchemy.orm import Session

import logging

from config import UPLOAD_DIR, MAX_UPLOAD_SIZE
from core.file_upload import stream_upload_to_disk
from database import get_db, SessionLocal
from core.security import check_pi_order_access, CurrentUser, get_current_user

logger = logging.getLogger(__name__)
from enums import JobStatus, JobType, OrderStatus, OrderItemStatus
from models import (
    ProcessingJob, Order, Product, OrderItem,
    ProductImage, ProformaInvoice, PIRevision, Client, Document,
)
from sqlalchemy import func as sql_func
from services.excel_parser import process_client_excel, process_factory_excel
from services.excel_apply import apply_parsed_data as _apply_parsed_data

router = APIRouter()
pi_router = APIRouter()  # PI endpoints — registered separately with get_current_user dep


# ========================================
# Pydantic Schemas
# ========================================
class JobOut(BaseModel):
    id: str
    order_id: Optional[str] = None
    job_type: str
    file_path: Optional[str] = None
    original_filename: Optional[str] = None
    file_size: Optional[int] = None
    status: str
    progress: int
    total_rows: Optional[int] = None
    processed_rows: Optional[int] = None
    result_summary: Optional[dict] = None
    result_data: Optional[list] = None
    error_message: Optional[str] = None
    created_at: Optional[str] = None
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    created_by: Optional[str] = None  # G-015: user ID of job creator

    class Config:
        from_attributes = True


def serialize_job(job: ProcessingJob) -> dict:
    """Convert job to dict, parsing JSON fields."""
    data = JobOut(
        id=job.id,
        order_id=job.order_id,
        job_type=job.job_type,
        file_path=job.file_path,
        original_filename=job.original_filename,
        file_size=job.file_size,
        status=job.status,
        progress=job.progress,
        total_rows=job.total_rows,
        processed_rows=job.processed_rows,
        result_summary=json.loads(job.result_summary) if job.result_summary else None,
        result_data=json.loads(job.result_data) if job.result_data else None,
        error_message=job.error_message,
        created_at=job.created_at.isoformat() if job.created_at else None,
        started_at=job.started_at.isoformat() if job.started_at else None,
        completed_at=job.completed_at.isoformat() if job.completed_at else None,
        created_by=getattr(job, 'created_by', None),
    ).model_dump()
    return data


def _validate_file_path(file_path: str) -> str:
    """Validate and resolve file_path, ensuring it stays within UPLOAD_DIR.

    Raises HTTPException on path traversal attempts.
    """
    if not os.path.isabs(file_path):
        resolved = (UPLOAD_DIR / file_path).resolve()
    else:
        resolved = Path(file_path).resolve()

    upload_dir_resolved = UPLOAD_DIR.resolve()
    if not str(resolved).startswith(str(upload_dir_resolved)):
        raise HTTPException(status_code=400, detail="Invalid file path")

    return str(resolved)


# ========================================
# Upload Endpoint
# ========================================
@router.post("/upload/")
@limiter.limit("3/minute")
async def upload_excel(
    request: FastAPIRequest,
    file: UploadFile = File(...),
    order_id: Optional[str] = Query(None),
    job_type: str = Query("CLIENT_EXCEL"),
    skip_processing: bool = Query(False),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Upload an Excel file. Streams to disk in chunks (not memory).
    Returns job_id immediately. Processing happens in background."""
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only .xlsx and .xls files are accepted")

    if job_type not in [jt.value for jt in JobType]:
        raise HTTPException(status_code=400, detail=f"Invalid job_type. Use: {[jt.value for jt in JobType]}")

    if order_id:
        order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

    temp_filename = f"{uuid_mod.uuid4()}.xlsx"
    temp_path = UPLOAD_DIR / "temp" / temp_filename

    try:
        file_size = await stream_upload_to_disk(file, temp_path, MAX_UPLOAD_SIZE)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=500, detail="Upload failed")

    job = ProcessingJob(
        order_id=order_id,
        job_type=job_type,
        file_path=str(temp_path),
        original_filename=file.filename,
        file_size=file_size,
        status=JobStatus.PENDING.value,
        progress=0,
        created_by=current_user.id,  # G-015: record job creator for ownership check
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    if not skip_processing:
        background_tasks.add_task(process_excel_job, job.id)

    # Enforce temp file FIFO limit (keep max 10, delete oldest)
    from services.cleanup import enforce_temp_file_limit
    enforce_temp_file_limit()

    return serialize_job(job)


# ========================================
# AI Column Analysis Endpoint
# ========================================
class AnalyzeColumnsRequest(BaseModel):
    file_path: str
    schema_type: str = "product"
    sheet_index: int = 0


@router.post("/analyze-columns/")
@limiter.limit("5/minute")
def analyze_columns_endpoint(
    request: FastAPIRequest,
    req: AnalyzeColumnsRequest,
    db: Session = Depends(get_db),
):
    """Analyze Excel column headers using AI and return mapping suggestions."""
    from services.column_mapper import analyze_columns

    file_path = _validate_file_path(req.file_path)

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.worksheets[req.sheet_index] if req.sheet_index < len(wb.worksheets) else wb.active
        rows = []
        for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
            rows.append(list(row))
            if ri >= 3:
                break
        wb.close()
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read Excel file")

    if not rows:
        raise HTTPException(status_code=400, detail="Empty Excel file")

    headers = [str(h) if h is not None else "" for h in rows[0]]
    sample_rows = rows[1:] if len(rows) > 1 else []

    result = analyze_columns(headers, sample_rows, req.schema_type)
    return result


# ========================================
# AI Conflict Resolution
# ========================================
class ConflictGroup(BaseModel):
    code: str = Field(..., max_length=200)
    rows: List[dict] = Field(..., max_length=500)
    existing_variants: List[dict] = Field(default=[], max_length=200)

class AnalyzeConflictsRequest(BaseModel):
    groups: List[ConflictGroup] = Field(..., max_length=1000)


@router.post("/analyze-conflicts/")
@limiter.limit("5/minute")
def analyze_conflicts_endpoint(request: FastAPIRequest, req: AnalyzeConflictsRequest):
    """Use AI to suggest resolution actions for conflict groups."""
    from services.conflict_resolver import analyze_conflicts
    groups_data = [g.model_dump() for g in req.groups]
    return analyze_conflicts(groups_data)


# ========================================
# Job Status Endpoints
# ========================================
@router.get("/jobs/{job_id}/")
def get_job_status(
    job_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Get current status of a processing job."""
    job = db.query(ProcessingJob).filter(ProcessingJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    # G-015: OPERATIONS users may only view their own jobs; ADMIN/SUPER_ADMIN see any
    if (
        current_user.role not in ("ADMIN", "SUPER_ADMIN")
        and job.created_by is not None
        and job.created_by != current_user.id
    ):
        raise HTTPException(status_code=403, detail="Access denied: not your job")
    return serialize_job(job)


@router.get("/jobs/")
def list_jobs(
    order_id: Optional[str] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """List processing jobs with optional filters."""
    query = db.query(ProcessingJob)
    if order_id:
        query = query.filter(ProcessingJob.order_id == order_id)
    if status:
        query = query.filter(ProcessingJob.status == status)

    total = query.count()
    jobs = query.order_by(desc(ProcessingJob.created_at)).offset(
        (page - 1) * per_page
    ).limit(per_page).all()

    return {
        "items": [serialize_job(j) for j in jobs],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": ceil(total / per_page) if per_page else 0,
    }


@router.delete("/jobs/{job_id}/")
def cancel_job(
    job_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Cancel a pending/processing job and clean up temp file."""
    job = db.query(ProcessingJob).filter(ProcessingJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    # G-015: OPERATIONS users may only cancel their own jobs; ADMIN/SUPER_ADMIN cancel any
    if (
        current_user.role not in ("ADMIN", "SUPER_ADMIN")
        and job.created_by is not None
        and job.created_by != current_user.id
    ):
        raise HTTPException(status_code=403, detail="Access denied: not your job")

    if job.file_path and os.path.exists(job.file_path):
        os.remove(job.file_path)

    db.delete(job)
    db.commit()
    return {"message": "Job cancelled and file deleted"}


class ReparseRequest(BaseModel):
    column_mapping: Optional[Dict[str, str]] = None


@router.post("/jobs/{job_id}/reparse/")
def reparse_job(
    job_id: str,
    body: Optional[ReparseRequest] = None,
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
):
    """Re-trigger parsing for an existing job."""
    job = db.query(ProcessingJob).filter(ProcessingJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if not job.file_path or not os.path.exists(job.file_path):
        raise HTTPException(status_code=400, detail="Original file no longer available")

    job.status = JobStatus.PENDING.value
    job.progress = 0
    job.result_summary = None
    job.result_data = None
    job.completed_at = None
    job.error_message = None

    if body and body.column_mapping:
        mapping_path = job.file_path + ".column_mapping.json"
        with open(mapping_path, "w", encoding="utf-8") as f:
            json.dump(body.column_mapping, f)
    db.commit()

    background_tasks.add_task(process_excel_job, job.id)
    return serialize_job(job)


# ========================================
# Background Processing
# ========================================
def process_excel_job(job_id: str):
    """Background task: Process uploaded Excel file."""
    db = SessionLocal()
    try:
        job = db.query(ProcessingJob).filter(ProcessingJob.id == job_id).first()
        if not job:
            return

        job.status = JobStatus.PROCESSING.value
        job.started_at = datetime.utcnow()
        db.commit()

        if job.job_type == JobType.CLIENT_EXCEL.value:
            process_client_excel(job, db)
        elif job.job_type == JobType.FACTORY_EXCEL.value:
            process_factory_excel(job, db)
        else:
            job.status = JobStatus.FAILED.value
            job.error_message = f"Unknown job type: {job.job_type}"
            db.commit()
    except Exception as e:
        job.status = JobStatus.FAILED.value
        job.error_message = str(e)
        job.completed_at = datetime.utcnow()
        db.commit()
    finally:
        db.close()


# ========================================
# Apply Parsed Data
# ========================================
class ApplyParsedDataRequest(BaseModel):
    job_id: str = Field(..., max_length=200)
    create_new_products: bool = True
    duplicate_resolutions: dict = {}
    image_conflict_resolutions: dict = {}
    variant_resolutions: dict = {}
    row_overrides: dict = {}

    @model_validator(mode='after')
    def validate_dict_sizes(self):
        for field_name in ['duplicate_resolutions', 'image_conflict_resolutions', 'variant_resolutions', 'row_overrides']:
            d = getattr(self, field_name)
            if len(d) > 10000:
                raise ValueError(f'{field_name} exceeds maximum of 10000 entries')
        return self


@router.post("/apply/{job_id}/")
def apply_parsed_data_endpoint(
    job_id: str,
    data: ApplyParsedDataRequest,
    db: Session = Depends(get_db),
):
    """Apply parsed Excel results to create products and order items."""
    job = db.query(ProcessingJob).filter(ProcessingJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.status != JobStatus.COMPLETED.value:
        raise HTTPException(status_code=400, detail="Job not completed yet")
    if not job.result_data:
        raise HTTPException(status_code=400, detail="Job has no result data")

    order = None
    if job.order_id:
        order = db.query(Order).filter(Order.id == job.order_id, Order.deleted_at.is_(None)).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

    results = json.loads(job.result_data)

    result = _apply_parsed_data(
        job=job,
        order=order,
        results=results,
        duplicate_resolutions=data.duplicate_resolutions,
        image_conflict_resolutions=data.image_conflict_resolutions,
        variant_resolutions=data.variant_resolutions,
        row_overrides=data.row_overrides,
        create_new_products=data.create_new_products,
        db=db,
    )

    # Auto-recalculate client_factory_price for transparency clients
    if order and order.client_id:
        from config import TRANSPARENCY_ENABLED
        if TRANSPARENCY_ENABLED:
            client = db.query(Client).filter(Client.id == order.client_id).first()
            if client and client.client_type == "TRANSPARENCY" and client.factory_markup_percent is not None:
                from decimal import Decimal, ROUND_HALF_UP
                markup = Decimal(str(client.factory_markup_percent))
                items = db.query(OrderItem).filter(
                    OrderItem.order_id == order.id,
                    OrderItem.status == OrderItemStatus.ACTIVE.value,
                    OrderItem.factory_price.isnot(None),
                    OrderItem.factory_price > 0,
                ).all()
                for item in items:
                    factory = Decimal(str(item.factory_price))
                    cfp = float((factory * (1 + markup / 100)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    if item.client_factory_price != cfp:
                        item.client_factory_price = cfp
                db.commit()

    # Clean up temp file after successful apply
    if job.file_path:
        from pathlib import Path
        temp_file = UPLOAD_DIR / job.file_path if not Path(job.file_path).is_absolute() else Path(job.file_path)
        if temp_file.exists():
            try:
                temp_file.unlink()
            except Exception:
                pass

    return result


# ========================================
# PI (Proforma Invoice) Generation
# ========================================
class GeneratePIRequest(BaseModel):
    advance_percent: float = 30.0


@pi_router.post("/generate-pi/{order_id}/")
def generate_pi(
    order_id: str,
    data: GeneratePIRequest = GeneratePIRequest(),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Generate a clean Proforma Invoice Excel for the client."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    check_pi_order_access(order, current_user)

    client = db.query(Client).filter(Client.id == order.client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    items = db.query(OrderItem).filter(
        OrderItem.order_id == order_id,
        OrderItem.status == OrderItemStatus.ACTIVE.value,
        or_(OrderItem.pi_item_status.is_(None), OrderItem.pi_item_status == "APPROVED"),
    ).all()

    if not items:
        raise HTTPException(status_code=400, detail="No active items in this order")

    # Determine if this is a transparency client
    is_transparency = (
        client.client_type == "TRANSPARENCY"
        and client.factory_markup_percent is not None
    )
    rate = order.exchange_rate or 1.0

    missing_prices = []
    for item in items:
        product = db.query(Product).filter(Product.id == item.product_id).first()
        is_aftersales_item = item.notes and "After-Sales" in item.notes
        if is_transparency:
            # Transparency clients need client_factory_price
            if item.client_factory_price is None and not is_aftersales_item and (item.factory_price or 0) > 0:
                pass  # factory price exists, cfp will be auto-calc'd — OK
            elif item.factory_price is None and not is_aftersales_item:
                code = item.product_code_snapshot or (product.product_code if product else None) or "Unknown"
                missing_prices.append(code)
        else:
            if item.selling_price_inr is None or (item.selling_price_inr <= 0 and not is_aftersales_item):
                code = item.product_code_snapshot or (product.product_code if product else None) or "Unknown"
                missing_prices.append(code)

    if missing_prices:
        raise HTTPException(
            status_code=400,
            detail=f"Missing {'factory' if is_transparency else 'selling'} prices for: {', '.join(missing_prices[:10])}"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proforma Invoice"

    styles = _pi_styles()
    _set_pi_column_widths(ws, with_images=False)
    _write_pi_header(ws, styles, merge_range='A1:I1')

    pi_number = order.order_number or f"PI-{datetime.utcnow().strftime('%Y%m%d')}-{uuid_mod.uuid4().hex[:4].upper()}"
    _write_pi_info(ws, styles, pi_number, client, date_col='G', date_val_col='H')

    if is_transparency:
        headers = ['S.No', 'Part Code', 'Description', 'Type', 'Size', 'Material', 'Qty', 'Unit Price (INR)', 'Total (INR)']
    else:
        headers = ['S.No', 'Part Code', 'Description', 'Type', 'Size', 'Material', 'Qty', 'Unit Price (INR)', 'Total (INR)']
    _write_pi_col_headers(ws, 7, headers, styles)

    grand_total = 0
    row = 7
    for idx, item in enumerate(items, 1):
        row += 1
        product = db.query(Product).filter(Product.id == item.product_id).first()
        _lazy_snapshot_backfill(item, product)

        part_code, part_name, pi_part_type, pi_dimension, pi_material = _snapshot_fields(item, product)

        # For transparency clients: use client_factory_price * exchange_rate
        # For regular clients: use selling_price_inr
        if is_transparency:
            cfp = float(item.client_factory_price or 0)
            unit_price_inr = round(cfp * rate, 2)
            line_total = round(unit_price_inr * item.quantity, 2)
        else:
            unit_price_inr = item.selling_price_inr or 0
            line_total = round(unit_price_inr * item.quantity, 2)
        grand_total += line_total
        row_fill = styles['alt_fill'] if idx % 2 == 0 else None

        values = [idx, part_code, part_name, pi_part_type, pi_dimension, pi_material,
                  item.quantity, unit_price_inr, line_total]
        # Text columns that should wrap: Code(2), Desc(3), Type(4), Size(5), Material(6)
        _wrap_cols_noimgs = {2, 3, 4, 5, 6}
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = styles['normal_font']
            cell.border = styles['thin_border']
            cell.alignment = Alignment(
                horizontal='left' if col_idx in _wrap_cols_noimgs else 'center',
                vertical='center',
                wrap_text=(col_idx in _wrap_cols_noimgs),
            )
            if col_idx in (8, 9):
                cell.number_format = styles['currency_format']
            if row_fill:
                cell.fill = row_fill

    row += 1
    _write_pi_total_row(ws, row, grand_total, len(items), styles, total_col=9, merge_range=f'A{row}:G{row}')
    _write_pi_advance(ws, row + 2, grand_total, data.advance_percent, styles, label_col='G', val_col=9)

    pi_dir = UPLOAD_DIR / "orders" / order_id
    pi_dir.mkdir(parents=True, exist_ok=True)
    pi_filename = f"PI_{pi_number.replace('/', '-')}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    pi_path = pi_dir / pi_filename
    wb.save(str(pi_path))

    _save_pi_record(db, order_id, order, pi_number, grand_total, data.advance_percent, pi_path, items)

    doc = Document(
        order_id=order_id,
        doc_type="PROFORMA_INVOICE",
        file_path=f"orders/{order_id}/{pi_filename}",
        filename=pi_filename,
        file_size=os.path.getsize(str(pi_path)),
    )
    db.add(doc)
    order.items_modified_at = None
    db.commit()

    advance_amount = grand_total * (data.advance_percent / 100)
    return {
        "pi_number": pi_number,
        "total_items": len(items),
        "grand_total_inr": round(grand_total, 2),
        "advance_percent": data.advance_percent,
        "advance_amount_inr": round(advance_amount, 2),
        "balance_amount_inr": round(grand_total - advance_amount, 2),
        "file_path": f"/uploads/orders/{order_id}/{pi_filename}",
        "filename": pi_filename,
    }


@pi_router.get("/download-pi/{order_id}/")
def download_pi(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Download the latest PI for an order."""
    pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order_id).first()
    if not pi:
        raise HTTPException(status_code=404, detail="No PI found for this order")
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    check_pi_order_access(order, current_user)
    if not pi.file_path or not os.path.exists(pi.file_path):
        raise HTTPException(status_code=404, detail="PI file not found on disk")

    return FileResponse(
        path=pi.file_path,
        filename=os.path.basename(pi.file_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pi_router.get("/download-pi-with-images/{order_id}/")
def download_pi_with_images(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Generate and download a PI Excel with product images embedded in cells."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    check_pi_order_access(order, current_user)

    client = db.query(Client).filter(Client.id == order.client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    items = db.query(OrderItem).filter(
        OrderItem.order_id == order_id,
        OrderItem.status == OrderItemStatus.ACTIVE.value,
        or_(OrderItem.pi_item_status.is_(None), OrderItem.pi_item_status == "APPROVED"),
    ).all()
    if not items:
        raise HTTPException(status_code=400, detail="No active items in this order")

    # Detect transparency client
    is_transparency = (
        client.client_type == "TRANSPARENCY"
        and client.factory_markup_percent is not None
    )
    rate = order.exchange_rate or 1.0
    currency = order.currency or "USD"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proforma Invoice"

    styles = _pi_styles()
    _set_pi_column_widths(ws, with_images=True)

    if is_transparency:
        # 11 columns: adds Factory (currency) before Price (INR)
        _write_pi_header(ws, styles, merge_range='A1:K1')
        pi_record = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order_id).first()
        pi_number = pi_record.pi_number if pi_record else (
            order.order_number or f"PI-{datetime.utcnow().strftime('%Y%m%d')}-{uuid_mod.uuid4().hex[:4].upper()}"
        )
        _write_pi_info(ws, styles, pi_number, client, date_col='I', date_val_col='J', client_merge='B4:D4')
        headers = ['S.No', 'Image', 'Part Code', 'Description', 'Type', 'Size', 'Material', 'Qty',
                   f'Factory ({currency})', 'Price (INR)', 'Total (INR)']
        _write_pi_col_headers(ws, 7, headers, styles)
        # Extra column width for Factory price
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 14
        ws.column_dimensions['K'].width = 16
    else:
        # 10 columns: standard layout
        _write_pi_header(ws, styles, merge_range='A1:J1')
        pi_record = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order_id).first()
        pi_number = pi_record.pi_number if pi_record else (
            order.order_number or f"PI-{datetime.utcnow().strftime('%Y%m%d')}-{uuid_mod.uuid4().hex[:4].upper()}"
        )
        _write_pi_info(ws, styles, pi_number, client, date_col='H', date_val_col='I', client_merge='B4:D4')
        headers = ['S.No', 'Image', 'Part Code', 'Description', 'Type', 'Size', 'Material', 'Qty', 'Unit Price (INR)', 'Total (INR)']
        _write_pi_col_headers(ws, 7, headers, styles)

    IMG_DISPLAY_PX = 60
    grand_total = 0
    snapshot_dirty = False
    row = 7

    for idx, item in enumerate(items, 1):
        row += 1
        product = db.query(Product).filter(Product.id == item.product_id).first()
        if _lazy_snapshot_backfill(item, product):
            snapshot_dirty = True

        part_code, part_name, pi_part_type, pi_dimension, pi_material = _snapshot_fields(item, product)

        # Calculate pricing based on client type
        if is_transparency:
            cfp = float(item.client_factory_price or 0)
            unit_price_inr = round(cfp * rate, 2)
            line_total = round(unit_price_inr * item.quantity, 2)
        else:
            cfp = None
            unit_price_inr = item.selling_price_inr or 0
            line_total = round(unit_price_inr * item.quantity, 2)

        grand_total += line_total
        row_fill = styles['alt_fill'] if idx % 2 == 0 else None

        ws.row_dimensions[row].height = 50

        # S.No
        _write_cell(ws, row, 1, idx, styles, row_fill)

        # Image cell
        cell = ws.cell(row=row, column=2)
        cell.border = styles['thin_border']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if row_fill:
            cell.fill = row_fill

        img_path_to_use = _resolve_item_image(item, product, order_id, db)
        if img_path_to_use and item.image_path_snapshot and not snapshot_dirty:
            pass
        elif img_path_to_use and not (item.image_path_snapshot and os.path.exists(os.path.join(str(UPLOAD_DIR), item.image_path_snapshot))):
            snapshot_dirty = True

        if img_path_to_use and os.path.exists(img_path_to_use):
            try:
                xl_img = XlImage(img_path_to_use)
                scale = min(IMG_DISPLAY_PX / xl_img.width, IMG_DISPLAY_PX / xl_img.height)
                img_w = int(xl_img.width * scale)
                img_h = int(xl_img.height * scale)
                xl_img.width = img_w
                xl_img.height = img_h

                EMU_PER_PX = 9525
                EMU_PER_PT = 12700
                cell_w_emu = 100 * EMU_PER_PX
                cell_h_emu = 50 * EMU_PER_PT
                img_emu_w = img_w * EMU_PER_PX
                img_emu_h = img_h * EMU_PER_PX
                off_x = max(0, (cell_w_emu - img_emu_w) // 2)
                off_y = max(0, (cell_h_emu - img_emu_h) // 2)

                marker = AnchorMarker(col=1, colOff=off_x, row=row - 1, rowOff=off_y)
                xl_img.anchor = OneCellAnchor(
                    _from=marker,
                    ext=XDRPositiveSize2D(cx=img_emu_w, cy=img_emu_h),
                )
                ws.add_image(xl_img)
            except Exception:
                pass

        # Build data columns based on client type
        if is_transparency:
            # 11 columns: cols 3-11
            values_with_cols = [
                (3, part_code), (4, part_name), (5, pi_part_type),
                (6, pi_dimension), (7, pi_material), (8, item.quantity),
                (9, cfp), (10, unit_price_inr), (11, line_total),
            ]
            _wrap_cols = {3, 4, 5, 6, 7}
            _currency_cols = {9, 10, 11}
        else:
            # 10 columns: cols 3-10
            values_with_cols = [
                (3, part_code), (4, part_name), (5, pi_part_type),
                (6, pi_dimension), (7, pi_material), (8, item.quantity),
                (9, unit_price_inr), (10, line_total),
            ]
            _wrap_cols = {3, 4, 5, 6, 7}
            _currency_cols = {9, 10}

        for col_idx, val in values_with_cols:
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = styles['normal_font']
            cell.border = styles['thin_border']
            cell.alignment = Alignment(
                horizontal='left' if col_idx in _wrap_cols else 'center',
                vertical='center',
                wrap_text=(col_idx in _wrap_cols),
            )
            if col_idx in _currency_cols:
                cell.number_format = styles['currency_format']
            if row_fill:
                cell.fill = row_fill

    row += 1
    if is_transparency:
        _write_pi_total_row(ws, row, grand_total, len(items), styles, total_col=11, merge_range=f'A{row}:I{row}')
        advance_percent = pi_record.advance_percent if pi_record else 30.0
        _write_pi_advance(ws, row + 2, grand_total, advance_percent, styles, label_col='I', val_col=11)
    else:
        _write_pi_total_row(ws, row, grand_total, len(items), styles, total_col=10, merge_range=f'A{row}:H{row}')
        advance_percent = pi_record.advance_percent if pi_record else 30.0
        _write_pi_advance(ws, row + 2, grand_total, advance_percent, styles, label_col='H', val_col=10)

    pi_dir = UPLOAD_DIR / "orders" / order_id
    pi_dir.mkdir(parents=True, exist_ok=True)
    pi_filename = f"PI_{pi_number.replace('/', '-')}_with_images_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    pi_path = pi_dir / pi_filename

    if snapshot_dirty:
        db.commit()

    wb.save(str(pi_path))

    return FileResponse(
        path=str(pi_path),
        filename=pi_filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ========================================
# PI Helper Functions
# ========================================

def _pi_styles():
    """Shared styles for PI generation."""
    return {
        'header_font': Font(name='Calibri', bold=True, size=14),
        'col_header_font': Font(name='Calibri', bold=True, size=10, color='FFFFFF'),
        'normal_font': Font(name='Calibri', size=10),
        'bold_font': Font(name='Calibri', bold=True, size=10),
        'currency_format': '#,##0.00',
        'thin_border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        ),
        'header_fill': PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid'),
        'alt_fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
    }


def _set_pi_column_widths(ws, with_images: bool):
    if with_images:
        #           S.No  Image  Code   Desc   Type   Size   Material  Qty  Price  Total
        widths = {'A': 6, 'B': 14, 'C': 26, 'D': 38, 'E': 14, 'F': 20, 'G': 16, 'H': 8, 'I': 15, 'J': 16}
    else:
        #           S.No  Code   Desc   Type   Size   Material  Qty  Price  Total
        widths = {'A': 6, 'B': 26, 'C': 38, 'D': 14, 'E': 20, 'F': 16, 'G': 8, 'H': 15, 'I': 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_pi_header(ws, styles, merge_range):
    ws.merge_cells(merge_range)
    cell = ws[merge_range.split(':')[0]]
    cell.value = 'PROFORMA INVOICE'
    cell.font = styles['header_font']
    cell.alignment = Alignment(horizontal='center')


def _write_pi_info(ws, styles, pi_number, client, date_col='G', date_val_col='H', client_merge=None):
    ws['A3'] = 'PI No:'
    ws['A3'].font = styles['bold_font']
    ws['B3'] = pi_number
    ws['B3'].font = styles['normal_font']
    ws[f'{date_col}3'] = 'Date:'
    ws[f'{date_col}3'].font = styles['bold_font']
    ws[f'{date_val_col}3'] = datetime.utcnow().strftime('%d-%b-%Y')
    ws[f'{date_val_col}3'].font = styles['normal_font']
    ws['A4'] = 'Client:'
    ws['A4'].font = styles['bold_font']
    ws.merge_cells(client_merge or 'B4:C4')
    ws['B4'] = client.company_name
    ws['B4'].font = styles['normal_font']
    if client.gstin:
        ws['A5'] = 'GSTIN:'
        ws['A5'].font = styles['bold_font']
        ws['B5'] = client.gstin
        ws['B5'].font = styles['normal_font']


def _write_pi_col_headers(ws, row, headers, styles):
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header_text)
        cell.font = styles['col_header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _write_cell(ws, row, col, value, styles, row_fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = styles['normal_font']
    cell.border = styles['thin_border']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if row_fill:
        cell.fill = row_fill


def _write_pi_total_row(ws, row, grand_total, item_count, styles, total_col, merge_range):
    ws.merge_cells(merge_range)
    cell = ws[merge_range.split(':')[0]]
    cell.value = f'Total ({item_count} items)'
    cell.font = styles['bold_font']
    cell.border = styles['thin_border']
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row=row, column=total_col - 1).border = styles['thin_border']

    cell = ws.cell(row=row, column=total_col, value=grand_total)
    cell.font = styles['bold_font']
    cell.border = styles['thin_border']
    cell.number_format = styles['currency_format']
    cell.alignment = Alignment(horizontal='center', vertical='center')


def _write_pi_advance(ws, row, grand_total, advance_percent, styles, label_col, val_col):
    advance_amount = grand_total * (advance_percent / 100)
    balance_amount = grand_total - advance_amount

    ws[f'{label_col}{row}'] = 'Advance:'
    ws[f'{label_col}{row}'].font = styles['bold_font']
    next_col = chr(ord(label_col) + 1)
    ws[f'{next_col}{row}'] = f'{advance_percent}%'
    ws[f'{next_col}{row}'].font = styles['normal_font']
    cell = ws.cell(row=row, column=val_col, value=advance_amount)
    cell.font = styles['bold_font']
    cell.number_format = styles['currency_format']

    row += 1
    ws[f'{label_col}{row}'] = 'Balance:'
    ws[f'{label_col}{row}'].font = styles['bold_font']
    cell = ws.cell(row=row, column=val_col, value=balance_amount)
    cell.font = styles['normal_font']
    cell.number_format = styles['currency_format']


def _lazy_snapshot_backfill(item, product) -> bool:
    """Backfill missing snapshots from live product data. Returns True if modified."""
    if product and not item.product_code_snapshot:
        item.product_code_snapshot = product.product_code
        item.product_name_snapshot = product.product_name
        item.material_snapshot = product.material
        item.category_snapshot = product.category
        item.part_type_snapshot = product.part_type
        item.dimension_snapshot = product.dimension
        item.variant_note_snapshot = product.variant_note
        return True
    return False


def _snapshot_fields(item, product):
    """Extract snapshot fields with live fallback."""
    part_code = item.product_code_snapshot or (product.product_code if product else None) or ''
    part_name = item.product_name_snapshot or (product.product_name if product else None) or ''
    pi_part_type = item.part_type_snapshot or (product.part_type if product else None) or ''
    pi_dimension = item.dimension_snapshot or (product.dimension if product else None) or ''
    pi_material = item.material_snapshot or (product.material if product else None) or ''
    return part_code, part_name, pi_part_type, pi_dimension, pi_material


def _resolve_item_image(item, product, order_id, db) -> Optional[str]:
    """Resolve the image path for an order item, with lazy snapshot."""
    img_path_to_use = None
    if item.image_path_snapshot:
        img_path_to_use = os.path.join(str(UPLOAD_DIR), item.image_path_snapshot)
        if not os.path.exists(img_path_to_use):
            img_path_to_use = None

    if not img_path_to_use and product:
        prod_img = db.query(ProductImage).filter(
            ProductImage.product_id == product.id
        ).order_by(ProductImage.is_primary.desc(), ProductImage.created_at).first()
        if prod_img:
            img_path_to_use = os.path.join(str(UPLOAD_DIR), prod_img.image_path)
            if img_path_to_use and os.path.exists(img_path_to_use):
                dest_dir = UPLOAD_DIR / "orders" / order_id
                dest_dir.mkdir(parents=True, exist_ok=True)
                src_name = os.path.basename(prod_img.image_path)
                dest = dest_dir / src_name
                if not dest.exists():
                    shutil.copy2(img_path_to_use, str(dest))
                item.image_path_snapshot = f"orders/{order_id}/{src_name}"

    return img_path_to_use


def _save_pi_record(db, order_id, order, pi_number, grand_total, advance_percent, pi_path, items):
    """Save or update PI database record."""
    advance_amount = grand_total * (advance_percent / 100)
    existing_pi = db.query(ProformaInvoice).filter(ProformaInvoice.order_id == order_id).first()

    if existing_pi:
        revision_count = db.query(sql_func.count(PIRevision.id)).filter(
            PIRevision.order_id == order_id
        ).scalar()
        db.add(PIRevision(
            order_id=order_id,
            pi_number=existing_pi.pi_number,
            revision_number=revision_count + 1,
            total_cny=existing_pi.total_cny,
            total_inr=existing_pi.total_inr,
            exchange_rate=existing_pi.exchange_rate,
            advance_percent=existing_pi.advance_percent,
            advance_amount_inr=existing_pi.advance_amount_inr,
            item_count=len(items),
            trigger="manual",
            file_path=existing_pi.file_path,
            generated_at=existing_pi.updated_at or existing_pi.generated_at,
        ))

        existing_pi.exchange_rate = order.exchange_rate or 1.0
        existing_pi.total_inr = grand_total
        existing_pi.advance_percent = advance_percent
        existing_pi.advance_amount_inr = advance_amount
        existing_pi.file_path = str(pi_path)
    else:
        db.add(ProformaInvoice(
            pi_number=pi_number,
            order_id=order_id,
            exchange_rate=order.exchange_rate or 1.0,
            total_cny=0,
            total_inr=grand_total,
            advance_percent=advance_percent,
            advance_amount_inr=advance_amount,
            file_path=str(pi_path),
        ))
