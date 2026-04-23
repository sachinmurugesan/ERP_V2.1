"""After-Sales Management API — objection tracking, resolution, carry-forward"""
from typing import Optional, List
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
import mimetypes
from sqlalchemy import or_
from sqlalchemy.orm import Session
from pathlib import Path

import shutil
import uuid

from database import get_db
from models import (
    AfterSalesItem, Order, OrderItem, Product, Client, Factory,
    PackingList, PackingListItem,
)
from enums import AfterSalesStatus, ResolutionType, CarryForwardStatus, CarryForwardType
from config import UPLOAD_DIR
from core.security import get_current_user, CurrentUser


def _check_order_access(order: Order, current_user: CurrentUser):
    """Enforce RLS: CLIENT can only access own orders, FACTORY own orders."""
    if current_user.user_type == "CLIENT" and order.client_id != current_user.client_id:
        raise HTTPException(status_code=403, detail="Access denied")
    if current_user.user_type == "FACTORY" and order.factory_id != getattr(current_user, 'factory_id', None):
        raise HTTPException(status_code=403, detail="Access denied")


def _require_admin(current_user: CurrentUser):
    """Only ADMIN/SUPER_ADMIN/OPERATIONS/FINANCE can perform this action."""
    if current_user.role not in ("ADMIN", "SUPER_ADMIN", "OPERATIONS", "FINANCE"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

router = APIRouter()


# ========================================
# Pydantic Schemas
# ========================================

class AfterSalesSaveItem(BaseModel):
    id: str
    received_qty: int = Field(..., ge=0, le=1_000_000)
    objection_type: Optional[str] = None
    description: Optional[str] = None
    resolution_type: Optional[str] = None
    affected_quantity: Optional[int] = Field(None, ge=0, le=1_000_000)
    compensation_amount: Optional[float] = Field(None, ge=0, le=50_000_000)
    resolution_notes: Optional[str] = None


class ResolveRequest(BaseModel):
    resolution_type: str
    resolution_notes: Optional[str] = None
    affected_quantity: Optional[int] = None
    compensation_amount: Optional[float] = None


# ========================================
# Helpers
# ========================================

def _classify_item_type(item: AfterSalesItem, db: Session) -> str:
    """Classify after-sales item type from its OrderItem (same logic as packing list)."""
    if (item.selling_price_inr or 0) < 0:
        return "compensation"
    if item.order_item_id:
        oi = db.query(OrderItem).filter(OrderItem.id == item.order_item_id).first()
        if oi and oi.notes:
            if oi.notes.startswith("After-Sales") and (oi.selling_price_inr is None or oi.selling_price_inr == 0):
                return "aftersales_replacement"
            if oi.notes.startswith("Carried from"):
                return "carried_forward"
    return "regular"


def _serialize_item(item: AfterSalesItem, db: Session = None) -> dict:
    """Convert AfterSalesItem to dict for JSON response."""
    item_type = _classify_item_type(item, db) if db else "regular"
    return {
        "id": item.id,
        "order_id": item.order_id,
        "order_number": item.order_number,
        "order_item_id": item.order_item_id,
        "product_id": item.product_id,
        "product_code": item.product_code,
        "product_name": item.product_name,
        "client_id": item.client_id,
        "client_name": item.client_name,
        "ordered_quantity": item.ordered_quantity,
        "delivered_quantity": item.delivered_quantity,
        "affected_quantity": item.affected_quantity,
        "selling_price_inr": item.selling_price_inr,
        "total_value_inr": item.total_value_inr,
        "objection_type": item.objection_type,
        "description": item.description,
        "client_remarks": item.client_remarks,
        "photos": item.photos,
        "status": item.status,
        "resolution_type": item.resolution_type,
        "resolution_notes": item.resolution_notes,
        "resolution_amount": item.resolution_amount,
        "resolved_at": item.resolved_at.isoformat() if item.resolved_at else None,
        "factory_id": item.factory_id,
        "sent_qty": item.sent_qty,
        "received_qty": item.received_qty,
        "carry_forward_status": item.carry_forward_status,
        "added_to_order_id": item.added_to_order_id,
        "added_to_order_number": None,
        "added_to_order_status": None,
        "carry_forward_type": item.carry_forward_type,
        "compensation_amount": item.compensation_amount,
        "source_aftersales_id": item.source_aftersales_id,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        "is_balance_only": bool((item.selling_price_inr or 0) < 0),
        "item_type": item_type,
    }


def _apply_resolution(item: AfterSalesItem, resolution_type: str,
                       affected_quantity: Optional[int] = None,
                       compensation_amount: Optional[float] = None,
                       resolution_notes: Optional[str] = None):
    """Apply resolution logic and set carry-forward fields."""
    item.resolution_type = resolution_type
    if resolution_notes is not None:
        item.resolution_notes = resolution_notes
    if affected_quantity is not None:
        item.affected_quantity = affected_quantity

    item.carry_forward_status = CarryForwardStatus.PENDING.value

    # Determine carry-forward type
    if resolution_type in (ResolutionType.REPLACE_NEXT_ORDER.value,
                           ResolutionType.PARTIAL_REPLACEMENT.value):
        item.carry_forward_type = CarryForwardType.REPLACEMENT.value
    elif resolution_type in (ResolutionType.COMPENSATE_BALANCE.value,
                             ResolutionType.PARTIAL_COMPENSATE.value):
        item.carry_forward_type = CarryForwardType.COMPENSATION.value

    # Calculate amounts
    if resolution_type in (ResolutionType.COMPENSATE_BALANCE.value,
                           ResolutionType.PARTIAL_COMPENSATE.value):
        # Compensation: amount = affected_qty * selling_price (if not provided)
        if compensation_amount is None and item.affected_quantity and item.selling_price_inr:
            item.compensation_amount = item.affected_quantity * item.selling_price_inr
        elif compensation_amount is not None:
            item.compensation_amount = compensation_amount
        item.resolution_amount = item.compensation_amount or 0
    else:
        # Replacement: total value
        if item.affected_quantity and item.selling_price_inr:
            item.resolution_amount = item.affected_quantity * item.selling_price_inr
        if compensation_amount is not None:
            item.compensation_amount = compensation_amount

    # Set total_value_inr
    if item.affected_quantity and item.selling_price_inr:
        item.total_value_inr = item.affected_quantity * item.selling_price_inr

    item.status = AfterSalesStatus.RESOLVED.value
    item.resolved_at = datetime.utcnow()


# ========================================
# Endpoints
# ========================================

@router.get("/orders/{order_id}/")
def list_order_aftersales(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """List after-sales items for an order. Auto-populate from packing list if none exist."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    _check_order_access(order, current_user)

    items = db.query(AfterSalesItem).filter(AfterSalesItem.order_id == order_id).all()

    if not items:
        # Auto-populate from packing list
        packing_list = db.query(PackingList).filter(PackingList.order_id == order_id).first()
        if packing_list:
            pl_items = (
                db.query(PackingListItem, OrderItem, Product)
                .join(OrderItem, OrderItem.id == PackingListItem.order_item_id)
                .outerjoin(Product, Product.id == PackingListItem.product_id)
                .filter(PackingListItem.packing_list_id == packing_list.id)
                .filter(PackingListItem.loaded_qty > 0)
                .all()
            )

            # Get client info
            client = db.query(Client).filter(Client.id == order.client_id).first()
            client_name = client.company_name if client else ""

            new_items = []
            for pl_item, oi, product in pl_items:
                product_code = oi.product_code_snapshot or (product.product_code if product else "")
                product_name = oi.product_name_snapshot or (product.product_name if product else "")

                asi = AfterSalesItem(
                    order_id=order_id,
                    order_number=order.order_number or "",
                    order_item_id=pl_item.order_item_id,
                    product_id=pl_item.product_id,
                    product_code=product_code,
                    product_name=product_name,
                    client_id=order.client_id,
                    client_name=client_name,
                    factory_id=order.factory_id,
                    sent_qty=pl_item.loaded_qty,
                    received_qty=pl_item.loaded_qty,
                    ordered_quantity=oi.quantity,
                    delivered_quantity=pl_item.loaded_qty,
                    selling_price_inr=oi.selling_price_inr or 0,
                    affected_quantity=0,
                    total_value_inr=0,
                    objection_type="",
                    description="",
                    status=AfterSalesStatus.OPEN.value,
                )
                db.add(asi)
                new_items.append(asi)

            db.commit()
            for asi in new_items:
                db.refresh(asi)
            items = new_items

    return {"items": [_serialize_item(i, db) for i in items], "total": len(items)}


@router.post("/orders/{order_id}/")
def batch_save_aftersales(
    order_id: str,
    payload: List[AfterSalesSaveItem],
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Batch save/update after-sales items for an order."""
    _require_admin(current_user)
    if len(payload) > 500:
        raise HTTPException(status_code=400, detail="Too many items (max 500)")
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    updated = []
    for data in payload:
        item = db.query(AfterSalesItem).filter(
            AfterSalesItem.id == data.id,
            AfterSalesItem.order_id == order_id,
        ).first()
        if not item:
            continue

        item.received_qty = data.received_qty

        if data.description is not None:
            item.description = data.description

        if data.objection_type:
            item.objection_type = data.objection_type
            item.status = AfterSalesStatus.IN_PROGRESS.value

            if data.affected_quantity is not None:
                item.affected_quantity = data.affected_quantity

            # Save resolution notes (approver/factory comments)
            if data.resolution_notes is not None:
                item.resolution_notes = data.resolution_notes

            # If resolution is provided, apply it
            if data.resolution_type and data.resolution_type in (
                ResolutionType.REPLACE_NEXT_ORDER.value,
                ResolutionType.COMPENSATE_BALANCE.value,
                ResolutionType.PARTIAL_COMPENSATE.value,
                ResolutionType.PARTIAL_REPLACEMENT.value,
            ):
                _apply_resolution(
                    item,
                    resolution_type=data.resolution_type,
                    affected_quantity=data.affected_quantity,
                    compensation_amount=data.compensation_amount,
                )
            else:
                # Resolution cleared (user picked "--") — reset resolution fields
                item.resolution_type = None
                item.resolution_amount = 0
                item.resolution_notes = None
                item.resolved_at = None
                item.carry_forward_status = None
                item.carry_forward_type = None
                item.compensation_amount = None
                item.status = AfterSalesStatus.IN_PROGRESS.value
        else:
            # No objection — mark as CLOSED, clear all issue fields
            # objection_type and description are NOT NULL in DB, use empty string
            item.objection_type = ""
            item.description = ""
            item.resolution_type = None
            item.resolution_amount = 0
            item.resolution_notes = None
            item.resolved_at = None
            item.affected_quantity = 0
            item.carry_forward_status = None
            item.carry_forward_type = None
            item.compensation_amount = None
            item.status = AfterSalesStatus.CLOSED.value

            # Check if this is a carry-forward item — update the original claim
            if item.source_aftersales_id:
                original_claim = db.query(AfterSalesItem).filter(
                    AfterSalesItem.id == item.source_aftersales_id
                ).first()
                if original_claim and original_claim.carry_forward_status == CarryForwardStatus.ADDED_TO_ORDER.value:
                    original_claim.carry_forward_status = CarryForwardStatus.FULFILLED.value

        updated.append(item)

    db.commit()
    for item in updated:
        db.refresh(item)

    return {"items": [_serialize_item(i, db) for i in updated], "total": len(updated)}


@router.get("/orders/{order_id}/download-excel/")
def download_aftersales_excel(order_id: str, db: Session = Depends(get_db), current_user: CurrentUser = Depends(get_current_user)):
    """Download consolidated after-sales Excel report with 2 sheets:
    Sheet 1: Stock Report (all items grouped by pallet)
    Sheet 2: After-Sales Issues (only flagged items, color-coded by issue type)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from fastapi.responses import StreamingResponse
    import io

    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Get all after-sales items
    items = db.query(AfterSalesItem).filter(AfterSalesItem.order_id == order_id).all()
    if not items:
        raise HTTPException(status_code=404, detail="No after-sales items found")

    # Get package_number from packing list via order_item_id
    packing_list = db.query(PackingList).filter(PackingList.order_id == order_id).first()
    package_map = {}  # order_item_id -> package_number
    if packing_list:
        pl_items = db.query(PackingListItem).filter(
            PackingListItem.packing_list_id == packing_list.id
        ).all()
        for pli in pl_items:
            package_map[pli.order_item_id] = pli.package_number or "Loose"

    # Get client & factory names
    client = db.query(Client).filter(Client.id == order.client_id).first()
    factory = db.query(Factory).filter(Factory.id == order.factory_id).first() if order.factory_id else None
    client_name = client.company_name if client else ""
    factory_name = factory.company_name if factory else ""

    # ── Styles ──
    title_font = Font(name='Calibri', bold=True, size=14)
    info_font = Font(name='Calibri', bold=True, size=10)
    info_val_font = Font(name='Calibri', size=10)
    header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    normal_font = Font(name='Calibri', size=10)
    section_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    summary_val_font = Font(name='Calibri', bold=True, size=12)

    header_fill = PatternFill(start_color='6B21A8', end_color='6B21A8', fill_type='solid')
    alt_fill = PatternFill(start_color='F5F3FF', end_color='F5F3FF', fill_type='solid')
    flagged_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    # (pallet header fill removed — items listed flat with Pallet # column)

    # Issue-specific colors
    missing_header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
    missing_row_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    mismatch_header_fill = PatternFill(start_color='EA580C', end_color='EA580C', fill_type='solid')
    mismatch_row_fill = PatternFill(start_color='FFF7ED', end_color='FFF7ED', fill_type='solid')
    price_header_fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
    price_row_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
    quality_header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    quality_row_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    total_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════
    # SHEET 1: Stock Report (all items by pallet)
    # ═══════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Stock Report"

    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 38
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 14

    # Title
    ws1.merge_cells('A1:G1')
    c = ws1['A1']
    c.value = 'AFTER-SALES STOCK REPORT'
    c.font = title_font
    c.alignment = Alignment(horizontal='center')

    # Info
    ws1['A3'] = 'Order:'
    ws1['A3'].font = info_font
    ws1['B3'] = order.order_number or order_id
    ws1['B3'].font = info_val_font
    ws1['D3'] = 'Client:'
    ws1['D3'].font = info_font
    ws1['E3'] = client_name
    ws1['E3'].font = info_val_font
    ws1['A4'] = 'Factory:'
    ws1['A4'].font = info_font
    ws1['B4'] = factory_name
    ws1['B4'].font = info_val_font
    ws1['D4'] = 'Date:'
    ws1['D4'].font = info_font
    ws1['E4'] = datetime.utcnow().strftime('%Y-%m-%d')
    ws1['E4'].font = info_val_font

    # Group items by package_number
    grouped = {}
    for item in items:
        pkg = package_map.get(item.order_item_id, "Loose")
        grouped.setdefault(pkg, []).append(item)

    row = 6
    s1_headers = ['#', 'Part Code', 'Product Name', 'Sent Qty', 'Received Qty', 'Pallet #', 'Status']
    center_cols_s1 = {1, 4, 5, 6, 7}

    for col_idx, h in enumerate(s1_headers, 1):
        cell = ws1.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    idx = 0
    for pkg_name in sorted(grouped.keys()):
        pkg_items = grouped[pkg_name]

        for item in pkg_items:
            idx += 1
            row += 1

            if item.objection_type:
                status_text = 'Resolved' if (item.status == 'RESOLVED' or item.carry_forward_status) else 'Flagged'
            else:
                status_text = 'OK'

            vals = [idx, item.product_code, item.product_name,
                    item.sent_qty, item.received_qty, pkg_name, status_text]
            is_flagged = bool(item.objection_type)

            for col_idx, val in enumerate(vals, 1):
                cell = ws1.cell(row=row, column=col_idx, value=val)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = center_align if col_idx in center_cols_s1 else left_align
                if is_flagged:
                    cell.fill = flagged_fill
                elif idx % 2 == 0:
                    cell.fill = alt_fill

    # ═══════════════════════════════════════════
    # SHEET 2: After-Sales Issues (flagged only)
    # ═══════════════════════════════════════════
    flagged_items = [i for i in items if i.objection_type]

    if flagged_items:
        ws2 = wb.create_sheet("After-Sales Issues")
        ws2.sheet_properties.tabColor = "E11D48"

        ws2.column_dimensions['A'].width = 6
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 32
        ws2.column_dimensions['D'].width = 18
        ws2.column_dimensions['E'].width = 28
        ws2.column_dimensions['F'].width = 12
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 12
        ws2.column_dimensions['I'].width = 14
        ws2.column_dimensions['J'].width = 16
        ws2.column_dimensions['K'].width = 22
        ws2.column_dimensions['L'].width = 28
        ws2.column_dimensions['M'].width = 14

        ws2.merge_cells('A1:M1')
        c = ws2['A1']
        c.value = 'AFTER-SALES ISSUES REPORT'
        c.font = title_font
        c.alignment = Alignment(horizontal='center')

        ws2['A3'] = 'Order:'
        ws2['A3'].font = info_font
        ws2['B3'] = order.order_number or order_id
        ws2['B3'].font = info_val_font
        ws2['D3'] = 'Client:'
        ws2['D3'].font = info_font
        ws2['E3'] = client_name
        ws2['E3'].font = info_val_font

        issue_order = ['PRODUCT_MISSING', 'PRODUCT_MISMATCH', 'PRICE_MISMATCH', 'QUALITY_ISSUE']
        issue_labels = {
            'PRODUCT_MISSING': 'MISSING ITEMS',
            'PRODUCT_MISMATCH': 'MISMATCHED ITEMS',
            'PRICE_MISMATCH': 'PRICE MISMATCH ITEMS',
            'QUALITY_ISSUE': 'QUALITY ISSUES',
        }
        issue_header_fills = {
            'PRODUCT_MISSING': missing_header_fill,
            'PRODUCT_MISMATCH': mismatch_header_fill,
            'PRICE_MISMATCH': price_header_fill,
            'QUALITY_ISSUE': quality_header_fill,
        }
        issue_row_fills = {
            'PRODUCT_MISSING': missing_row_fill,
            'PRODUCT_MISMATCH': mismatch_row_fill,
            'PRICE_MISMATCH': price_row_fill,
            'QUALITY_ISSUE': quality_row_fill,
        }
        resolution_labels = {
            'REPLACE_NEXT_ORDER': 'Replace (Next Order)',
            'COMPENSATE_BALANCE': 'Compensate Balance',
            'PARTIAL_COMPENSATE': 'Partial Compensate',
            'PARTIAL_REPLACEMENT': 'Partial Replacement',
        }

        categorized = {}
        for it in flagged_items:
            categorized.setdefault(it.objection_type, []).append(it)

        # Summary section
        row = 5
        ws2.merge_cells(f'A{row}:C{row}')
        ws2[f'A{row}'] = 'CONSOLIDATED SUMMARY'
        ws2[f'A{row}'].font = Font(name='Calibri', bold=True, size=11)

        row = 6
        summary_data = []
        for otype in issue_order:
            grp = categorized.get(otype, [])
            if grp:
                val = sum((i.affected_quantity or 0) * (i.selling_price_inr or 0) for i in grp)
                summary_data.append((issue_labels[otype], len(grp), val))

        total_claim = 0
        for label, count, val in summary_data:
            ws2.cell(row=row, column=1, value=label).font = info_font
            ws2.cell(row=row, column=2, value=f'{count} items').font = info_val_font
            val_cell = ws2.cell(row=row, column=3)
            val_cell.value = val
            val_cell.font = summary_val_font
            val_cell.number_format = '#,##0.00'
            total_claim += val
            row += 1

        ws2.cell(row=row, column=1, value='TOTAL CLAIM VALUE').font = Font(name='Calibri', bold=True, size=11)
        total_cell = ws2.cell(row=row, column=3)
        total_cell.value = total_claim
        total_cell.font = Font(name='Calibri', bold=True, size=12, color='DC2626')
        total_cell.number_format = '#,##0.00'

        # Issue detail sections
        row += 2
        s2_headers = ['#', 'Part Code', 'Product Name', 'Issue Type', 'Explain',
                       'Sent Qty', 'Received Qty', 'Claim Qty',
                       'Unit Price (₹)', 'Claim Value (₹)', 'Resolution', 'Notes', 'Status']
        center_cols_s2 = {1, 6, 7, 8, 13}
        right_cols_s2 = {9, 10}
        grand_total = 0

        for otype in issue_order:
            grp = categorized.get(otype, [])
            if not grp:
                continue

            ws2.merge_cells(f'A{row}:M{row}')
            sec_cell = ws2.cell(row=row, column=1, value=f'{issue_labels[otype]} ({len(grp)})')
            sec_cell.font = section_font
            sec_cell.fill = issue_header_fills[otype]
            sec_cell.border = thin_border
            row += 1

            for col_idx, h in enumerate(s2_headers, 1):
                cell = ws2.cell(row=row, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = issue_header_fills[otype]
                cell.border = thin_border
                cell.alignment = center_align
            row += 1

            row_fill = issue_row_fills[otype]
            for gi, item in enumerate(grp, 1):
                claim_qty = item.affected_quantity or 0
                unit_price = item.selling_price_inr or 0
                claim_value = claim_qty * unit_price
                grand_total += claim_value

                issue_label = issue_labels.get(item.objection_type, item.objection_type)
                res_label = resolution_labels.get(item.resolution_type, item.resolution_type or '-')
                status_text = 'Resolved' if item.status == 'RESOLVED' else ('Carry Forward' if item.carry_forward_status == 'PENDING' else 'Flagged')

                vals = [gi, item.product_code, item.product_name,
                        issue_label, item.description or '',
                        item.sent_qty, item.received_qty, claim_qty,
                        unit_price, claim_value, res_label,
                        item.resolution_notes or '', status_text]

                for col_idx, val in enumerate(vals, 1):
                    cell = ws2.cell(row=row, column=col_idx, value=val)
                    cell.font = normal_font
                    cell.border = thin_border
                    if col_idx in right_cols_s2:
                        cell.alignment = right_align
                        cell.number_format = '#,##0.00'
                    elif col_idx in center_cols_s2:
                        cell.alignment = center_align
                    elif col_idx in (5, 12):
                        cell.alignment = wrap_align
                    else:
                        cell.alignment = left_align
                    if gi % 2 == 0:
                        cell.fill = row_fill
                row += 1
            row += 1

        # Grand total row
        ws2.merge_cells(f'A{row}:I{row}')
        gt_label = ws2.cell(row=row, column=1, value='GRAND TOTAL CLAIM VALUE')
        gt_label.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        gt_label.fill = total_fill
        gt_label.border = thin_border
        gt_label.alignment = Alignment(horizontal='right', vertical='center')
        for c in range(2, 10):
            ws2.cell(row=row, column=c).fill = total_fill
            ws2.cell(row=row, column=c).border = thin_border

        gt_val = ws2.cell(row=row, column=10, value=grand_total)
        gt_val.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        gt_val.fill = total_fill
        gt_val.border = thin_border
        gt_val.alignment = right_align
        gt_val.number_format = '#,##0.00'

        for c in [11, 12, 13]:
            ws2.cell(row=row, column=c).fill = total_fill
            ws2.cell(row=row, column=c).border = thin_border

    # ── Save & return ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"AfterSales_{order.order_number or order_id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.put("/orders/{order_id}/{item_id}/")
def update_single_item(
    order_id: str,
    item_id: str,
    data: AfterSalesSaveItem,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Update a single after-sales item."""
    _require_admin(current_user)
    item = db.query(AfterSalesItem).filter(
        AfterSalesItem.id == item_id,
        AfterSalesItem.order_id == order_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="After-sales item not found")

    item.received_qty = data.received_qty

    if data.description is not None:
        item.description = data.description

    if data.objection_type:
        item.objection_type = data.objection_type
        item.status = AfterSalesStatus.IN_PROGRESS.value

        if data.affected_quantity is not None:
            item.affected_quantity = data.affected_quantity

        if data.resolution_type and data.resolution_type in (
            ResolutionType.REPLACE_NEXT_ORDER.value,
            ResolutionType.COMPENSATE_BALANCE.value,
            ResolutionType.PARTIAL_COMPENSATE.value,
            ResolutionType.PARTIAL_REPLACEMENT.value,
        ):
            _apply_resolution(
                item,
                resolution_type=data.resolution_type,
                affected_quantity=data.affected_quantity,
                compensation_amount=data.compensation_amount,
            )
    else:
        item.status = AfterSalesStatus.CLOSED.value

        # Check if this is a carry-forward item — update the original claim
        if hasattr(item, 'source_aftersales_id') and item.source_aftersales_id:
            original_claim = db.query(AfterSalesItem).filter(
                AfterSalesItem.id == item.source_aftersales_id
            ).first()
            if original_claim and original_claim.carry_forward_status == CarryForwardStatus.ADDED_TO_ORDER.value:
                original_claim.carry_forward_status = CarryForwardStatus.FULFILLED.value

    db.commit()
    db.refresh(item)

    return _serialize_item(item, db)


@router.post("/orders/{order_id}/{item_id}/photos/")
def upload_photo(
    order_id: str,
    item_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Upload a photo for an after-sales item."""
    item = db.query(AfterSalesItem).filter(
        AfterSalesItem.id == item_id,
        AfterSalesItem.order_id == order_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="After-sales item not found")

    # Validate file type
    from core.file_upload import validate_file_type, ALLOWED_IMAGE_EXTENSIONS
    ext = validate_file_type(file.filename, ALLOWED_IMAGE_EXTENSIONS)

    # Save file to disk
    dest_dir = UPLOAD_DIR / "orders" / order_id / "aftersales"
    dest_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{uuid.uuid4().hex[:12]}{ext}"
    save_path = dest_dir / filename

    with open(str(save_path), "wb") as f:
        shutil.copyfileobj(file.file, f)

    # Update photos JSON field (append)
    photos = item.photos or []
    if not isinstance(photos, list):
        photos = []
    photos.append(filename)
    item.photos = photos

    db.commit()
    db.refresh(item)

    return _serialize_item(item, db)


@router.delete("/orders/{order_id}/{item_id}/photos/{filename}")
def delete_photo(
    order_id: str,
    item_id: str,
    filename: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Delete a photo from an after-sales item."""
    _require_admin(current_user)
    item = db.query(AfterSalesItem).filter(
        AfterSalesItem.id == item_id,
        AfterSalesItem.order_id == order_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="After-sales item not found")

    photos = list(item.photos or [])
    if filename not in photos:
        raise HTTPException(status_code=404, detail="Photo not found")

    # Remove from list — reassign a NEW list so SQLAlchemy detects the mutation
    photos.remove(filename)
    from sqlalchemy.orm.attributes import flag_modified
    item.photos = photos
    flag_modified(item, "photos")

    # Delete file from disk
    file_path = UPLOAD_DIR / "orders" / order_id / "aftersales" / filename
    if file_path.exists():
        file_path.unlink()

    db.commit()
    db.refresh(item)

    return _serialize_item(item, db)


@router.put("/orders/{order_id}/{item_id}/resolve/")
def resolve_item(
    order_id: str,
    item_id: str,
    data: ResolveRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Resolve an after-sales item with resolution type and carry-forward. ADMIN only."""
    _require_admin(current_user)
    item = db.query(AfterSalesItem).filter(
        AfterSalesItem.id == item_id,
        AfterSalesItem.order_id == order_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="After-sales item not found")

    _apply_resolution(
        item,
        resolution_type=data.resolution_type,
        affected_quantity=data.affected_quantity,
        compensation_amount=data.compensation_amount,
        resolution_notes=data.resolution_notes,
    )

    db.commit()
    db.refresh(item)

    return _serialize_item(item, db)


@router.get("/pending/")
def get_pending_claims(
    client_id: str = Query(...),
    factory_id: str = Query(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Get PENDING carry-forward claims for a client+factory pair."""
    # RLS: clients can only see their own pending claims
    if current_user.user_type == "CLIENT" and client_id != current_user.client_id:
        raise HTTPException(status_code=403, detail="Access denied")
    rows = (
        db.query(AfterSalesItem)
        .outerjoin(Order, Order.id == AfterSalesItem.order_id)
        .filter(
            AfterSalesItem.carry_forward_status == CarryForwardStatus.PENDING.value,
            AfterSalesItem.client_id == client_id,
            AfterSalesItem.factory_id == factory_id,
            or_(Order.deleted_at.is_(None), Order.id.is_(None)),
        )
        .order_by(AfterSalesItem.created_at.desc())
        .all()
    )

    result = []
    for item in rows:
        result.append({
            "id": item.id,
            "order_id": item.order_id,
            "order_number": item.order_number,
            "product_id": item.product_id,
            "product_code": item.product_code,
            "product_name": item.product_name,
            "affected_quantity": item.affected_quantity,
            "selling_price_inr": item.selling_price_inr,
            "total_value_inr": item.total_value_inr,
            "objection_type": item.objection_type,
            "resolution_type": item.resolution_type,
            "carry_forward_type": item.carry_forward_type,
            "compensation_amount": item.compensation_amount,
            "resolution_amount": item.resolution_amount,
        })

    return {"items": result}


@router.get("/")
def list_all_aftersales(
    client_id: Optional[str] = Query(None),
    factory_id: Optional[str] = Query(None),
    order_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    objection_type: Optional[str] = Query(None),
    resolution_type: Optional[str] = Query(None),
    carry_forward_only: Optional[bool] = Query(None),
    include_no_issue: Optional[bool] = Query(False),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Global list of after-sales items with optional filters.
    By default, excludes CLOSED items with no issue (OK items)."""
    if current_user.user_type == "CLIENT":
        client = db.query(Client).filter(Client.id == current_user.client_id).first()
        if not client:
            raise HTTPException(status_code=403, detail="Access denied")
        perms = client.portal_permissions or {}
        if not perms.get("show_after_sales", False):
            raise HTTPException(status_code=403, detail="After-sales access not enabled for your account")
    q = (
        db.query(AfterSalesItem, Order)
        .outerjoin(Order, Order.id == AfterSalesItem.order_id)
        .filter(
            or_(Order.deleted_at.is_(None), Order.id.is_(None))
        )
    )

    # Always hide carry-forward tracking records from the list view.
    # These are internal bookkeeping items on destination orders — not real claims.
    # The original claims (source) are the ones users interact with.
    q = q.filter(AfterSalesItem.source_aftersales_id.is_(None))

    # RLS: CLIENT users see only their own claims
    if current_user.user_type == "CLIENT" and current_user.client_id:
        q = q.filter(AfterSalesItem.client_id == current_user.client_id)
    # FACTORY users see only their factory's claims
    elif current_user.user_type == "FACTORY" and current_user.factory_id:
        q = q.filter(AfterSalesItem.factory_id == current_user.factory_id)

    # Default: exclude items with no issue (CLOSED with empty objection_type)
    if not include_no_issue:
        q = q.filter(
            or_(
                AfterSalesItem.objection_type.isnot(None),
                AfterSalesItem.objection_type != "",
                AfterSalesItem.status != "CLOSED",
                AfterSalesItem.affected_quantity > 0,
            )
        )
        # Further exclude truly empty items
        q = q.filter(
            ~(
                (AfterSalesItem.status == "CLOSED") &
                ((AfterSalesItem.objection_type == None) | (AfterSalesItem.objection_type == "")) &
                ((AfterSalesItem.affected_quantity == None) | (AfterSalesItem.affected_quantity == 0))
            )
        )

    if client_id:
        q = q.filter(AfterSalesItem.client_id == client_id)
    if factory_id:
        q = q.filter(AfterSalesItem.factory_id == factory_id)
    if order_id:
        q = q.filter(AfterSalesItem.order_id == order_id)
    if status:
        q = q.filter(AfterSalesItem.status == status)
    if objection_type:
        q = q.filter(AfterSalesItem.objection_type == objection_type)
    if resolution_type:
        q = q.filter(AfterSalesItem.resolution_type == resolution_type)
    if carry_forward_only:
        q = q.filter(AfterSalesItem.carry_forward_status.isnot(None))

    rows = q.order_by(AfterSalesItem.created_at.desc()).all()

    # Summary stats
    total_claims = len(rows)
    open_count = sum(1 for item, _ in rows if item.status == "OPEN")
    in_progress_count = sum(1 for item, _ in rows if item.status == "IN_PROGRESS")
    resolved_count = sum(1 for item, _ in rows if item.status == "RESOLVED")
    pending_carry = sum(1 for item, _ in rows if item.carry_forward_status == "PENDING")
    total_claim_value = sum(
        (item.selling_price_inr or 0) * (item.affected_quantity or 0)
        for item, _ in rows if item.objection_type
    )

    # Batch-load destination orders for carry-forward items
    added_order_ids = {item.added_to_order_id for item, _ in rows if item.added_to_order_id}
    added_orders_map = {}
    if added_order_ids:
        for ao in db.query(Order).filter(Order.id.in_(added_order_ids)).all():
            added_orders_map[ao.id] = ao

    result = []
    for item, order in rows:
        d = _serialize_item(item, db)
        d["order_number"] = order.order_number if order else item.order_number
        d["factory_name"] = None
        if item.factory_id:
            factory = db.query(Factory).filter(Factory.id == item.factory_id).first()
            d["factory_name"] = factory.company_name if factory else None
        # Destination order info for carry-forward tracking
        if item.added_to_order_id and item.added_to_order_id in added_orders_map:
            dest = added_orders_map[item.added_to_order_id]
            d["added_to_order_number"] = dest.order_number
            d["added_to_order_status"] = dest.status
        # Claim value
        d["claim_value"] = round((item.selling_price_inr or 0) * (item.affected_quantity or 0), 2) if item.objection_type else 0
        result.append(d)

    return {
        "items": result,
        "total": total_claims,
        "summary": {
            "open": open_count,
            "in_progress": in_progress_count,
            "resolved": resolved_count,
            "pending_carry_forward": pending_carry,
            "total_claim_value": round(total_claim_value, 2),
        },
    }


# ========================================
# CLIENT PORTAL ENDPOINTS
# ========================================

def _client_safe_serialize(item: AfterSalesItem, db: Session, pallet_map: dict = None) -> dict:
    """Serialize after-sales item for client portal — hides factory data."""
    item_type = _classify_item_type(item, db)
    pallet = None
    if pallet_map and item.order_item_id:
        pallet = pallet_map.get(item.order_item_id)
    return {
        "id": item.id,
        "order_id": item.order_id,
        "order_number": item.order_number,
        "product_code": item.product_code,
        "product_name": item.product_name,
        "ordered_quantity": item.ordered_quantity,
        "sent_qty": item.sent_qty,
        "received_qty": item.received_qty,
        "objection_type": item.objection_type,
        "description": item.description,
        "client_remarks": item.client_remarks,
        "affected_quantity": item.affected_quantity,
        "selling_price_inr": item.selling_price_inr,
        "total_value_inr": item.total_value_inr,
        "photos": item.photos,
        "status": item.status,
        "resolution_type": item.resolution_type,
        "resolution_notes": item.resolution_notes,
        "resolved_at": item.resolved_at.isoformat() if item.resolved_at else None,
        "carry_forward_status": item.carry_forward_status,
        "carry_forward_type": item.carry_forward_type,
        "is_balance_only": bool((item.selling_price_inr or 0) < 0),
        "item_type": item_type,
        "pallet_number": pallet,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        # Deliberately excluded: factory_id, compensation_amount, resolution_amount
    }


class ClientClaimItem(BaseModel):
    """Client can only edit these fields when submitting a claim."""
    id: str
    received_qty: int
    objection_type: Optional[str] = None  # PRODUCT_MISMATCH, PRODUCT_MISSING, QUALITY_ISSUE, PRICE_MISMATCH
    description: Optional[str] = None
    affected_quantity: Optional[int] = None
    client_remarks: Optional[str] = None


@router.get("/client/orders/{order_id}/")
def client_get_aftersales(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Client view of after-sales items — read-only for resolution fields."""
    if current_user.user_type == "CLIENT":
        client = db.query(Client).filter(Client.id == current_user.client_id).first()
        if not client:
            raise HTTPException(status_code=403, detail="Access denied")
        perms = client.portal_permissions or {}
        if not perms.get("show_after_sales", False):
            raise HTTPException(status_code=403, detail="After-sales access not enabled for your account")
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    # RLS: client can only view their own orders
    if current_user.user_type == "CLIENT" and order.client_id != current_user.client_id:
        raise HTTPException(status_code=403, detail="Access denied")

    items = db.query(AfterSalesItem).filter(
        AfterSalesItem.order_id == order_id,
    ).order_by(AfterSalesItem.product_code).all()

    # If no items yet, the admin hasn't opened the after-sales tab yet
    if not items:
        return {"items": [], "total": 0, "message": "After-sales review has not started yet. Please wait for our team to initiate the process."}

    # Build pallet map from packing list
    pallet_map = {}
    packing_list = db.query(PackingList).filter(PackingList.order_id == order_id).first()
    if packing_list:
        packing_items = db.query(PackingListItem).filter(
            PackingListItem.packing_list_id == packing_list.id,
        ).all()
        for pli in packing_items:
            if pli.order_item_id and pli.package_number:
                pallet_map[pli.order_item_id] = pli.package_number

    return {
        "items": [_client_safe_serialize(i, db, pallet_map) for i in items],
        "total": len(items),
    }


@router.post("/client/orders/{order_id}/claims/")
def client_submit_claims(
    order_id: str,
    claims: List[ClientClaimItem],
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Client submits claims — can only set received_qty, issue, description, claim_qty."""
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if current_user.user_type == "CLIENT" and order.client_id != current_user.client_id:
        raise HTTPException(status_code=403, detail="Access denied")

    updated = []
    for claim in claims:
        item = db.query(AfterSalesItem).filter(
            AfterSalesItem.id == claim.id,
            AfterSalesItem.order_id == order_id,
        ).first()
        if not item:
            continue

        # Client can only edit these fields
        item.received_qty = claim.received_qty

        if claim.objection_type:
            item.objection_type = claim.objection_type
            item.description = claim.description or item.description
            item.affected_quantity = claim.affected_quantity or item.affected_quantity
            item.client_remarks = claim.client_remarks or item.client_remarks
            # Auto-set status to OPEN when client files a claim
            if item.status in (None, "", "CLOSED"):
                item.status = AfterSalesStatus.OPEN.value
        else:
            # Client clearing an issue (no objection)
            if item.status == AfterSalesStatus.OPEN.value:
                item.objection_type = None
                item.description = None
                item.affected_quantity = None
                item.status = "CLOSED"

        item.updated_at = datetime.utcnow()
        updated.append(item)

    # Notify admin about client's after-sales submission
    open_claims = [i for i in updated if i.status == AfterSalesStatus.OPEN.value]
    if open_claims:
        from models import Notification
        db.add(Notification(
            user_role="ADMIN",
            title="After-Sales Claims Submitted",
            message=f"Client submitted {len(open_claims)} claim(s) for {order.order_number}",
            notification_type="AFTER_SALES_SUBMIT",
            resource_type="order",
            resource_id=order_id,
        ))

    db.commit()

    return {
        "items": [_client_safe_serialize(i, db) for i in updated],
        "saved": len(updated),
    }


@router.get("/orders/{order_id}/photos/{filename}")
def download_aftersales_photo(
    order_id: str,
    filename: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Authenticated download of an after-sales evidence photo.
    Applies the same RLS as the after-sales item endpoints via _check_order_access.
    """
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at.is_(None)).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    _check_order_access(order, current_user)

    # Prevent path traversal — filename must not contain directory components
    safe_filename = Path(filename).name
    if not safe_filename or safe_filename != filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    full_path = UPLOAD_DIR / "orders" / order_id / "aftersales" / safe_filename
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="Photo not found")

    mime_type = mimetypes.guess_type(str(full_path))[0] or "image/jpeg"
    return FileResponse(path=str(full_path), filename=safe_filename, media_type=mime_type)


@router.post("/client/orders/{order_id}/{item_id}/photos/")
def client_upload_photo(
    order_id: str,
    item_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Client uploads evidence photo for a claim."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if current_user.user_type == "CLIENT" and (not order or order.client_id != current_user.client_id):
        raise HTTPException(status_code=403, detail="Access denied")
    item = db.query(AfterSalesItem).filter(
        AfterSalesItem.id == item_id,
        AfterSalesItem.order_id == order_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    # Save file
    save_dir = Path(UPLOAD_DIR) / "orders" / order_id / "aftersales"
    save_dir.mkdir(parents=True, exist_ok=True)

    ext = file.filename.rsplit(".", 1)[-1] if "." in file.filename else "jpg"
    filename = f"{uuid.uuid4().hex[:12]}.{ext}"
    dest = save_dir / filename

    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # Add to photos list (immutable)
    photos = list(item.photos or [])
    photos.append(filename)
    item.photos = photos
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(item, "photos")
    item.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(item)

    return _client_safe_serialize(item, db)
